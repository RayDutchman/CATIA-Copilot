"""
重量、重心、惯量统计对话框模块。

提供：
- MassPropsDialog – 遍历产品树，展示每个零件实例的质量/重心/转动惯量，
                    支持：
                      • 手动编辑重量（等比缩放惯量，联动同型号零件）
                      • 完整 BOM / 汇总 BOM 切换
                      • 重量单位 g/kg 独立选择
                      • 长度单位 mm/m 独立选择
                      • 惯量单位 g·mm²/g·m²/kg·mm²/kg·m² 独立选择（4 种）
                      • 惯量包络体读取模式：只读.1 / 最大编号 / 全部汇总
                      • 文件名 / 零件编号 / 术语 / 版本列可隐藏
                      • 自动汇总产品总质量特性并导出 Excel
"""

import csv
import ctypes
import logging
import math
import uuid
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PySide6.QtCore import QByteArray, QSettings, Qt, QUrl
from PySide6.QtGui import QBrush, QDesktopServices, QFont, QKeySequence, QShortcut
from PySide6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QButtonGroup,
    QCheckBox,
    QComboBox,
    QDialog,
    QFileDialog,
    QFrame,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QRadioButton,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from catia_copilot.catia.connection import open_document
from catia_copilot.catia.mass_props_calc import rollup_mass_properties
from catia_copilot.catia.mass_props_collect import (
    _compute_root_mp_from_placement,
    _identity_4x4,
    _mat4_mul,
    _measure_part_mass_props,
    _measure_part_mass_props_analyze,
    _position_to_mat4,
    _rollup_one_product,
    _row_inertia_to_root,
    collect_mass_props_rows,
    load_rows,
    merge_rows,
    recompute_product_rows,
    save_rows,
)
from catia_copilot.constants import (
    FILENAME_UNSAVED,
    MASS_PROPS_COLUMN_DISPLAY_NAMES,
    MASS_PROPS_HIDEABLE_COLUMNS,
    MASS_PROPS_READONLY_COLUMNS,
    MAX_INERTIA_INDEX,
    TYPE_DISPLAY_NAMES,
    BomNodeType,
)
from catia_copilot.ui.bom_widgets import (
    _BomSortItem,
    _BomTreeWidget,
    _RowHeightDelegate,
)
from catia_copilot.ui.theme_manager import theme_manager, theme_signal
from catia_copilot.ui.ui_colors import (
    get_colors as _get_colors,
)
from catia_copilot.ui.ui_layout import L
from catia_copilot.utils import estimate_column_width

logger = logging.getLogger(__name__)

# UserRole：行索引（映射到 self._rows）
_ROW_IDX_ROLE = Qt.ItemDataRole.UserRole
# UserRole+1：锁定标志位（不可编辑行）
_ITEM_LOCKED_ROLE = Qt.ItemDataRole.UserRole + 1
# UserRole+2：密度锁定（密度值为 -1 或 None，不允许编辑密度列）
_DENSITY_LOCKED_ROLE = Qt.ItemDataRole.UserRole + 2
# UserRole+3：该行已被用户排除（不参与计算）
_EXCLUDED_ROLE = Qt.ItemDataRole.UserRole + 3

# 惯量列名 → (行索引, 列索引)，对应 3×3 张量位置
_INERTIA_IDX: dict[str, tuple[int, int]] = {
    "Ixx": (0, 0), "Iyy": (1, 1), "Izz": (2, 2),
    "Ixy": (0, 1), "Ixz": (0, 2), "Iyz": (1, 2),
}
_COG_IDX: dict[str, int] = {"CogX": 0, "CogY": 1, "CogZ": 2}

# 显示值随当前单位制变化的列名
_UNIT_SENSITIVE_COLUMNS: tuple[str, ...] = (
    "Weight",
) + tuple(_INERTIA_IDX.keys()) + ("CogX", "CogY", "CogZ")

# 数值格式化：判断"接近整数"的绝对容差（用于 _fmt / _fmt_scaled）
_INTEGER_ABS_TOL: float = 1e-9

# 排除行视觉样式（斜体字体；背景/前景色已移至 ui_colors.py）
_EXCL_FONT: QFont = QFont()
_EXCL_FONT.setItalic(True)

# 对称件（虚拟行）视觉样式
_MIRROR_TOOLTIP: str = "对称件（虚拟行），相对 ZX 平面与原件对称，不可直接编辑。"


class _MassPropsDelegate(_RowHeightDelegate):
    """只允许对未锁定零件行的"Weight"列进行编辑，其余列一律只读。

    继承 :class:`_RowHeightDelegate`，获得行高保证与列 0 防溢出能力。
    """

    def __init__(self, cols_fn, tree) -> None:
        super().__init__(tree)
        self._cols_fn = cols_fn  # callable: () -> list[str]

    def createEditor(self, parent, option, index):
        tree = self.parent()
        item = tree.itemFromIndex(index)
        if item is None:
            return None
        if item.data(0, _ITEM_LOCKED_ROLE):
            return None
        cols = self._cols_fn()
        if index.column() >= len(cols):
            return None
        col_name = cols[index.column()]
        if col_name in MASS_PROPS_READONLY_COLUMNS:
            return None
        # 密度列：当该行密度不统一或无数据时不允许编辑
        if col_name == "Density" and item.data(0, _DENSITY_LOCKED_ROLE):
            return None
        return super().createEditor(parent, option, index)


def _fmt(value) -> str:
    """数值 → 字符串，None → '—'（不含单位换算）。"""
    if value is None:
        return "—"
    try:
        v = float(value)
        rv = round(v)
        if math.isclose(v, rv, rel_tol=0.0, abs_tol=_INTEGER_ABS_TOL):
            return f"{rv:.0f}"
        if abs(v) >= 1e5 or abs(v) < 0.001:
            return f"{v:.3e}"
        return f"{v:.3f}"
    except (TypeError, ValueError):
        return str(value)


class MassPropsDialog(QDialog):
    """重量、重心、惯量统计对话框。

    - 遍历 CATProduct 树，每个节点（零件/产品/部件实例）单独显示一行（完整 BOM 模式）。
      Weight / CogX / CogY / CogZ / Ixx–Iyz 均在根产品坐标系下显示，与装配位置有关。
    - 汇总 BOM 模式：相同零件编号的零件实例合并为一行，并显示数量（Quantity）；
      仅列出零件（不含产品和部件）；Weight / CogX / CogY / CogZ / Ixx–Iyz
      在零件自身坐标系下显示，与装配位置无关。
    - 仅零件节点的"重量"列可编辑；修改后等比缩放该行惯量，
      并同步更新所有相同零件编号的行（及 _rows 中全部同PN数据）。
    - 单位可在 kg/g 间切换（影响重量列与转动惯量列的显示和导出）。
    - 修改密度或重量后自动重新计算产品总质量特性，汇总结果实时更新。
    - "导出表格"将当前数据（含汇总行）写入 Excel 。
    """

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("质量特性工作台")
        self.setMinimumSize(900, 600)
        self.resize(1100, 700)

        self._settings = QSettings("CATIACompanion", "MassPropsDialog")
        self._last_browse_dir: str = self._settings.value("last_browse_dir", "")

        # ── 持久化显示选项 ─────────────────────────────────────────────────
        saved_hid = self._settings.value("visible_hideable_cols", list(MASS_PROPS_HIDEABLE_COLUMNS))
        if saved_hid is None:
            saved_hid = list(MASS_PROPS_HIDEABLE_COLUMNS)
        elif isinstance(saved_hid, str):
            saved_hid = [saved_hid]
        else:
            saved_hid = list(saved_hid)
        self._visible_hideable_cols: set[str] = {
            c for c in saved_hid if c in MASS_PROPS_HIDEABLE_COLUMNS
        }

        self._summarize: bool = self._settings.value("summarize", False, type=bool)

        # ── 单位制 ────────────────────────────────────────────────────────────
        self._mass_unit: str = self._settings.value("mass_unit", "g")
        self._cog_unit: str = self._settings.value("cog_unit", "mm")
        if self._mass_unit not in ("g", "kg"):
            self._mass_unit = "g"
        if self._cog_unit not in ("mm", "m"):
            self._cog_unit = "mm"

        # 惯量单位独立选择（4 种）
        _valid_inertia_units = ("g\u00b7mm\u00b2", "g\u00b7m\u00b2", "kg\u00b7mm\u00b2", "kg\u00b7m\u00b2")
        self._inertia_unit: str = self._settings.value("inertia_unit", "g\u00b7mm\u00b2")
        if self._inertia_unit not in _valid_inertia_units:
            self._inertia_unit = "g\u00b7mm\u00b2"

        # 内部单位为 SI（kg / m / kg·m²）；根据所选显示单位制设置换算因子
        self._unit_factor, _, self._cog_unit_factor = (
            self._calc_unit_factors(self._mass_unit, self._cog_unit)
        )
        self._inertia_unit_factor = self._calc_inertia_factor(self._inertia_unit)

        # ── 读取模式 ─────────────────────────────────────────────────────────
        self._read_mode: str = self._settings.value("read_mode", "all")
        if self._read_mode not in ("first", "last", "all"):
            self._read_mode = "all"

        # ── 数据来源 ──────────────────────────────────────────────────────────
        self._source: str = self._settings.value("source", "analyze")
        if self._source not in ("keep_inertia", "analyze"):
            self._source = "analyze"

        # ── 忽略隐藏节点 ──────────────────────────────────────────────────────
        self._skip_hidden: bool = self._settings.value("skip_hidden", False, type=bool)

        # ── 内部状态 ──────────────────────────────────────────────────────
        self._rows: list[dict] = []
        # display_row_idx → QTreeWidgetItem
        self._item_by_row: list[QTreeWidgetItem] = []
        # Part Number → list[QTreeWidgetItem] (all visible items with that PN)
        self._pn_to_items: dict[str, list[QTreeWidgetItem]] = {}
        self._is_updating: bool = False
        self._rollup_result: dict | None = None
        self._loaded: bool = False
        self._col_widths: dict[str, int] = {}
        # 未保存到磁盘的编辑标志（标题栏 * 指示符）
        self._is_dirty: bool = False
        # 筛选/搜索框当前文本（小写后的模式串）
        self._filter_text: str = ""

        # 按钮引用（_build_ui 中赋值，此处声明以供类型提示）
        self._append_data_btn: QPushButton
        self._append_active_btn: QPushButton

        # 列名列表在可见性或模式改变时重建
        self._columns: list[str] = self._build_columns()

        self._build_ui()

        # ── 恢复窗口几何（位置与尺寸）────────────────────────────────────────
        saved_geom = self._settings.value("geometry")
        if isinstance(saved_geom, QByteArray) and not saved_geom.isEmpty():
            self.restoreGeometry(saved_geom)

        # ── 主题切换：重新着色所有行及提示标签 ───────────────────────────────
        theme_signal.theme_changed.connect(self._on_theme_changed)

        # ── 默认使用活动文档（_build_ui 已完成，控件已存在）────────────────────
        self._use_active_chk.setChecked(True)

    def done(self, result: int) -> None:
        """保存窗口几何后执行标准关闭流程。"""
        self._settings.setValue("geometry", self.saveGeometry())
        super().done(result)

    # ── 主题切换响应 ──────────────────────────────────────────────────────────

    @staticmethod
    def _prereq_label_style(mode: str) -> str:
        """惯量前置条件说明标签的主题样式。"""
        if mode == "dark":
            return (
                "QLabel { background-color: #3a2e00; border: 1px solid #8a6800;"
                f" border-radius: 4px; padding: 4px 8px; color: #e0b840; font-size: {L.SMALL_FONT_SIZE}px; }}"
            )
        return (
            "QLabel { background-color: #FFF8E1; border: 1px solid #F9A825;"
            f" border-radius: 4px; padding: 4px 8px; color: #5D4037; font-size: {L.SMALL_FONT_SIZE}px; }}"
        )

    @staticmethod
    def _bom_desc_label_style(mode: str) -> str:
        """BOM 描述说明标签的主题样式。"""
        if mode == "dark":
            return (
                "QLabel { background-color: #102040; border: 1px solid #204878;"
                f" border-radius: 4px; padding: 4px 8px; color: #80b0e0; font-size: {L.SMALL_FONT_SIZE}px; }}"
            )
        return (
            "QLabel { background-color: #EEF4FC; border: 1px solid #B8D0F0;"
            f" border-radius: 4px; padding: 4px 8px; color: #2B4C7E; font-size: {L.SMALL_FONT_SIZE}px; }}"
        )

    def _on_theme_changed(self, mode: str) -> None:
        """主题切换时刷新提示标签样式并重新着色所有树形行。"""
        # 更新提示标签
        if hasattr(self, "_prereq_lbl"):
            self._prereq_lbl.setStyleSheet(self._prereq_label_style(mode))
        if hasattr(self, "_bom_desc_lbl"):
            self._bom_desc_lbl.setStyleSheet(self._bom_desc_label_style(mode))
        # 重新着色树形行
        if self._item_by_row:
            self._recolor_all_items(mode)

    def _recolor_all_items(self, mode: str) -> None:
        """按当前主题颜色重新为所有树形行设置背景/前景色。"""
        for item in self._item_by_row:
            r_idx = item.data(0, _ROW_IDX_ROLE)
            if r_idx is None:
                continue
            row_data = self._rows[r_idx]
            # 排除行优先级最高
            if item.data(0, _EXCLUDED_ROLE):
                c = _get_colors(mode)
                for ci in range(len(self._columns)):
                    item.setBackground(ci, c.EXCL_BG)
                    item.setForeground(ci, c.EXCL_FG)
            else:
                # 先清除，再按行状态重新应用
                default_brush = QBrush()
                for ci in range(len(self._columns)):
                    item.setBackground(ci, default_brush)
                    item.setForeground(ci, default_brush)
                self._apply_row_state_style(item, row_data)

    # ── 列管理 ──────────────────────────────────────────────────────────────

    @staticmethod
    def _calc_unit_factors(mass_unit: str, cog_unit: str) -> tuple[float, float, float]:
        """根据重量单位和长度单位返回 (mass_factor, inertia_factor_derived, cog_factor)。

        内部存储单位为：质量 kg、坐标 mm、惯量 kg·mm²。
        inertia_factor_derived = mass_factor × cog_factor²（用于向后兼容推导）。
        独立惯量单位换算请使用 _calc_inertia_factor()。
        """
        mf = 1e3 if mass_unit == "g" else 1.0
        cf = 1e-3 if cog_unit == "m" else 1.0   # 内部已是 mm；选 m 显示则缩小 1000 倍
        return mf, mf * cf * cf, cf

    @staticmethod
    def _calc_inertia_factor(inertia_unit: str) -> float:
        """从内部单位 kg·mm² 换算到 inertia_unit 字符串对应显示单位的换算因子。

        支持的惯量单位字符串（Unicode 上标²）：
          "kg·mm²" → 1.0    (内部单位，无需换算)
          "kg·m²"  → 1e-6   (mm²→m²=÷1e6)
          "g·mm²"  → 1e3    (kg→g=×1e3)
          "g·m²"   → 1e-3   (kg→g=×1e3, mm²→m²=÷1e6, 合计×1e-3)
        """
        _map = {
            "kg\u00b7mm\u00b2": 1.0,
            "kg\u00b7m\u00b2":  1e-6,
            "g\u00b7mm\u00b2":  1e3,
            "g\u00b7m\u00b2":   1e-3,
        }
        return _map.get(inertia_unit, 1.0)

    def _build_columns(self) -> list[str]:
        """根据当前可见性设置和 BOM 模式，构建列名列表。

        完整 BOM ：Level 在第 0 列（承载树形装饰线），# 在第 1 列。
        汇总 BOM ：无 Level 列，# 在第 0 列（无装饰线需求），增加 Quantity 列。
                   Instance Name 在汇总 BOM 中无意义（多实例合并为一行），强制隐藏。
        """
        if self._summarize:
            base = ["#", "Type"]
            for c in MASS_PROPS_HIDEABLE_COLUMNS:
                if c == "Instance Name":
                    continue  # 汇总 BOM 不显示实例名（多实例合并，无从区分）
                if c in self._visible_hideable_cols:
                    base.append(c)
            base += ["Quantity", "Density", "Weight", "CogX", "CogY", "CogZ",
                     "Ixx", "Iyy", "Izz", "Ixy", "Ixz", "Iyz"]
        else:
            base = ["Level", "#", "Type"]
            for c in MASS_PROPS_HIDEABLE_COLUMNS:
                if c in self._visible_hideable_cols:
                    base.append(c)
            base += ["Density", "Weight", "CogX", "CogY", "CogZ",
                     "Ixx", "Iyy", "Izz", "Ixy", "Ixz", "Iyz"]
        return base

    def _column_header(self, col_name: str) -> str:
        """返回列名的中文显示名（含当前单位后缀）。"""
        if col_name == "Density":
            return "密度 (kg/m³)"
        if col_name == "Weight":
            return f"重量 ({self._mass_unit})"
        if col_name in _INERTIA_IDX:
            return f"{col_name} ({self._inertia_unit})"
        if col_name in ("CogX", "CogY", "CogZ"):
            return f"{col_name} ({self._cog_unit})"
        return MASS_PROPS_COLUMN_DISPLAY_NAMES.get(col_name, col_name)

    def _display_headers(self) -> list[str]:
        return [self._column_header(c) for c in self._columns]

    @staticmethod
    def _fmt_scaled(value, factor: float) -> str:
        """将原始 SI 值乘以换算因子后格式化为字符串。

        None → '—'；数值格式化规则与模块级 _fmt() 完全一致（委托调用）；
        乘法本身若抛出 TypeError/ValueError 则回退为 str(value)。
        """
        if value is None:
            return "—"
        try:
            return _fmt(float(value) * factor)
        except (TypeError, ValueError):
            return str(value)

    def _fmt_mass_val(self, value) -> str:
        """将质量原始值（kg，SI 内部单位）乘以 _unit_factor 并格式化为字符串（重量列专用）。"""
        return self._fmt_scaled(value, self._unit_factor)

    def _fmt_inertia_val(self, value) -> str:
        """将惯量原始值（kg·mm²，内部单位）乘以 _inertia_unit_factor 并格式化为字符串（惯量列专用）。"""
        return self._fmt_scaled(value, self._inertia_unit_factor)

    def _fmt_cog_val(self, value) -> str:
        """将重心坐标原始值（mm，内部单位）乘以 _cog_unit_factor 并格式化为字符串（CogX/Y/Z 列专用）。"""
        return self._fmt_scaled(value, self._cog_unit_factor)

    # ── UI 构建 ────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        layout = QVBoxLayout(self)
        layout.setSpacing(8)
        layout.setContentsMargins(16, 16, 16, 16)

        # ── 前提条件说明（窗口过窄时允许截断）──────────────────────────────
        prereq_lbl = QLabel(
            "⚠ 使用说明：本功能用于统计产品的重量、重心、转动惯量。基本原理是读取指定产品树下的每个零件的'属性-机械'中的结果或'测量惯量'结果，"
            "和每个零件在根产品中的位置，计算出产品重量、重心、转动惯量。"
            "当使用'Analyze'模式时，读取的是零件的'属性-机械'中的数据，需要零件的主几何体被正确赋予材料，且不得有多余的几何体或包络体（因为它们也会被一同纳入统计）。"
            "当使用'惯量包络体'模式时，读取的是零件中勾选 '保持测量' 的惯量测量结果（惯量包络体），需要在 CATIA 中 <b>单独打开</b> 每个零件,执行'测量惯量'并勾选 <b>保持测量</b>。"
            "在产品窗口中建立的惯量包络体的参考坐标系为根产品坐标系（即使当前工作对象是零件），"
            "这会导致坐标系与根产品不重合的零件的测量结果不正确。"
            f"测量结果必须命名为 <b>惯量包络体.x</b>（x 为 1–{MAX_INERTIA_INDEX} 的整数），"
            "支持一个零件具有多个惯量包络体，产品中的惯量包络体将不被读取。"
        )
        prereq_lbl.setWordWrap(True)
        self._prereq_lbl = prereq_lbl
        prereq_lbl.setStyleSheet(self._prereq_label_style(theme_manager.current_mode()))
        layout.addWidget(prereq_lbl)

        # ── 数据来源选择 ────────────────────────────────────────────────────
        self._use_active_chk = QCheckBox("使用当前 CATIA 活动文档（无需手动选择文件）")
        self._use_active_chk.toggled.connect(self._toggle_file_row)
        layout.addWidget(self._use_active_chk)

        file_row = QHBoxLayout()
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("选择一个 CATProduct 文件…")
        self._file_edit.setReadOnly(True)
        self._file_browse_btn = QPushButton("浏览…")
        self._file_browse_btn.clicked.connect(self._browse_file)
        self._load_btn = QPushButton("加载")
        self._load_btn.clicked.connect(self._load_data)
        self._load_json_btn = QPushButton("载入已保存数据…")
        self._load_json_btn.setToolTip("从之前保存的数据文件中载入质量特性（无需打开 CATIA）")
        self._load_json_btn.clicked.connect(self._load_data_from_json)
        self._append_data_btn = QPushButton("追加数据…")
        self._append_data_btn.setToolTip(
            "从已保存的数据文件（.mpd）追加分总成数据并合并汇总\n"
            "（适用于主产品过大、分批读取各分总成的场景；各分总成坐标系须与主产品一致）"
        )
        self._append_data_btn.setEnabled(False)
        self._append_data_btn.clicked.connect(self._append_data_from_file)
        self._append_active_btn = QPushButton("追加活动文档…")
        self._append_active_btn.setToolTip(
            "将 CATIA 当前活动文档（分总成）的质量特性追加到现有数据中\n"
            "（各分总成坐标系须与主产品一致）"
        )
        self._append_active_btn.setEnabled(False)
        self._append_active_btn.clicked.connect(self._append_from_active)
        file_row.addWidget(self._file_edit)
        file_row.addWidget(self._file_browse_btn)
        file_row.addWidget(self._load_btn)
        file_row.addWidget(self._load_json_btn)
        file_row.addWidget(self._append_data_btn)
        file_row.addWidget(self._append_active_btn)
        layout.addLayout(file_row)

        # ── 选项面板（2 行）────────────────────────────────────────────────
        opts_group = QGroupBox("读取与显示选项")
        opts_main = QVBoxLayout(opts_group)
        opts_main.setSpacing(4)
        opts_main.setContentsMargins(8, 6, 8, 6)

        # ── 第一行：BOM类型 ｜ 读取模式 ｜ 忽略隐藏 ──────────────────────────
        row1 = QHBoxLayout()
        row1.setSpacing(6)

        # BOM 类型
        self._bom_type_group = QButtonGroup(self)
        self._radio_hier = QRadioButton("完整 BOM")
        self._radio_summ = QRadioButton("汇总 BOM")
        self._radio_summ.setToolTip(
            "汇总 BOM ：按零件编号合并同种零件，仅显示零件行。\n"
            "产品、部件、对称件不在此视图中显示。"
        )
        self._radio_summ.setMinimumHeight(24)
        self._radio_hier.setChecked(not self._summarize)
        self._radio_summ.setChecked(self._summarize)
        self._bom_type_group.addButton(self._radio_hier)
        self._bom_type_group.addButton(self._radio_summ)
        self._radio_summ.toggled.connect(self._on_bom_type_changed)
        row1.addWidget(QLabel("BOM:"))
        row1.addWidget(self._radio_hier)
        row1.addWidget(self._radio_summ)

        _sep1 = QFrame(); _sep1.setFrameShape(QFrame.Shape.VLine)
        _sep1.setFrameShadow(QFrame.Shadow.Sunken)
        row1.addSpacing(4); row1.addWidget(_sep1); row1.addSpacing(4)

        # 数据来源
        self._radio_src_keep = QRadioButton("惯量包络体")
        self._radio_src_analyze = QRadioButton("Analyze")
        self._radio_src_keep.setToolTip(
            "读取 SPA 保持测量写入的「惯量包络体.N」参数。\n"
            "需用户预先在 SPA 中执行「测量惯量 + 保持测量」操作。"
        )
        self._radio_src_analyze.setToolTip(
            "通过 CATIA Analyze API 实时计算零件质量特性。\n"
            "需零件已赋材料；无需手动创建保持测量。\n"
            "选中时「惯量包络体读取」选项不可用。"
        )
        self._radio_src_keep.setChecked(self._source == "keep_inertia")
        self._radio_src_analyze.setChecked(self._source == "analyze")
        self._source_group = QButtonGroup(self)
        self._source_group.addButton(self._radio_src_keep)
        self._source_group.addButton(self._radio_src_analyze)
        self._radio_src_keep.toggled.connect(self._on_source_changed)
        self._radio_src_analyze.toggled.connect(self._on_source_changed)
        row1.addWidget(QLabel("数据来源:"))
        row1.addWidget(self._radio_src_analyze)
        row1.addWidget(self._radio_src_keep)

        _sep2 = QFrame(); _sep2.setFrameShape(QFrame.Shape.VLine)
        _sep2.setFrameShadow(QFrame.Shadow.Sunken)
        row1.addSpacing(4); row1.addWidget(_sep2); row1.addSpacing(4) 

        # 读取模式
        self._read_mode_group = QButtonGroup(self)
        self._radio_read_first = QRadioButton("只读.1")
        self._radio_read_last  = QRadioButton("最大编号")
        self._radio_read_all   = QRadioButton("全部汇总")
        self._radio_read_first.setToolTip(
            '仅读取名为"惯量包络体.1"的保持测量结果。\n'
            "速度最快：只进行一次参数查询，不扫描其余编号。"
        )
        self._radio_read_last.setToolTip(
            f"扫描编号 1 到 {MAX_INERTIA_INDEX} 的全部惯量包络体，使用编号最大的有效保持测量结果。\n"
            f"速度较慢：每个缺失的编号均会产生一次 COM 异常，最多 {MAX_INERTIA_INDEX - 1} 次。"
        )
        self._radio_read_all.setToolTip(
            f"扫描编号 1 到 {MAX_INERTIA_INDEX} 的全部惯量包络体，读取所有有效测量并按平行轴定理汇总为单一质量特性。\n"
            f"速度较慢：每个缺失的编号均会产生一次 COM 异常，最多 {MAX_INERTIA_INDEX - 1} 次。"
        )
        self._radio_read_first.setChecked(self._read_mode == "first")
        self._radio_read_last.setChecked(self._read_mode == "last")
        self._radio_read_all.setChecked(self._read_mode == "all")
        self._read_mode_group.addButton(self._radio_read_first)
        self._read_mode_group.addButton(self._radio_read_last)
        self._read_mode_group.addButton(self._radio_read_all)
        self._radio_read_first.toggled.connect(self._on_read_mode_changed)
        self._radio_read_last.toggled.connect(self._on_read_mode_changed)
        self._radio_read_all.toggled.connect(self._on_read_mode_changed)
        row1.addWidget(QLabel("惯量包络体读取:"))
        row1.addWidget(self._radio_read_first)
        row1.addWidget(self._radio_read_last)
        row1.addWidget(self._radio_read_all)   

        # 初始状态：analyze 时禁用读取模式控件
        _read_mode_enabled = (self._source == "keep_inertia")
        for _w in (self._radio_read_first, self._radio_read_last, self._radio_read_all):
            _w.setEnabled(_read_mode_enabled)

        _sep3 = QFrame(); _sep3.setFrameShape(QFrame.Shape.VLine)
        _sep3.setFrameShadow(QFrame.Shadow.Sunken)
        row1.addSpacing(4); row1.addWidget(_sep3); row1.addSpacing(4)

        # 忽略隐藏节点
        self._skip_hidden_chk = QCheckBox("忽略隐藏的节点")
        self._skip_hidden_chk.setChecked(self._skip_hidden)
        self._skip_hidden_chk.setToolTip(
            "勾选时：零件处于隐藏状态则跳过；产品/部件处于隐藏状态则连同其子孙一并跳过"
        )
        self._skip_hidden_chk.toggled.connect(self._on_skip_hidden_changed)
        row1.addWidget(self._skip_hidden_chk)

        row1.addStretch()
        opts_main.addLayout(row1)

        # ── 第二行：重量单位 ｜ 长度单位 ｜ 惯量单位（4选1）｜ 显示列 ──
        row2 = QHBoxLayout()
        row2.setSpacing(6)

        # 重量单位
        self._mass_unit_group = QButtonGroup(self)
        self._radio_mass_g  = QRadioButton("g")
        self._radio_mass_kg = QRadioButton("kg")
        self._radio_mass_g.setMinimumHeight(24)
        self._radio_mass_g.setChecked(self._mass_unit == "g")
        self._radio_mass_kg.setChecked(self._mass_unit == "kg")
        self._mass_unit_group.addButton(self._radio_mass_g)
        self._mass_unit_group.addButton(self._radio_mass_kg)
        self._radio_mass_g.toggled.connect(self._on_unit_changed)
        self._radio_mass_kg.toggled.connect(self._on_unit_changed)
        row2.addWidget(QLabel("重量:"))
        row2.addWidget(self._radio_mass_g)
        row2.addWidget(self._radio_mass_kg)

        _sep3 = QFrame(); _sep3.setFrameShape(QFrame.Shape.VLine)
        _sep3.setFrameShadow(QFrame.Shadow.Sunken)
        row2.addSpacing(4); row2.addWidget(_sep3); row2.addSpacing(4)

        # 长度单位
        self._cog_unit_group = QButtonGroup(self)
        self._radio_cog_mm = QRadioButton("mm")
        self._radio_cog_m  = QRadioButton("m")
        self._radio_cog_mm.setChecked(self._cog_unit == "mm")
        self._radio_cog_m.setChecked(self._cog_unit == "m")
        self._cog_unit_group.addButton(self._radio_cog_mm)
        self._cog_unit_group.addButton(self._radio_cog_m)
        self._radio_cog_mm.toggled.connect(self._on_unit_changed)
        self._radio_cog_m.toggled.connect(self._on_unit_changed)
        row2.addWidget(QLabel("长度:"))
        row2.addWidget(self._radio_cog_mm)
        row2.addWidget(self._radio_cog_m)

        _sep4 = QFrame(); _sep4.setFrameShape(QFrame.Shape.VLine)
        _sep4.setFrameShadow(QFrame.Shadow.Sunken)
        row2.addSpacing(4); row2.addWidget(_sep4); row2.addSpacing(4)

        # 惯量单位（4 选 1，独立）— QComboBox 节省水平空间
        _IU = ("g\u00b7mm\u00b2", "g\u00b7m\u00b2", "kg\u00b7mm\u00b2", "kg\u00b7m\u00b2")
        self._inertia_combo = QComboBox()
        for iu in _IU:
            self._inertia_combo.addItem(iu)
        self._inertia_combo.setCurrentText(self._inertia_unit)
        self._inertia_combo.setToolTip("选择转动惯量的显示单位")
        self._inertia_combo.currentTextChanged.connect(self._on_inertia_unit_changed)
        row2.addWidget(QLabel("惯量:"))
        row2.addWidget(self._inertia_combo)

        _sep5 = QFrame(); _sep5.setFrameShape(QFrame.Shape.VLine)
        _sep5.setFrameShadow(QFrame.Shadow.Sunken)
        row2.addSpacing(4); row2.addWidget(_sep5); row2.addSpacing(4)

        # 显示列
        row2.addWidget(QLabel("显示列:"))
        self._hid_col_checks: dict[str, QCheckBox] = {}
        for col_name in MASS_PROPS_HIDEABLE_COLUMNS:
            cb = QCheckBox(MASS_PROPS_COLUMN_DISPLAY_NAMES.get(col_name, col_name))
            cb.setChecked(col_name in self._visible_hideable_cols)
            cb.setProperty("col_name", col_name)
            cb.toggled.connect(self._on_col_visibility_changed)
            row2.addWidget(cb)
            self._hid_col_checks[col_name] = cb

        row2.addStretch()
        opts_main.addLayout(row2)

        layout.addWidget(opts_group)

        # ── BOM说明标签 ─────────────────────────────────────────────────────
        self._bom_desc_lbl = QLabel(self._bom_desc_text())
        self._bom_desc_lbl.setWordWrap(True)
        self._bom_desc_lbl.setStyleSheet(self._bom_desc_label_style(theme_manager.current_mode()))
        layout.addWidget(self._bom_desc_lbl)

        # ── 搜索筛选框（Ctrl+F） ──────────────────────────────────────────────
        filter_row = QHBoxLayout()
        filter_row.setSpacing(6)
        filter_row.addWidget(QLabel("筛选:"))
        self._filter_edit = QLineEdit()
        self._filter_edit.setPlaceholderText("按零件编号、文件名、术语等关键字搜索行… (Ctrl+F)")
        self._filter_edit.setClearButtonEnabled(True)
        self._filter_edit.textChanged.connect(self._on_filter_changed)
        filter_row.addWidget(self._filter_edit)
        layout.addLayout(filter_row)
        QShortcut(QKeySequence("Ctrl+F"), self, self._focus_filter)

        # ── 树形表格 ────────────────────────────────────────────────────────
        self._table = _BomTreeWidget()
        self._table.setColumnCount(len(self._columns))
        self._table.setHeaderLabels(self._display_headers())
        hdr = self._table.header()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(False)
        hdr.setFixedHeight(L.TABLE_ROW_HEIGHT)
        self._table.setUniformRowHeights(True)
        self._table.setRootIsDecorated(True)
        self._table.setSortingEnabled(False)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        # 不使用交替行色：Qt QSS 的 branch 伪元素不支持 :alternate，
        # 开启后 branch 列背景无法同步，且选中行颜色因奇偶行底色不同而出现色差。
        self._table.setAlternatingRowColors(True)
        self._table.setIndentation(16)
        self._table.setItemDelegate(_MassPropsDelegate(lambda: self._columns, self._table))
        self._table.itemChanged.connect(self._on_item_changed)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_tree_context_menu)
        hdr.sectionResized.connect(self._on_section_resized)
        layout.addWidget(self._table, 1)

        # ── 汇总面板 ────────────────────────────────────────────────────────
        summary_group = QGroupBox("汇总结果（基于根产品坐标系）")
        summary_h = QHBoxLayout(summary_group)
        summary_h.setSpacing(12)
        summary_h.setContentsMargins(10, 8, 10, 8)
        summary_h.setAlignment(Qt.AlignmentFlag.AlignBottom)

        def _sec_lbl(text: str) -> QLabel:
            lb = QLabel(text)
            lb.setObjectName("sectionLabel")
            return lb

        def _fld_lbl(text: str) -> QLabel:
            return QLabel(text)

        def _val_edit(min_w: int = 80) -> QLineEdit:
            le = QLineEdit("—")
            le.setReadOnly(True)
            le.setMinimumWidth(min_w)
            return le

        # ── 左列：总重量 + 重心 (G) ─────────────────────────────────────────
        left_w = QWidget()
        left_g = QGridLayout(left_w)
        left_g.setSpacing(4)
        left_g.setContentsMargins(0, 0, 0, 0)

        left_g.addWidget(_sec_lbl("总重量"), 0, 0, 1, 2)
        left_g.addWidget(_fld_lbl("m"), 1, 0)
        self._edit_weight = _val_edit()
        left_g.addWidget(self._edit_weight, 1, 1)

        left_g.addWidget(_sec_lbl("重心 (G)"), 2, 0, 1, 2)
        for r, (text, attr) in enumerate([("Gx", "_edit_cx"), ("Gy", "_edit_cy"), ("Gz", "_edit_cz")]):
            left_g.addWidget(_fld_lbl(text), r + 3, 0)
            edit = _val_edit()
            setattr(self, attr, edit)
            left_g.addWidget(edit, r + 3, 1)
        left_g.setRowStretch(left_g.rowCount(), 1)
        summary_h.addWidget(left_w, 0, Qt.AlignmentFlag.AlignBottom)

        # ── 中列：惯量矩阵 (3×3) ────────────────────────────────────────────
        mid_w = QWidget()
        mid_g = QGridLayout(mid_w)
        mid_g.setSpacing(4)
        mid_g.setContentsMargins(0, 0, 0, 0)

        mid_g.addWidget(_sec_lbl("惯量矩阵"), 0, 0, 1, 6)
        _inertia_rows = [
            ("Ixx", "_edit_ixx", "Ixy", "_edit_ixy", "Ixz", "_edit_ixz"),
            ("Iyx", "_edit_iyx", "Iyy", "_edit_iyy", "Iyz", "_edit_iyz"),
            ("Izx", "_edit_izx", "Izy", "_edit_izy", "Izz", "_edit_izz"),
        ]
        for r, row_items in enumerate(_inertia_rows):
            for ci in range(3):
                mid_g.addWidget(_fld_lbl(row_items[ci * 2]), r + 1, ci * 2)
                edit = _val_edit()
                setattr(self, row_items[ci * 2 + 1], edit)
                mid_g.addWidget(edit, r + 1, ci * 2 + 1)
        mid_g.setRowStretch(mid_g.rowCount(), 1)
        summary_h.addWidget(mid_w, 0, Qt.AlignmentFlag.AlignBottom)

        # ── 右列：重心主惯量矩 + 主轴 ───────────────────────────────────────
        right_w = QWidget()
        right_g = QGridLayout(right_w)
        right_g.setSpacing(4)
        right_g.setContentsMargins(0, 0, 0, 0)

        right_g.addWidget(_sec_lbl("重心主惯量矩"), 0, 0, 1, 6)
        for c, (text, attr) in enumerate([("M1", "_edit_m1"), ("M2", "_edit_m2"), ("M3", "_edit_m3")]):
            right_g.addWidget(_fld_lbl(text), 1, c * 2)
            edit = _val_edit()
            setattr(self, attr, edit)
            right_g.addWidget(edit, 1, c * 2 + 1)

        right_g.addWidget(_sec_lbl("主轴"), 2, 0, 1, 6)
        _axes_rows = [
            ("A1x", "_edit_a1x", "A2x", "_edit_a2x", "A3x", "_edit_a3x"),
            ("A1y", "_edit_a1y", "A2y", "_edit_a2y", "A3y", "_edit_a3y"),
            ("A1z", "_edit_a1z", "A2z", "_edit_a2z", "A3z", "_edit_a3z"),
        ]
        for r, row_items in enumerate(_axes_rows):
            for ci in range(3):
                right_g.addWidget(_fld_lbl(row_items[ci * 2]), r + 3, ci * 2)
                edit = _val_edit(min_w=80)
                setattr(self, row_items[ci * 2 + 1], edit)
                right_g.addWidget(edit, r + 3, ci * 2 + 1)
        right_g.setRowStretch(right_g.rowCount(), 1)
        summary_h.addWidget(right_w, 0, Qt.AlignmentFlag.AlignBottom)

        layout.addWidget(summary_group)

        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(line)

        # ── 底部按钮行 ──────────────────────────────────────────────────────
        btn_row = QHBoxLayout()

        autofit_btn = QPushButton("自适应列宽")
        autofit_btn.clicked.connect(self._autofit_columns)
        btn_row.addWidget(autofit_btn)

        expand_btn = QPushButton("全部展开")
        expand_btn.clicked.connect(self._table.expandAll)
        btn_row.addWidget(expand_btn)

        collapse_btn = QPushButton("全部折叠")
        collapse_btn.clicked.connect(self._table.collapseAll)
        btn_row.addWidget(collapse_btn)

        btn_row.addStretch()

        self._save_json_btn = QPushButton("保存数据…")
        self._save_json_btn.setToolTip("将当前行数据保存为数据文件，可在不打开 CATIA 的情况下重新载入")
        self._save_json_btn.setEnabled(False)
        self._save_json_btn.clicked.connect(self._save_data_to_json)
        btn_row.addWidget(self._save_json_btn)

        self._export_btn = QPushButton("导出表格")
        self._export_btn.setToolTip(
            "将完整 BOM 数据（含汇总行）导出为 Excel （.xlsx）或 CSV 文件。\n"
            "无论当前显示完整 BOM 还是汇总 BOM ，导出内容始终为完整 BOM 。"
        )
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_table)
        btn_row.addWidget(self._export_btn)

        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.reject)
        btn_row.addWidget(close_btn)

        layout.addLayout(btn_row)

    # ── BOM 说明文字 ───────────────────────────────────────────────────────

    def _bom_desc_text(self) -> str:
        """返回当前 BOM 模式对应的说明文字。"""
        if self._summarize:
            return (
                "【汇总 BOM】按零件编号合并，仅列出零件（不含产品、部件和对称件）。"
                "Weight / CogX / CogY / CogZ / Ixx–Iyz "
                "在零件自身坐标系下显示，与装配位置无关。"
                "底部「汇总结果」在根产品坐标系下计算。"
            )
        return (
            "【完整 BOM】展示零件节点和产品/部件节点。"
            "Weight / CogX / CogY / CogZ / Ixx–Iyz "
            "在根产品坐标系下显示，与零件的装配位置有关。"
            "底部「汇总结果」在根产品坐标系下计算。"
        )

    # ── 文件/活动文档切换 ──────────────────────────────────────────────────

    def _toggle_file_row(self, use_active: bool) -> None:
        self._file_edit.setEnabled(not use_active)
        self._file_browse_btn.setEnabled(not use_active)

    def _browse_file(self) -> None:
        file, _ = QFileDialog.getOpenFileName(
            self, "选择 CATProduct 文件",
            self._last_browse_dir,
            "*.CATProduct (*.CATProduct);;All Files (*)",
        )
        if file:
            self._file_edit.setText(file)
            self._last_browse_dir = str(Path(file).parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)

    # ── 标题栏脏标志 ────────────────────────────────────────────────────────

    def _update_title(self) -> None:
        """在标题栏末尾追加 ' *' 表示有未保存到磁盘的编辑；清除则恢复原标题。"""
        base = "质量特性工作台"
        self.setWindowTitle(f"{base} *" if self._is_dirty else base)

    # ── 搜索筛选 ───────────────────────────────────────────────────────────

    def _focus_filter(self) -> None:
        """将输入焦点移至搜索框（Ctrl+F）。"""
        self._filter_edit.setFocus()
        self._filter_edit.selectAll()

    def _on_filter_changed(self, text: str) -> None:
        """搜索框文本变化时，重新过滤表格中的行。"""
        self._filter_text = text.strip().lower()
        self._apply_filter()

    def _apply_filter(self) -> None:
        """根据 _filter_text 显示/隐藏树形控件中的各行。"""
        pattern = self._filter_text
        root = self._table.invisibleRootItem()
        for i in range(root.childCount()):
            self._apply_filter_to_item(root.child(i), pattern)

    def _apply_filter_to_item(self, item: QTreeWidgetItem, pattern: str) -> bool:
        """递归地对 *item* 及其子项应用过滤。

        返回 True 表示该项或其任意后代匹配 *pattern*，
        同时设置 item.setHidden() 的可见性。
        """
        if not pattern:
            item.setHidden(False)
            for i in range(item.childCount()):
                self._apply_filter_to_item(item.child(i), pattern)
            return True
        self_match = self._item_matches_filter(item, pattern)
        child_match = False
        for i in range(item.childCount()):
            child_match |= self._apply_filter_to_item(item.child(i), pattern)
        visible = self_match or child_match
        item.setHidden(not visible)
        return visible

    def _item_matches_filter(self, item: QTreeWidgetItem, pattern: str) -> bool:
        """判断 *item* 任意列的文本是否包含 *pattern*（大小写不敏感）。"""
        for col_idx in range(len(self._columns)):
            if pattern in item.text(col_idx).lower():
                return True
        return False

    # ── 显示选项响应 ───────────────────────────────────────────────────────

    def _on_bom_type_changed(self, checked: bool) -> None:
        self._summarize = self._radio_summ.isChecked()
        self._settings.setValue("summarize", self._summarize)
        self._bom_desc_lbl.setText(self._bom_desc_text())
        self._rebuild_columns_and_table()

    def _on_read_mode_changed(self, checked: bool) -> None:
        if not checked:
            return
        if self._radio_read_first.isChecked():
            self._read_mode = "first"
        elif self._radio_read_last.isChecked():
            self._read_mode = "last"
        else:
            self._read_mode = "all"
        self._settings.setValue("read_mode", self._read_mode)

    def _on_source_changed(self, checked: bool) -> None:
        if not checked:
            return
        self._source = "analyze" if self._radio_src_analyze.isChecked() else "keep_inertia"
        self._settings.setValue("source", self._source)
        # 读取模式控件仅在 keep_inertia 时有意义
        _enabled = (self._source == "keep_inertia")
        for _w in (self._radio_read_first, self._radio_read_last, self._radio_read_all):
            _w.setEnabled(_enabled)

    def _on_skip_hidden_changed(self, checked: bool) -> None:
        self._skip_hidden = self._skip_hidden_chk.isChecked()
        self._settings.setValue("skip_hidden", self._skip_hidden)

    def _on_unit_changed(self, checked: bool) -> None:
        if not checked:
            return
        self._mass_unit = "g" if self._radio_mass_g.isChecked() else "kg"
        self._cog_unit  = "mm" if self._radio_cog_mm.isChecked() else "m"
        self._unit_factor, _, self._cog_unit_factor = (
            self._calc_unit_factors(self._mass_unit, self._cog_unit)
        )
        self._settings.setValue("mass_unit", self._mass_unit)
        self._settings.setValue("cog_unit", self._cog_unit)
        if self._rows:
            self._refresh_unit_display()

    def _on_inertia_unit_changed(self, text: str = "") -> None:
        self._inertia_unit = self._inertia_combo.currentText()
        self._inertia_unit_factor = self._calc_inertia_factor(self._inertia_unit)
        self._settings.setValue("inertia_unit", self._inertia_unit)
        if self._rows:
            self._refresh_unit_display()

    def _on_col_visibility_changed(self, checked: bool) -> None:
        cb = self.sender()
        if cb is not None:
            col_name = cb.property("col_name")
            if col_name:
                if checked:
                    self._visible_hideable_cols.add(col_name)
                else:
                    self._visible_hideable_cols.discard(col_name)
        self._settings.setValue("visible_hideable_cols",
                                list(self._visible_hideable_cols))
        self._rebuild_columns_and_table()

    def _rebuild_columns_and_table(self) -> None:
        """重建列列表并重新填充表格（保留列宽）。"""
        if self._rows:
            # 重建前保存各列宽度
            for col_idx, col_name in enumerate(self._columns):
                self._col_widths[col_name] = self._table.columnWidth(col_idx)
        self._columns = self._build_columns()
        self._populate_table()
        # 恢复各列宽度
        for col_idx, col_name in enumerate(self._columns):
            if col_name in self._col_widths:
                self._table.setColumnWidth(col_idx, self._col_widths[col_name])

    def _refresh_unit_display(self) -> None:
        """仅更新列标题和重量/惯量单元格的显示值（单位切换时调用，避免全量重建）。"""
        # 更新列标题
        self._table.setHeaderLabels(self._display_headers())

        mass_col_indices: list[tuple[str, int]] = []
        for col_name in _UNIT_SENSITIVE_COLUMNS:
            if col_name in self._columns:
                mass_col_indices.append((col_name, self._columns.index(col_name)))

        if not mass_col_indices:
            return

        self._is_updating = True
        display_rows = self._get_display_rows()
        try:
            for di, row_data in enumerate(display_rows):
                if di >= len(self._item_by_row):
                    break
                item = self._item_by_row[di]
                if not any(row_data.get(c) is not None for c in _UNIT_SENSITIVE_COLUMNS):
                    continue
                for col_name, col_idx in mass_col_indices:
                    raw = row_data.get(col_name)
                    if raw is not None:
                        if col_name == "Weight":
                            item.setText(col_idx, self._fmt_mass_val(raw))
                        elif col_name in _INERTIA_IDX:
                            item.setText(col_idx, self._fmt_inertia_val(raw))
                        else:
                            item.setText(col_idx, self._fmt_cog_val(raw))
        finally:
            self._is_updating = False

        # 若已有汇总结果，更新底部汇总标签
        if self._rollup_result:
            self._update_summary_labels(self._rollup_result)

    # ── 加载数据 ───────────────────────────────────────────────────────────

    def _load_data(self) -> None:
        if self._use_active_chk.isChecked():
            file_path = None
        else:
            file_path = self._file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请先选择一个 CATProduct 文件。")
                return
            if not Path(file_path).exists():
                QMessageBox.warning(self, "文件不存在", f"文件不存在：\n{file_path}")
                return

        self._load_btn.setEnabled(False)
        self._load_btn.setText("加载中…")
        QApplication.processEvents()

        progress = QProgressDialog("正在加载产品树，请稍候…", None, 0, 0, self)
        progress.setWindowTitle("加载质量特性")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(300)
        progress.setValue(0)

        def _on_row_collected(count: int) -> None:
            progress.setLabelText(f"正在加载产品树，请稍候… 已读取 {count} 个节点")
            progress.repaint()
            QApplication.processEvents()

        try:
            rows = collect_mass_props_rows(file_path, progress_callback=_on_row_collected,
                                           read_mode=self._read_mode,
                                           skip_hidden=self._skip_hidden,
                                           source=self._source)
        except Exception as e:
            progress.close()
            logger.error(f"加载质量特性失败: {e}")
            QMessageBox.critical(
                self, "加载失败",
                f"加载产品树时出错：\n{e}\n\n请确保 CATIA 已启动。",
            )
            self._load_btn.setEnabled(True)
            self._load_btn.setText("加载")
            return
        finally:
            progress.close()

        self._load_btn.setEnabled(True)
        self._load_btn.setText("重新加载")

        self._apply_loaded_rows(rows)

    def _apply_loaded_rows(self, rows: list[dict], new_rows: list[dict] | None = None) -> None:
        """将已就绪的行列表应用到对话框：重建表格、调整列宽、启用按钮并计算。

        由 :meth:`_load_data`、:meth:`_load_data_from_json`、
        :meth:`_append_data_from_file`、:meth:`_append_from_active` 共用。

        参数：
            rows:      要设置为 ``self._rows`` 的完整行列表。
            new_rows:  本次操作新增的行（用于"部分零件测量失败"提示）。
                       ``None`` 表示整个 *rows* 均为新增（全量载入场景）。
        """
        # 重新填充前保存列宽
        if self._loaded:
            for col_idx, col_name in enumerate(self._columns):
                self._col_widths[col_name] = self._table.columnWidth(col_idx)

        self._rows = rows
        self._rollup_result = None
        self._clear_summary_labels()
        self._columns = self._build_columns()
        self._populate_table()

        if not self._loaded:
            # 首次加载时自适应列宽
            for _c, col_name in enumerate(self._columns):
                if col_name == "#":
                    self._table.setColumnWidth(_c, 40)
                    self._col_widths[col_name] = 40
                else:
                    self._table.resizeColumnToContents(_c)
                    self._col_widths[col_name] = self._table.columnWidth(_c)
            self._loaded = True
        else:
            for col_idx, col_name in enumerate(self._columns):
                if col_name in self._col_widths:
                    self._table.setColumnWidth(col_idx, self._col_widths[col_name])

        self._export_btn.setEnabled(True)
        self._save_json_btn.setEnabled(True)
        self._append_data_btn.setEnabled(True)
        self._append_active_btn.setEnabled(True)

        failed_count = sum(
            1 for r in (new_rows if new_rows is not None else rows)
            if r.get("_meas_failed") and r.get("Type") == BomNodeType.PART
        )
        if failed_count:
            _read_mode_desc = {
                "first": "「惯量包络体.1」",
                "last":  f"编号最大的「惯量包络体.N」（N ≤ {MAX_INERTIA_INDEX}）",
                "all":   f"「惯量包络体.1」至「惯量包络体.{MAX_INERTIA_INDEX}」",
            }.get(self._read_mode, "惯量包络体")
            QMessageBox.information(
                self, "部分零件测量失败",
                f"有 {failed_count} 个零件节点无法完成质量特性测量（显示橙色背景）。\n\n"
                "可能原因：\n"
                "  • 零件文档无法加载到 CATIA 会话中\n"
                f"  • 当前读取模式要求的 {_read_mode_desc} 保持测量不存在\n"
                "  • 测量结果保存在产品中而非零件中\n\n"
                "未能测量的零件不参与最终汇总计算。可以对其右键选择「在 CATIA 中打开」\n"
                "检查问题所在，然后右键选择「重新读取质量特性」。",
            )


        # 加载完成后自动计算汇总结果
        self._calculate()
        # 新数据加载后重置脏标志
        self._is_dirty = False
        self._update_title()

    def _save_data_to_json(self) -> None:
        """将当前行数据保存为压缩二进制数据文件（不包含 _root_mp，可重新计算）。"""
        if not self._rows:
            return

        # ── 默认文件名：根产品零件编号 + "_惯量汇总" ───────────────────────
        root_pn = str(self._rows[0].get("Part Number", "")).strip()
        default_name = f"{root_pn}_惯量汇总" if root_pn else "惯量汇总"

        # ── 默认目录：上次浏览目录 → 根产品文件所在目录 → 空 ──────────────
        if self._last_browse_dir and Path(self._last_browse_dir).is_dir():
            default_dir = self._last_browse_dir
        else:
            root_fp = str(self._rows[0].get("_filepath", "")).strip()
            default_dir = str(Path(root_fp).parent) if root_fp else ""

        default_path = str(Path(default_dir) / default_name) if default_dir else default_name

        dest, _ = QFileDialog.getSaveFileName(
            self, "保存质量特性数据", default_path, "质量特性数据文件 (*.mpd)"
        )
        if not dest:
            return
        if not dest.lower().endswith(".mpd"):
            dest += ".mpd"
        try:
            save_rows(self._rows, dest)
            self._last_browse_dir = str(Path(dest).parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)
            # 保存成功后清除脏标志
            self._is_dirty = False
            self._update_title()
        except Exception as e:
            logger.error(f"保存质量特性数据失败: {e}")
            QMessageBox.critical(self, "保存失败", f"保存数据时出错：\n{e}")

    def _load_data_from_json(self) -> None:
        """从压缩二进制数据文件载入行数据（无需 CATIA ，_root_mp 由后处理重建）。"""
        src, _ = QFileDialog.getOpenFileName(
            self, "载入质量特性数据", "", "质量特性数据文件 (*.mpd)"
        )
        if not src:
            return
        if not Path(src).exists():
            QMessageBox.warning(self, "文件不存在", f"文件不存在：\n{src}")
            return
        try:
            rows = load_rows(src)
        except Exception as e:
            logger.error(f"载入质量特性数据失败: {e}")
            QMessageBox.critical(self, "载入失败", f"载入数据时出错：\n{e}")
            return
        self._apply_loaded_rows(rows)

    def _append_data_from_file(self) -> None:
        """从一个或多个已保存的数据文件（.mpd）追加分总成数据并合并。

        支持多选文件，每个文件的行数据依次追加到 ``self._rows`` 中。
        适用于主产品过大、分批读取各分总成并在此汇总的工作流。
        前提：各分总成坐标系须与主产品（及彼此）一致，无须额外坐标变换。
        """
        if not self._rows:
            QMessageBox.warning(self, "无基础数据", "请先加载基础产品数据，再追加分总成数据。")
            return
        srcs, _ = QFileDialog.getOpenFileNames(
            self, "追加质量特性数据", self._last_browse_dir,
            "质量特性数据文件 (*.mpd)"
        )
        if not srcs:
            return

        combined = list(self._rows)
        errors: list[str] = []
        appended_rows: list[dict] = []
        for src in srcs:
            if not Path(src).exists():
                errors.append(f"文件不存在：{src}")
                continue
            try:
                extra = load_rows(src)
                combined = merge_rows(combined, extra)
                appended_rows.extend(extra)
                self._last_browse_dir = str(Path(src).parent)
                self._settings.setValue("last_browse_dir", self._last_browse_dir)
            except Exception as e:
                logger.error(f"追加质量特性数据失败 ({src}): {e}")
                errors.append(f"{Path(src).name}：{e}")

        if errors:
            QMessageBox.warning(
                self, "部分文件追加失败",
                "以下文件追加时出错：\n\n" + "\n".join(errors),
            )

        if appended_rows:
            self._apply_loaded_rows(combined, new_rows=appended_rows)

    def _append_from_active(self) -> None:
        """将 CATIA 当前活动文档（分总成）的质量特性追加到现有数据中。

        调用 :func:`collect_mass_props_rows` 读取当前 CATIA 活动文档（需已在
        CATIA 中打开对应的 CATProduct），然后用 :func:`merge_rows` 追加到
        ``self._rows`` 并刷新显示与汇总。
        前提：活动文档坐标系须与已加载数据的坐标系一致。
        """
        if not self._rows:
            QMessageBox.warning(self, "无基础数据", "请先加载基础产品数据，再追加分总成数据。")
            return

        self._append_active_btn.setEnabled(False)
        self._append_active_btn.setText("读取中…")
        QApplication.processEvents()

        progress = QProgressDialog("正在读取当前活动文档，请稍候…", None, 0, 0, self)
        progress.setWindowTitle("追加质量特性")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(300)
        progress.setValue(0)

        def _on_row_collected(count: int) -> None:
            progress.setLabelText(f"正在读取当前活动文档，请稍候… 已读取 {count} 个节点")
            progress.repaint()
            QApplication.processEvents()

        extra: list[dict] = []
        try:
            extra = collect_mass_props_rows(
                None,
                progress_callback=_on_row_collected,
                read_mode=self._read_mode,
                skip_hidden=self._skip_hidden,
                source=self._source,
            )
        except Exception as e:
            logger.error(f"追加活动文档质量特性失败: {e}")
            QMessageBox.critical(
                self, "读取失败",
                f"读取当前活动文档时出错：\n{e}\n\n请确保 CATIA 已启动且有活动文档。",
            )
            return
        finally:
            progress.close()
            self._append_active_btn.setEnabled(True)
            self._append_active_btn.setText("追加活动文档…")

        if not extra:
            QMessageBox.information(self, "无数据", "当前活动文档未读取到任何节点，未进行追加。")
            return

        combined = merge_rows(self._rows, extra)
        self._apply_loaded_rows(combined, new_rows=extra)

    # ── 构建显示行 ─────────────────────────────────────────────────────────

    def _get_display_rows(self) -> list[dict]:
        """返回当前模式下应显示的行列表。"""
        if self._summarize:
            return self._build_summary_rows()
        # 层级BOM：展示全部节点（零件、产品、部件），使用根产品坐标系下的值。
        # 每行附加 _rows_idx，指向 self._rows 中的原始索引，
        # 以确保 _make_item / _on_item_changed 能正确回写数据。
        # 对零件行，将 _root_mp 中的根坐标系 COG / 惯量值覆盖显示字段；
        # 产品/部件行的显示字段已由 _post_process_rows() 写入根坐标系汇总值。
        result = []
        for i, row in enumerate(self._rows):
            r = dict(row)
            r["_rows_idx"] = i
            if r.get("Type") in BomNodeType.LEAF_TYPES:
                rmp = r.get("_root_mp")
                if rmp:
                    cog = rmp.get("cog", [None, None, None])
                    r["CogX"] = cog[0]
                    r["CogY"] = cog[1]
                    r["CogZ"] = cog[2]
                    I = rmp.get("inertia")
                    if I:
                        r["Ixx"] = I[0][0]
                        r["Iyy"] = I[1][1]
                        r["Izz"] = I[2][2]
                        r["Ixy"] = I[0][1]
                        r["Ixz"] = I[0][2]
                        r["Iyz"] = I[1][2]
            result.append(r)
        return result

    def _build_hierarchy_columns(self) -> list[str]:
        """返回完整 BOM 的列名列表（导出时始终使用，与当前显示模式无关）。"""
        base = ["Level", "#", "Type"]
        for c in MASS_PROPS_HIDEABLE_COLUMNS:
            if c in self._visible_hideable_cols:
                base.append(c)
        base += ["Density", "Weight", "CogX", "CogY", "CogZ",
                 "Ixx", "Iyy", "Izz", "Ixy", "Ixz", "Iyz"]
        return base

    def _get_hierarchy_rows(self) -> list[dict]:
        """返回完整 BOM 行列表（导出时始终使用，与当前显示模式无关）。

        与 _get_display_rows() 的非汇总分支相同：包含所有节点（零件、产品、
        部件、对称件），使用根产品坐标系下的 COG / 惯量值。
        """
        result = []
        for i, row in enumerate(self._rows):
            r = dict(row)
            r["_rows_idx"] = i
            if r.get("Type") in BomNodeType.LEAF_TYPES:
                rmp = r.get("_root_mp")
                if rmp:
                    cog = rmp.get("cog", [None, None, None])
                    r["CogX"] = cog[0]
                    r["CogY"] = cog[1]
                    r["CogZ"] = cog[2]
                    I = rmp.get("inertia")
                    if I:
                        r["Ixx"] = I[0][0]
                        r["Iyy"] = I[1][1]
                        r["Izz"] = I[2][2]
                        r["Ixy"] = I[0][1]
                        r["Ixz"] = I[0][2]
                        r["Iyz"] = I[1][2]
            result.append(r)
        return result

    def _build_summary_rows(self) -> list[dict]:
        """汇总模式：将相同零件编号的行合并，增加 Quantity 字段。

        每个唯一 PN 保留第一次出现的未排除行数据（含 _rows 中的索引），
        Quantity = 该 PN 在 _rows 中未被排除的实例数量。
        被排除（_excluded=True）的实例不计入数量，
        若某 PN 的全部实例均被排除，则该 PN 不出现在汇总 BOM 中。
        对称件的重心/惯量依赖根产品坐标系（位置相关），无法在汇总 BOM 中有意义地
        合并展示，故汇总 BOM 仅包含零件行，不含对称件。
        """
        seen_pn: dict[str, dict] = {}    # pn → 首次出现的未排除规范行副本
        qty: dict[str, int] = {}
        order: list[str] = []

        for i, row in enumerate(self._rows):
            if row.get("_excluded"):
                continue
            # 汇总BOM仅统计零件行；产品/部件/对称件不计入数量，也不占用 PN 的 seen_pn 位置，
            # 否则产品行会成为该 PN 的"规范行"，随后被类型过滤器删除，导致该 PN 的
            # 零件实例在汇总BOM中完全消失，且数量也会被错误地计入产品实例。
            # 对称件：其重心/惯量是位置相关量（根坐标系），在汇总上下文中无意义，故排除。
            if row.get("Type") != BomNodeType.PART:
                continue
            pn = str(row.get("Part Number", ""))
            if not pn:
                pn = str(row.get("Filename", "")) or "(未分组)"
            if pn not in seen_pn:
                r = dict(row)
                r["_rows_idx"] = i   # 映射回 _rows 的规范索引
                seen_pn[pn] = r
                qty[pn] = 1
                order.append(pn)
            else:
                qty[pn] += 1

        result = []
        for pn in order:
            r = dict(seen_pn[pn])
            r["Quantity"] = qty[pn]
            result.append(r)

        return result

    # ── 填充表格 ───────────────────────────────────────────────────────────

    def _populate_table(self) -> None:
        self._is_updating = True
        self._table.blockSignals(True)
        
        # 汇总模式：所有行均为无子项的顶层项。
        # 若保持 setRootIsDecorated(True)，每行都会预留展开箭头的空间，
        # 使第0列内容向右偏移。汇总模式下禁用，层级模式下重新启用以显示展开箭头。
        self._table.setRootIsDecorated(not self._summarize)

        # 插入行前必须先禁用排序，否则 addTopLevelItem/addChild 会立即按当前排序列重排，
        # 导致完整 BOM / 层级 BOM 的树形顺序被破坏。末尾再按模式决定是否重新启用。
        self._table.setSortingEnabled(False)
    
        self._table.clear()
        self._table.setColumnCount(len(self._columns))
        self._table.setHeaderLabels(self._display_headers())
        self._table.setRootIsDecorated(not self._summarize)
        self._item_by_row = []
        self._pn_to_items.clear()

        display_rows = self._get_display_rows()

        if self._summarize:
            self._populate_flat(display_rows)
        else:
            self._populate_tree(display_rows)

        self._table.expandAll()
        self._table.blockSignals(False)
        self._is_updating = False

        # 汇总模式：启用表头点击排序；层级模式：禁用（保持树结构）
        self._table.setSortingEnabled(self._summarize)

        # 保持当前搜索筛选状态
        if self._filter_text:
            self._apply_filter()

    def _make_item(self, row_idx: int, row_data: dict) -> QTreeWidgetItem:
        """构建并填充一行的 QTreeWidgetItem。

        row_idx: 对应的 self._rows 索引（汇总模式用 _rows_idx 字段）。
        """
        item = _BomSortItem()
        item.setData(0, _ROW_IDX_ROLE, row_idx)

        pn          = str(row_data.get("Part Number", ""))
        not_found   = bool(row_data.get("_not_found"))
        no_file     = bool(row_data.get("_no_file"))
        unreadable  = bool(row_data.get("_unreadable"))
        meas_failed = bool(row_data.get("_meas_failed"))
        node_type   = str(row_data.get("Type", ""))
        is_mirror   = bool(row_data.get("_is_mirror"))
        # 对称件视为锁定行（不可编辑质量/密度）
        row_locked  = unreadable or not_found or meas_failed or is_mirror

        if pn:
            self._pn_to_items.setdefault(pn, []).append(item)

        seq_no = str(len(self._item_by_row) + 1)

        for col_idx, col_name in enumerate(self._columns):
            if col_name == "#":
                item.setText(col_idx, seq_no)
            elif col_name == "Level":
                item.setText(col_idx, str(row_data.get("Level", 0)))
            elif col_name == "Filename":
                fp = str(row_data.get("_filepath", ""))
                fn = str(row_data.get("Filename", ""))
                if no_file:
                    value = FILENAME_UNSAVED
                else:
                    value = Path(fp).name if fp else fn
                item.setText(col_idx, value)
                if no_file:
                    pass  # tooltip 由下方 no_file 块统一设置
                elif fp:
                    item.setToolTip(col_idx, fp)
            elif col_name == "Quantity":
                item.setText(col_idx, str(row_data.get("Quantity", 1)))
            elif col_name == "Density":
                density = row_data.get("Density")
                if density is None:
                    # 对称件：按原件类型决定空显示（产品/部件→""，零件→"—"）
                    effective_type = row_data.get("_mirror_src_type") if node_type == BomNodeType.MIRROR else node_type
                    item.setText(col_idx, "" if (effective_type if effective_type is not None else node_type) in BomNodeType.ASSEMBLY_TYPES else "—")
                elif density < 0:
                    item.setText(col_idx, "不统一")
                else:
                    item.setText(col_idx, _fmt(density))
            elif col_name == "Weight":
                raw = row_data.get("Weight")
                if raw is None:
                    effective_type = row_data.get("_mirror_src_type") if node_type == BomNodeType.MIRROR else node_type
                    item.setText(col_idx, "" if (effective_type if effective_type is not None else node_type) in BomNodeType.ASSEMBLY_TYPES else "—")
                else:
                    item.setText(col_idx, self._fmt_mass_val(raw))
            elif col_name in _INERTIA_IDX or col_name in ("CogX", "CogY", "CogZ"):
                raw = row_data.get(col_name)
                if raw is None:
                    effective_type = row_data.get("_mirror_src_type") if node_type == BomNodeType.MIRROR else node_type
                    item.setText(col_idx, "" if (effective_type if effective_type is not None else node_type) in BomNodeType.ASSEMBLY_TYPES else "—")
                else:
                    if col_name in _INERTIA_IDX:
                        item.setText(col_idx, self._fmt_inertia_val(raw))
                    else:
                        item.setText(col_idx, self._fmt_cog_val(raw))
            elif col_name == "Type":
                # 存储英文 key，显示时转为中文
                raw = str(row_data.get("Type", ""))
                item.setText(col_idx, TYPE_DISPLAY_NAMES.get(raw, raw))
            else:
                item.setText(col_idx, str(row_data.get(col_name, "")))

        # 可编辑性：仅未锁定零件行的 Weight 和 Density（有效值）列可编辑
        if node_type == BomNodeType.PART and not row_locked:
            item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
            item.setData(0, _ITEM_LOCKED_ROLE, False)
        else:
            item.setData(0, _ITEM_LOCKED_ROLE, True)

        # 密度列锁定：密度为不统一（-1）或无数据（None）时不允许编辑密度列
        density_val = row_data.get("Density")
        density_locked = (density_val is None) or (density_val < 0)
        item.setData(0, _DENSITY_LOCKED_ROLE, density_locked)

        # 行背景色设置
        self._apply_row_state_style(item, row_data)

        # 排除状态：覆盖背景色、设置斜体灰色前景
        is_excluded = bool(row_data.get("_excluded", False))
        item.setData(0, _EXCLUDED_ROLE, is_excluded)
        if is_excluded:
            c = _get_colors(theme_manager.current_mode())
            excl_tip = "该行已被排除，不参与计算。"
            for ci in range(len(self._columns)):
                item.setBackground(ci, c.EXCL_BG)
                item.setForeground(ci, c.EXCL_FG)
                item.setFont(ci, _EXCL_FONT)
                item.setToolTip(ci, excl_tip)

        self._item_by_row.append(item)
        return item

    def _populate_flat(self, display_rows: list[dict]) -> None:
        """汇总 BOM 模式：所有行为顶级项（无树形层级）。"""
        for di, row_data in enumerate(display_rows):
            rows_idx = row_data.get("_rows_idx", di)
            item = self._make_item(rows_idx, row_data)
            self._table.addTopLevelItem(item)

    def _populate_tree(self, display_rows: list[dict]) -> None:
        """完整 BOM 模式：按 Level 构建树形结构。"""
        parent_stack: list[tuple[int, QTreeWidgetItem | None]] = [(-1, None)]

        for di, row_data in enumerate(display_rows):
            level = int(row_data.get("Level", 0))
            # 使用 _rows_idx（若存在）映射回 self._rows，保持与 _populate_flat 一致
            rows_idx = row_data.get("_rows_idx", di)

            while len(parent_stack) > 1 and parent_stack[-1][0] >= level:
                parent_stack.pop()

            parent_item = parent_stack[-1][1]
            item = self._make_item(rows_idx, row_data)

            if parent_item is None:
                self._table.addTopLevelItem(item)
            else:
                parent_item.addChild(item)

            parent_stack.append((level, item))

    # ── 单元格编辑 ─────────────────────────────────────────────────────────

    def _on_item_changed(self, item: QTreeWidgetItem, col_idx: int) -> None:
        if self._is_updating:
            return
        row_idx = item.data(0, _ROW_IDX_ROLE)
        if row_idx is None:
            return

        col_name = self._columns[col_idx]
        if col_name not in ("Weight", "Density"):
            return

        if item.data(0, _ITEM_LOCKED_ROLE):
            return

        # 密度列额外检查（-1 或 None 时不可编辑）
        if col_name == "Density" and item.data(0, _DENSITY_LOCKED_ROLE):
            return

        row_data = self._rows[row_idx]
        if row_data.get("Type") != BomNodeType.PART:
            return

        new_text = item.text(col_idx).strip()

        # ── 解析用户输入，计算新的内部存储重量和缩放比例 ─────────────────────
        if col_name == "Weight":
            try:
                # 输入值为当前显示单位；除以 _unit_factor 还原到内部单位（kg）
                new_weight_stored = float(new_text) / self._unit_factor
            except (ValueError, TypeError):
                self._is_updating = True
                item.setText(col_idx, self._fmt_mass_val(row_data.get("Weight")))
                self._is_updating = False
                return
            if new_weight_stored <= 0.0:
                QMessageBox.warning(
                    self, "重量不合法",
                    "重量必须为正数，请输入大于 0 的值。",
                )
                self._is_updating = True
                item.setText(col_idx, self._fmt_mass_val(row_data.get("Weight")))
                self._is_updating = False
                return
            # 重量变动时，按相同比例更新密度（密度 = 质量 / 体积，体积不变）
            _old_density = row_data.get("Density")
            _old_weight  = row_data.get("Weight")
            try:
                _od = float(_old_density) if _old_density is not None else 0.0
            except (ValueError, TypeError):
                _od = 0.0
            try:
                _ow = float(_old_weight) if _old_weight is not None else 0.0
            except (ValueError, TypeError):
                _ow = 0.0
            if _od > 0.0 and _ow > 0.0:
                new_density_stored = _od * (new_weight_stored / _ow)
            else:
                new_density_stored = _old_density  # 无法计算（旧密度/重量无效），密度保持不变
        else:  # col_name == "Density"
            try:
                new_density_stored = float(new_text)
            except (ValueError, TypeError):
                self._is_updating = True
                density_old = row_data.get("Density")
                item.setText(col_idx, _fmt(density_old) if density_old is not None and density_old >= 0 else "—")
                self._is_updating = False
                return
            if new_density_stored <= 0.0:
                QMessageBox.warning(
                    self, "密度不合法",
                    "密度必须为正数，请输入大于 0 的值。",
                )
                self._is_updating = True
                density_old = row_data.get("Density")
                item.setText(col_idx, _fmt(density_old) if density_old is not None and density_old >= 0 else "—")
                self._is_updating = False
                return
            # 密度按比例缩放重量（体积不变，密度×体积=质量）
            _raw_density = row_data.get("Density")
            _raw_weight  = row_data.get("Weight")
            try:
                old_density = float(_raw_density) if _raw_density is not None else 0.0
            except (ValueError, TypeError):
                old_density = 0.0
            try:
                old_weight = float(_raw_weight) if _raw_weight is not None else 0.0
            except (ValueError, TypeError):
                old_weight = 0.0
            if old_density > 0.0 and old_weight > 0.0:
                new_weight_stored = old_weight * (new_density_stored / old_density)
            else:
                new_weight_stored = old_weight  # 无法计算缩放，重量保持不变

        pn = str(row_data.get("Part Number", ""))

        # ── 更新 _rows 中所有相同 PN 的实例 ────────────────────────────────
        #
        # 同一零件的所有实例共享同一个 _mass_props dict（来自 _mass_cache）。
        # 若在循环内对每个实例各乘一次 scale，则第 n 个实例的惯量会被放大 scale^n 倍。
        # 正确做法：先从第一个匹配行计算 scale，对共享 dict 仅缩放一次，
        # 再遍历各实例分别用各自的 _placement 矩阵重新旋转到根坐标系。
        #
        # Step 1：从第一个匹配行取 scale 和共享的 _mass_props。
        scale: float = 1.0
        mp_shared: dict | None = None
        for r in self._rows:
            if str(r.get("Part Number", "")) == pn and r.get("Type") == BomNodeType.PART:
                try:
                    old_w_f_0 = float(r.get("Weight") or 0.0)
                except (ValueError, TypeError):
                    old_w_f_0 = 0.0
                if old_w_f_0 > 0.0:
                    scale = new_weight_stored / old_w_f_0
                mp_shared = r.get("_mass_props")
                break

        # Step 2：对共享 _mass_props 的惯量只缩放一次。
        if mp_shared is not None:
            mp_shared["weight"] = new_weight_stored
            if new_density_stored is not None and new_density_stored >= 0:
                mp_shared["density"] = new_density_stored
            if scale != 1.0:
                orig_i = mp_shared.get("inertia", [[0.0] * 3 for _ in range(3)])
                mp_shared["inertia"] = [[orig_i[ir][ic] * scale for ic in range(3)]
                                        for ir in range(3)]

        # Step 3：遍历所有实例，更新 Weight / Density / 行级显示字段 / _root_mp。
        # 对称件虚拟行（_is_mirror=True）不参与此处的 PN 批量更新：
        # 其数据由 _sync_all_mirrors() 在 _calculate() 中统一重算。
        for r in self._rows:
            if (str(r.get("Part Number", "")) != pn
                    or r.get("Type") != BomNodeType.PART
                    or r.get("_is_mirror")):
                continue
            r["Weight"] = new_weight_stored
            if new_density_stored is not None and new_density_stored >= 0:
                r["Density"] = new_density_stored
            mp = r.get("_mass_props")
            if mp:
                # 对于从文件载入的数据，各实例的 _mass_props 独立序列化，
                # 与 mp_shared 不共享同一 dict 对象，需在此独立更新内部字段。
                # 对于从 CATIA 实时读取的数据，所有实例共享同一 dict（_mass_cache），
                # Step 2 已完成更新，此处跳过以避免重复放大惯量。
                if mp is not mp_shared:
                    mp["weight"] = new_weight_stored
                    if new_density_stored is not None and new_density_stored >= 0:
                        mp["density"] = new_density_stored
                    if scale != 1.0:
                        orig_i = mp.get("inertia", [[0.0] * 3 for _ in range(3)])
                        mp["inertia"] = [[orig_i[ir][ic] * scale for ic in range(3)]
                                         for ir in range(3)]
                if scale != 1.0:
                    # 更新行级惯量显示字段（零件自身坐标系）
                    I_local_new = mp.get("inertia", [[0.0] * 3 for _ in range(3)])
                    for ic_name, (ir2, ic2) in _INERTIA_IDX.items():
                        r[ic_name] = I_local_new[ir2][ic2]
                    # 同步更新 _root_mp 中的惯量（缩放后重新旋转到根坐标系）
                    I_root = _row_inertia_to_root(r)
                    rmp = r.get("_root_mp")
                    if rmp is not None:
                        rmp["inertia"] = I_root
                        rmp["weight"]  = new_weight_stored
                else:
                    rmp = r.get("_root_mp")
                    if rmp is not None:
                        rmp["weight"] = new_weight_stored
            else:
                # 无 _mass_props（各实例行独立存储惯量值），逐行缩放显示字段。
                if scale != 1.0:
                    for ic_name in _INERTIA_IDX:
                        cur = r.get(ic_name)
                        if cur is not None:
                            r[ic_name] = float(cur) * scale
                rmp = r.get("_root_mp")
                if rmp is not None:
                    rmp["weight"] = new_weight_stored

        # ── 更新可见树节点中同 PN 的所有行 ────────────────────────────────
        self._is_updating = True
        w_idx = self._columns.index("Weight") if "Weight" in self._columns else -1
        d_idx = self._columns.index("Density") if "Density" in self._columns else -1
        try:
            for vis_item in self._pn_to_items.get(pn, []):
                vis_row_idx = vis_item.data(0, _ROW_IDX_ROLE)
                if vis_row_idx is None:
                    continue
                vis_row = self._rows[vis_row_idx]
                # 对称件虚拟行的显示刷新由 _sync_all_mirrors() 负责
                if vis_row.get("Type") != BomNodeType.PART or vis_row.get("_is_mirror"):
                    continue
                if w_idx >= 0:
                    vis_item.setText(w_idx, self._fmt_mass_val(vis_row.get("Weight")))
                if d_idx >= 0:
                    d_val = vis_row.get("Density")
                    if d_val is not None and d_val >= 0:
                        vis_item.setText(d_idx, _fmt(d_val))
                for ic_name, (ir, ic) in _INERTIA_IDX.items():
                    if ic_name in self._columns:
                        ic_idx = self._columns.index(ic_name)
                        if self._summarize:
                            # 汇总BOM：显示零件自身坐标系值
                            raw_i = vis_row.get(ic_name)
                        else:
                            # 层级BOM：显示根产品坐标系值
                            rmp = vis_row.get("_root_mp")
                            raw_i = (
                                rmp["inertia"][ir][ic]
                                if rmp and rmp.get("inertia")
                                else vis_row.get(ic_name)
                            )
                        if raw_i is not None:
                            vis_item.setText(ic_idx, self._fmt_inertia_val(raw_i))
        finally:
            self._is_updating = False
        # 编辑密度/重量后立即重新计算汇总结果（无需手动点击"计算"）
        # 对称件同步在 _calculate() 内部进行（_sync_all_mirrors），支持任意 Type 原件
        self._calculate()
        # 标记为有未保存的编辑
        self._is_dirty = True
        self._update_title()

    # ── 计算 ───────────────────────────────────────────────────────────────

    def _calculate(self) -> None:
        if not self._rows:
            return
        # 自底向上联合汇总并同步对称件（详见 _sync_all_mirrors 文档）：
        # 深层产品/部件先内联重算后同步其对称件，父层再汇总时即可读到最新子层数据，
        # 一劳永逸地解决任意嵌套深度下产品对称件 _root_mp 陈旧的问题。
        self._sync_all_mirrors()
        # 最终全量汇总：将所有已更新对称件的贡献纳入各产品/部件显示字段
        recompute_product_rows(self._rows)
        self._refresh_product_items()
        try:
            result = rollup_mass_properties(self._rows)
        except Exception as e:
            logger.error(f"质量特性计算失败: {e}")
            QMessageBox.critical(self, "计算失败", f"计算总质量特性时出错：\n{e}")
            return
        self._rollup_result = result
        self._update_summary_labels(result)

    def _refresh_product_items(self) -> None:
        """刷新树形表格中所有产品/部件行的显示值（仅完整 BOM 模式有效）。

        在 _calculate() 调用 recompute_product_rows() 更新 self._rows 后，
        调用本方法将新的汇总值写回对应的 QTreeWidgetItem，以保持表格与数据同步。
        汇总 BOM 不含产品/部件行，故直接返回。
        """
        if self._summarize:
            return
        self._is_updating = True
        try:
            for item in self._item_by_row:
                row_idx = item.data(0, _ROW_IDX_ROLE)
                if row_idx is None:
                    continue
                row_data = self._rows[row_idx]
                if row_data.get("Type") not in BomNodeType.ASSEMBLY_TYPES:
                    continue
                for col_idx, col_name in enumerate(self._columns):
                    if col_name == "Weight":
                        raw = row_data.get("Weight")
                        item.setText(col_idx, self._fmt_mass_val(raw) if raw is not None else "")
                    elif col_name in ("CogX", "CogY", "CogZ"):
                        raw = row_data.get(col_name)
                        item.setText(col_idx, self._fmt_cog_val(raw) if raw is not None else "")
                    elif col_name in _INERTIA_IDX:
                        raw = row_data.get(col_name)
                        item.setText(col_idx, self._fmt_inertia_val(raw) if raw is not None else "")
        finally:
            self._is_updating = False

    def _clear_summary_labels(self) -> None:
        for attr in (
            "_edit_weight",
            "_edit_cx", "_edit_cy", "_edit_cz",
            "_edit_ixx", "_edit_ixy", "_edit_ixz",
            "_edit_iyx", "_edit_iyy", "_edit_iyz",
            "_edit_izx", "_edit_izy", "_edit_izz",
            "_edit_m1", "_edit_m2", "_edit_m3",
            "_edit_a1x", "_edit_a1y", "_edit_a1z",
            "_edit_a2x", "_edit_a2y", "_edit_a2z",
            "_edit_a3x", "_edit_a3y", "_edit_a3z",
        ):
            getattr(self, attr).setText("—")

    def _update_summary_labels(self, result: dict) -> None:
        unit_lbl     = self._mass_unit
        inertia_unit = self._inertia_unit
        cog_unit     = self._cog_unit

        w_val = result.get("total_weight", 0.0)
        self._edit_weight.setText(f"{self._fmt_mass_val(w_val)} {unit_lbl}")

        cog = result.get("cog", [0.0, 0.0, 0.0])
        self._edit_cx.setText(f"{self._fmt_cog_val(cog[0])} {cog_unit}")
        self._edit_cy.setText(f"{self._fmt_cog_val(cog[1])} {cog_unit}")
        self._edit_cz.setText(f"{self._fmt_cog_val(cog[2])} {cog_unit}")

        I = result.get("inertia", [[0.0] * 3 for _ in range(3)])
        self._edit_ixx.setText(f"{self._fmt_inertia_val(I[0][0])} {inertia_unit}")
        self._edit_ixy.setText(f"{self._fmt_inertia_val(I[0][1])} {inertia_unit}")
        self._edit_ixz.setText(f"{self._fmt_inertia_val(I[0][2])} {inertia_unit}")
        self._edit_iyx.setText(f"{self._fmt_inertia_val(I[1][0])} {inertia_unit}")
        self._edit_iyy.setText(f"{self._fmt_inertia_val(I[1][1])} {inertia_unit}")
        self._edit_iyz.setText(f"{self._fmt_inertia_val(I[1][2])} {inertia_unit}")
        self._edit_izx.setText(f"{self._fmt_inertia_val(I[2][0])} {inertia_unit}")
        self._edit_izy.setText(f"{self._fmt_inertia_val(I[2][1])} {inertia_unit}")
        self._edit_izz.setText(f"{self._fmt_inertia_val(I[2][2])} {inertia_unit}")

        pm = result.get("principal_moments", [0.0, 0.0, 0.0])
        self._edit_m1.setText(f"{self._fmt_inertia_val(pm[0])} {inertia_unit}")
        self._edit_m2.setText(f"{self._fmt_inertia_val(pm[1])} {inertia_unit}")
        self._edit_m3.setText(f"{self._fmt_inertia_val(pm[2])} {inertia_unit}")

        # principal_axes[row][col]: col=主轴序号(0=A1,1=A2,2=A3), row=分量(0=x,1=y,2=z)
        pa = result.get("principal_axes", [[1.0, 0.0, 0.0], [0.0, 1.0, 0.0], [0.0, 0.0, 1.0]])
        _fmt6 = lambda v: f"{v:.6g}"
        self._edit_a1x.setText(_fmt6(pa[0][0]))
        self._edit_a2x.setText(_fmt6(pa[0][1]))
        self._edit_a3x.setText(_fmt6(pa[0][2]))
        self._edit_a1y.setText(_fmt6(pa[1][0]))
        self._edit_a2y.setText(_fmt6(pa[1][1]))
        self._edit_a3y.setText(_fmt6(pa[1][2]))
        self._edit_a1z.setText(_fmt6(pa[2][0]))
        self._edit_a2z.setText(_fmt6(pa[2][1]))
        self._edit_a3z.setText(_fmt6(pa[2][2]))

    # ── 导出 ───────────────────────────────────────────────────────────────

    def _export_table(self) -> None:
        if not self._rows:
            QMessageBox.warning(self, "无数据", "请先加载产品树数据。")
            return

        # ── 默认文件名：根产品零件编号 + "_惯量汇总"（与"保存数据"对话框一致）──
        root_pn = str(self._rows[0].get("Part Number", "")).strip()
        default_stem = f"{root_pn}_惯量汇总" if root_pn else "惯量汇总"

        # ── 默认目录：上次浏览目录 → 根产品文件所在目录 → 空（与"保存数据"对话框一致）──
        if self._last_browse_dir and Path(self._last_browse_dir).is_dir():
            default_dir = self._last_browse_dir
        else:
            root_fp = str(self._rows[0].get("_filepath", "")).strip()
            default_dir = str(Path(root_fp).parent) if root_fp else ""

        default_path = str(Path(default_dir) / f"{default_stem}.xlsx") if default_dir else f"{default_stem}.xlsx"

        dest, _ = QFileDialog.getSaveFileName(
            self, "导出质量特性表格",
            default_path,
            "Excel 文件 (*.xlsx);;CSV 文件 (*.csv)",
        )
        if not dest:
            return

        dest_path = Path(dest)
        suffix = dest_path.suffix.lower()
        if suffix not in (".xlsx", ".csv"):
            dest_path = dest_path.with_suffix(".xlsx")
            suffix = ".xlsx"

        try:
            if suffix == ".csv":
                self._do_export_csv(dest_path)
            else:
                self._do_export(str(dest_path))
            self._last_browse_dir = str(dest_path.parent)
            self._settings.setValue("last_browse_dir", self._last_browse_dir)
            self._show_export_success(dest_path)
        except Exception as e:
            logger.error(f"导出失败: {e}")
            QMessageBox.critical(self, "导出失败", f"导出时出错：\n{e}")

    def _show_export_success(self, dest_path: Path) -> None:
        """导出成功后弹出含"打开文件"和"打开所在文件夹"按钮的提示框。"""
        msg = QMessageBox(self)
        msg.setWindowTitle("导出成功")
        msg.setText(f"文件已成功导出：\n{dest_path}")
        msg.setIcon(QMessageBox.Icon.Information)
        open_file_btn   = msg.addButton("打开文件", QMessageBox.ButtonRole.ActionRole)
        open_folder_btn = msg.addButton("打开所在文件夹", QMessageBox.ButtonRole.ActionRole)
        msg.addButton(QMessageBox.StandardButton.Ok)
        msg.exec()
        clicked = msg.clickedButton()
        if clicked == open_file_btn:
            try:
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(dest_path)))
            except Exception as exc:
                logger.warning(f"Failed to open exported file: {exc}")
        elif clicked == open_folder_btn:
            self._open_path(str(dest_path))

    @staticmethod
    def _row_status(row_data: dict) -> str:
        """Return a pipe-separated status string for a row.

        Possible tokens: mirror / excluded / no_file / not_found / unreadable / meas_failed.
        Empty string means the row has no special state.
        """
        tokens = []
        if row_data.get("_is_mirror"):   tokens.append("mirror")
        if row_data.get("_excluded"):    tokens.append("excluded")
        if row_data.get("_no_file"):     tokens.append("no_file")
        if row_data.get("_not_found"):   tokens.append("not_found")
        if row_data.get("_unreadable"):  tokens.append("unreadable")
        if row_data.get("_meas_failed"): tokens.append("meas_failed")
        return " | ".join(tokens)

    def _do_export(self, dest: str) -> None:

        # 导出始终使用层级BOM（含产品/部件/对称件），排除内部序号列 "#"，末尾追加 Status 列
        export_cols = [c for c in self._build_hierarchy_columns() if c != "#"] + ["Status"]

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "质量特性"

        center      = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        thin_side   = Side(style="thin")
        thin_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side,
        )

        # 特殊行背景色
        excl_fill     = PatternFill(fill_type="solid", fgColor="D8D8E8")  # 排除：灰紫
        error_fill    = PatternFill(fill_type="solid", fgColor="FFB3B3")  # not_found/unreadable：红
        warning_fill  = PatternFill(fill_type="solid", fgColor="FFF2CC")  # no_file/meas_failed：黄

        # 写入表头
        for ci, col_name in enumerate(export_cols, start=1):
            cell = ws.cell(row=1, column=ci, value=self._column_header(col_name))
            cell.font   = Font(bold=True)
            cell.fill   = header_fill
            cell.border = thin_border

        # 写入数据行（始终导出层级BOM，含所有特殊状态行）
        display_rows = self._get_hierarchy_rows()
        for ri, row_data in enumerate(display_rows, start=2):
            status_val = self._row_status(row_data)

            # 确定行背景色
            if row_data.get("_excluded"):
                row_fill = excl_fill
            elif row_data.get("_not_found") or row_data.get("_unreadable"):
                row_fill = error_fill
            elif row_data.get("_no_file") or row_data.get("_meas_failed"):
                row_fill = warning_fill
            else:
                row_fill = None

            for ci, col_name in enumerate(export_cols, start=1):
                if col_name == "Status":
                    value = status_val
                else:
                    raw = row_data.get(col_name)
                    if raw is None:
                        value = ""
                    elif col_name == "Density":
                        if raw < 0:
                            value = "不统一"
                        else:
                            try:
                                value = float(raw)
                            except (TypeError, ValueError):
                                value = ""
                    elif col_name == "Weight":
                        try:
                            value = float(raw) * self._unit_factor + 0.0
                        except (TypeError, ValueError):
                            value = ""
                    elif col_name in _INERTIA_IDX:
                        try:
                            value = float(raw) * self._inertia_unit_factor + 0.0
                        except (TypeError, ValueError):
                            value = ""
                    elif col_name in ("CogX", "CogY", "CogZ"):
                        try:
                            value = float(raw) * self._cog_unit_factor + 0.0
                        except (TypeError, ValueError):
                            value = ""
                    else:
                        value = raw
                cell = ws.cell(row=ri, column=ci, value=value)
                cell.border = thin_border
                if col_name == "Level":
                    cell.alignment = center
                if row_fill is not None:
                    cell.fill = row_fill

        # 汇总行（若已计算）
        if self._rollup_result:
            summary_row_idx = len(display_rows) + 2
            cog = self._rollup_result.get("cog", [0.0, 0.0, 0.0])
            I   = self._rollup_result.get("inertia", [[0.0] * 3 for _ in range(3)])
            w   = self._rollup_result.get("total_weight", 0.0)
            summary = {
                "Part Number":  "总计 (根产品)",
                "Weight":       w * self._unit_factor,
                "CogX":         cog[0] * self._cog_unit_factor,
                "CogY":         cog[1] * self._cog_unit_factor,
                "CogZ":         cog[2] * self._cog_unit_factor,
                "Ixx":          I[0][0] * self._inertia_unit_factor,
                "Iyy":          I[1][1] * self._inertia_unit_factor,
                "Izz":          I[2][2] * self._inertia_unit_factor,
                "Ixy":          I[0][1] * self._inertia_unit_factor,
                "Ixz":          I[0][2] * self._inertia_unit_factor,
                "Iyz":          I[1][2] * self._inertia_unit_factor,
            }
            summary_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
            for ci, col_name in enumerate(export_cols, start=1):
                val  = summary.get(col_name, "")
                cell = ws.cell(row=summary_row_idx, column=ci, value=val)
                cell.font   = Font(bold=True)
                cell.fill   = summary_fill
                cell.border = thin_border

        ws.freeze_panes = "A2"

        # 自适应列宽
        for ci, col_name in enumerate(export_cols, start=1):
            col_letter = ws.cell(row=1, column=ci).column_letter
            header     = self._column_header(col_name)
            max_width  = max(estimate_column_width(header), 8)
            for row_i in range(2, ws.max_row + 1):
                cv = ws.cell(row=row_i, column=ci).value
                if cv is not None:
                    max_width = max(max_width, estimate_column_width(str(cv)))
            ws.column_dimensions[col_letter].width = max_width + 2

        wb.save(dest)

    def _do_export_csv(self, dest: Path) -> None:
        """将当前表格数据（含汇总行）写入 UTF-8 with BOM 的 CSV 文件。"""

        # 导出始终使用层级BOM（含产品/部件/对称件），排除内部序号列 "#"，末尾追加 Status 列
        export_cols = [c for c in self._build_hierarchy_columns() if c != "#"] + ["Status"]
        display_rows = self._get_hierarchy_rows()

        def _cell_value(col_name: str, row_data: dict) -> str:
            if col_name == "Status":
                return self._row_status(row_data)
            raw = row_data.get(col_name)
            if raw is None:
                return ""
            if col_name == "Density":
                if isinstance(raw, (int, float)) and raw < 0:
                    return "不统一"
                try:
                    return str(float(raw))
                except (TypeError, ValueError):
                    return ""
            if col_name == "Weight":
                try:
                    val = float(raw) * self._unit_factor
                    return str(val + 0.0)  # normalize IEEE 754 -0.0 → +0.0
                except (TypeError, ValueError):
                    return ""
            if col_name in _INERTIA_IDX:
                try:
                    val = float(raw) * self._inertia_unit_factor
                    return str(val + 0.0)
                except (TypeError, ValueError):
                    return ""
            if col_name in ("CogX", "CogY", "CogZ"):
                try:
                    val = float(raw) * self._cog_unit_factor
                    return str(val + 0.0)
                except (TypeError, ValueError):
                    return ""
            return str(raw)

        with open(dest, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([self._column_header(c) for c in export_cols])
            for row_data in display_rows:
                writer.writerow([_cell_value(c, row_data) for c in export_cols])
            if self._rollup_result:
                cog = self._rollup_result.get("cog", [0.0, 0.0, 0.0])
                I   = self._rollup_result.get("inertia", [[0.0] * 3 for _ in range(3)])
                w   = self._rollup_result.get("total_weight", 0.0)
                summary = {
                    "Part Number":  "总计 (根产品)",
                    "Weight":       str(w * self._unit_factor),
                    "CogX":         str(cog[0] * self._cog_unit_factor),
                    "CogY":         str(cog[1] * self._cog_unit_factor),
                    "CogZ":         str(cog[2] * self._cog_unit_factor),
                    "Ixx":          str(I[0][0] * self._inertia_unit_factor),
                    "Iyy":          str(I[1][1] * self._inertia_unit_factor),
                    "Izz":          str(I[2][2] * self._inertia_unit_factor),
                    "Ixy":          str(I[0][1] * self._inertia_unit_factor),
                    "Ixz":          str(I[0][2] * self._inertia_unit_factor),
                    "Iyz":          str(I[1][2] * self._inertia_unit_factor),
                }
                writer.writerow([summary.get(c, "") for c in export_cols])
        logger.info(f"质量特性表格已导出 (csv) -> {dest}")

    def _autofit_columns(self) -> None:
        min_width = 60
        for col_idx in range(len(self._columns)):
            self._table.resizeColumnToContents(col_idx)
            if self._table.columnWidth(col_idx) < min_width:
                self._table.setColumnWidth(col_idx, min_width)
        for col_idx, col_name in enumerate(self._columns):
            self._col_widths[col_name] = self._table.columnWidth(col_idx)

    def _on_section_resized(self, logical_index: int, _old: int, new_size: int) -> None:
        if logical_index < len(self._columns):
            self._col_widths[self._columns[logical_index]] = new_size

    # ── 行操作辅助（删除 / 排除） ──────────────────────────────────────────

    def _get_subtree_indices(self, row_idx: int) -> list[int]:
        """返回 self._rows 中以 row_idx 为根的子树行索引列表（含 row_idx 自身）。

        通过比较相邻行的 Level 字段确定子树范围：子孙行的 Level 严格大于根行的
        Level，遇到 Level ≤ 根行时停止。
        """
        level = int(self._rows[row_idx].get("Level", 0))
        indices = [row_idx]
        for j in range(row_idx + 1, len(self._rows)):
            if int(self._rows[j].get("Level", 0)) > level:
                indices.append(j)
            else:
                break
        return indices

    def _delete_rows_multi(self, root_idxs: list[int]) -> None:
        """删除多个根行（及各自全部子孙行），一次重建表格并刷新汇总。

        合并所有根行的子树索引后统一删除，避免因索引失效导致误删；
        级联删除所有关联对称件（_mirror_child_id / _mirror_id 机制）。
        """
        indices: set[int] = set()
        for ri in root_idxs:
            indices.update(self._get_subtree_indices(ri))

        # 级联删除：收集被删除行的 _mirror_child_id，找出关联的对称件行
        mirror_child_ids: set[str] = {
            self._rows[i]["_mirror_child_id"]
            for i in indices
            if self._rows[i].get("_mirror_child_id")
        }
        if mirror_child_ids:
            for i, row in enumerate(self._rows):
                if (
                    i not in indices
                    and row.get("_is_mirror")
                    and row.get("_mirror_id") in mirror_child_ids
                ):
                    indices.add(i)

        # 反向关联清理：若被删除的行中包含对称件行，将对应源行的
        # _mirror_child_id 清除，使源行可以再次添加对称件。
        deleted_mirror_ids: set[str] = {
            self._rows[i]["_mirror_id"]
            for i in indices
            if self._rows[i].get("_is_mirror") and self._rows[i].get("_mirror_id")
        }
        if deleted_mirror_ids:
            for i, row in enumerate(self._rows):
                if i not in indices and row.get("_mirror_child_id") in deleted_mirror_ids:
                    row.pop("_mirror_child_id", None)

        self._rows = [r for i, r in enumerate(self._rows) if i not in indices]

        if not self._rows:
            self._table.clear()
            self._item_by_row = []
            self._pn_to_items.clear()
            self._rollup_result = None
            self._clear_summary_labels()
            return

        self._rebuild_columns_and_table()
        self._calculate()

    def _toggle_excluded_multi(self, row_idxs: list[int]) -> None:
        """批量切换多个行（及各自子孙行）的"参与计算"状态。

        若所有选中行均已排除，则全部恢复参与；否则（全部参与或混合状态）
        则将全部行标记为排除，使多选操作结果可预期且一致。
        切换完成后局部更新 QTreeWidgetItem 的视觉样式，并立即重新计算。
        """
        all_excluded = all(bool(self._rows[ri].get("_excluded", False)) for ri in row_idxs)
        new_val = not all_excluded  # 全部排除→恢复参与(False)，否则→排除(True)

        all_indices: set[int] = set()
        for ri in row_idxs:
            all_indices.update(self._get_subtree_indices(ri))
        for i in all_indices:
            self._rows[i]["_excluded"] = new_val
        self._apply_excluded_style_for_indices(all_indices, new_val)
        self._calculate()

    # ── 对称件 ────────────────────────────────────────────────────────────

    def _make_mirror_row(self, row_idx: int) -> dict:
        """根据 row_idx 对应的源行，生成相对 ZX 平面对称的虚拟行字典。

        对称规则（ZX 平面对称 → Y 轴分量取反）：
        - 重心：CogX/Z 不变，CogY → -CogY（根坐标系）
        - 转动惯量：Ixx/Iyy/Izz/Ixz 不变，Ixy → -Ixy，Iyz → -Iyz（根坐标系）

        对称件行的 _placement 设为单位矩阵、_mass_props.cog 设为根坐标系下的
        镜像重心，使 rollup_mass_properties() 可直接累加其贡献。
        """
        source_row = self._rows[row_idx]
        node_type  = str(source_row.get("Type", BomNodeType.PART))

        # ── 获取根坐标系质量特性 ─────────────────────────────────────────
        if node_type == BomNodeType.PART:
            rmp = source_row.get("_root_mp")
            mp  = source_row.get("_mass_props")
            if rmp:
                weight    = float(rmp.get("weight", 0.0))
                cog_root  = list(rmp.get("cog", [0.0, 0.0, 0.0]))
                I_root    = [list(r) for r in rmp.get("inertia", [[0.0] * 3 for _ in range(3)])]
            elif mp:
                weight    = float(mp.get("weight", 0.0))
                cog_root  = list(mp.get("cog", [0.0, 0.0, 0.0]))
                I_root    = [list(r) for r in mp.get("inertia", [[0.0] * 3 for _ in range(3)])]
            else:
                weight    = float(source_row.get("Weight") or 0.0)
                cog_root  = [0.0, 0.0, 0.0]
                I_root    = [[0.0] * 3 for _ in range(3)]
            density = mp.get("density") if mp else None
        else:
            # 产品/部件：使用已汇总的根坐标系显示字段
            weight   = float(source_row.get("Weight") or 0.0)
            cog_root = [
                float(source_row.get("CogX") or 0.0),
                float(source_row.get("CogY") or 0.0),
                float(source_row.get("CogZ") or 0.0),
            ]
            I_root = [
                [float(source_row.get("Ixx") or 0.0),
                 float(source_row.get("Ixy") or 0.0),
                 float(source_row.get("Ixz") or 0.0)],
                [float(source_row.get("Ixy") or 0.0),
                 float(source_row.get("Iyy") or 0.0),
                 float(source_row.get("Iyz") or 0.0)],
                [float(source_row.get("Ixz") or 0.0),
                 float(source_row.get("Iyz") or 0.0),
                 float(source_row.get("Izz") or 0.0)],
            ]
            density = None

        # ── 无有效数据检测 ────────────────────────────────────────────────
        # 当 weight <= 0.0 时视为"无测量数据"，COG/惯量全部置 None，
        # 保留原件的空值语义（零件显示"—"，产品/部件显示""）。
        has_data = weight > 0.0

        # ── ZX 平面对称变换：Y 分量取反 ──────────────────────────────────
        # 使用 (0.0 - x) 代替 (-x)，避免对 0.0 取负产生 IEEE 754 负零（-0.0），
        # 防止负零进入导出文件（CSV/xlsx）或引发下游比较歧义。
        if has_data:
            cog_mirror = [cog_root[0], 0.0 - cog_root[1], cog_root[2]]
            I_mirror = [
                [ I_root[0][0], 0.0 - I_root[0][1],  I_root[0][2]],
                [0.0 - I_root[1][0],  I_root[1][1], 0.0 - I_root[1][2]],
                [ I_root[2][0], 0.0 - I_root[2][1],  I_root[2][2]],
            ]
        else:
            cog_mirror = None
            I_mirror   = None

        # 单位矩阵 placement：对称件的"局部坐标系"等于根坐标系
        identity_placement = [
            [1.0, 0.0, 0.0, 0.0],
            [0.0, 1.0, 0.0, 0.0],
            [0.0, 0.0, 1.0, 0.0],
            [0.0, 0.0, 0.0, 1.0],
        ]
        mirror_mp = {
            "weight":  weight,
            "cog":     cog_mirror if cog_mirror is not None else [0.0, 0.0, 0.0],
            "inertia": I_mirror if I_mirror is not None else [[0.0] * 3 for _ in range(3)],
            "density": density,
        }

        source_pn            = str(source_row.get("Part Number", ""))
        source_instance_name = str(source_row.get("Instance Name", ""))
        return {
            "Level":         source_row.get("Level", 0),
            "Type":          BomNodeType.MIRROR,    # 虚拟叶节点：对称件以独立类型标识，
                                        # 直接贡献质量特性汇总（不参与层级汇总）
            "Part Number":   source_pn + " (对称件)",
            "Instance Name": (source_instance_name + " (对称件)") if source_instance_name else " (对称件)",
            "Filename":      "(虚拟)",
            "Nomenclature": source_row.get("Nomenclature", ""),
            "Revision":     source_row.get("Revision", ""),
            "Density":      density,
            "Weight":       weight if has_data else None,
            "CogX":         cog_mirror[0] if cog_mirror is not None else None,
            "CogY":         cog_mirror[1] if cog_mirror is not None else None,
            "CogZ":         cog_mirror[2] if cog_mirror is not None else None,
            "Ixx":          I_mirror[0][0] if I_mirror is not None else None,
            "Iyy":          I_mirror[1][1] if I_mirror is not None else None,
            "Izz":          I_mirror[2][2] if I_mirror is not None else None,
            "Ixy":          I_mirror[0][1] if I_mirror is not None else None,
            "Ixz":          I_mirror[0][2] if I_mirror is not None else None,
            "Iyz":          I_mirror[1][2] if I_mirror is not None else None,
            "_filepath":    "",
            "_placement":   identity_placement,
            "_not_found":   False,
            "_no_file":     False,
            "_unreadable":  False,
            "_meas_failed": False,
            "_mass_props":  mirror_mp,
            "_root_mp": {
                "weight":  weight,
                "cog":     cog_mirror,
                "inertia": I_mirror,
            } if has_data else None,
            "_is_mirror":        True,
            "_mirror_source_pn": source_pn,
            "_mirror_src_type":  node_type,  # 原件类型，用于 None 值的空白/破折号显示判断
        }

    def _add_mirror_row(self, row_idx: int) -> None:
        """在 row_idx 行的上方插入其 ZX 平面对称件虚拟行。

        使用 UUID 建立原件↔对称件双向关联，以便原件被删除时自动级联删除对称件。
        插入后重建表格并重新计算汇总。
        """
        mirror_id = uuid.uuid4().hex
        # 在源行上标记其对称件 ID（供 _delete_rows_multi 级联删除用）
        self._rows[row_idx]["_mirror_child_id"] = mirror_id

        mirror_row = self._make_mirror_row(row_idx)
        mirror_row["_mirror_id"] = mirror_id   # 对称件侧的匹配键

        # 插入到索引 row_idx，原件下移至 row_idx + 1，
        # 视觉上对称件显示在原件上方一行。
        self._rows.insert(row_idx, mirror_row)

        self._rebuild_columns_and_table()
        self._calculate()

    def _sync_all_mirrors(self) -> None:
        """自底向上联合重算产品/部件行并同步所有对称件虚拟行。

        **与旧版（自上而下）的区别**：
        旧版先执行全量 recompute_product_rows()，再从头到尾遍历源行同步对称件。
        当存在嵌套产品对称件时（例如 Sub-A 有对称件、Root 也有对称件），
        Root 的对称件会在 Sub-A 的对称件刷新之前就被计算，导致 Root 对称件
        的 _root_mp 滞后一轮，rollup_mass_properties() 读到陈旧值。

        **新版算法（自底向上）**：
        1. 一次 O(n) 扫描建立 mirror_id → 行索引映射。
        2. 收集所有带 _mirror_child_id 的源行，按 Level 降序排列（深层先处理）。
        3. 对每个源行：
           a. 若为产品/部件：先内联重算其子树汇总值（此时更深层的对称件已在
              前面的迭代中刷新，_root_mp 均为最新），再调用 _make_mirror_row()。
           b. 若为零件：_root_mp 由 _on_item_changed 负责维护，直接同步即可。
        4. 原地合并镜像数据，保留标识字段；刷新对应 QTreeWidgetItem。

        由此保证：无论对称件嵌套多少层，每个对称件都从完全最新的原件数据
        生成，rollup_mass_properties() 结果始终正确。
        """
        rows = self._rows
        n    = len(rows)

        # Step 1：建立 mirror_id → 行索引 的快速查找表
        mirror_idx_by_id: dict[str, int] = {
            r["_mirror_id"]: i
            for i, r in enumerate(rows)
            if r.get("_is_mirror") and r.get("_mirror_id")
        }
        if not mirror_idx_by_id:
            return

        # Step 2：收集所有源行，按 Level 降序（深层先）
        src_indices = sorted(
            [
                i for i, r in enumerate(rows)
                if r.get("_mirror_child_id") and not r.get("_is_mirror")
            ],
            key=lambda i: -int(rows[i].get("Level", 0)),
        )

        self._is_updating = True
        try:
            for src_idx in src_indices:
                src_row  = rows[src_idx]
                child_id = src_row.get("_mirror_child_id")
                if not child_id or child_id not in mirror_idx_by_id:
                    continue

                # Step 3a：若原件为产品/部件，先内联重算其子树汇总值
                # （此时更深层对称件的 _root_mp 已在前面迭代中刷新）
                if src_row.get("Type") in BomNodeType.ASSEMBLY_TYPES:
                    level = int(src_row.get("Level", 0))
                    child_parts: list[dict] = []
                    for j in range(src_idx + 1, n):
                        desc = rows[j]
                        if int(desc.get("Level", 0)) <= level:
                            break
                        if desc.get("_excluded"):
                            continue
                        rmp = desc.get("_root_mp")
                        if rmp and float(rmp.get("weight", 0.0)) > 0.0:
                            child_parts.append(rmp)
                    if child_parts:
                        result = _rollup_one_product(child_parts)
                        if result:
                            src_row["Weight"] = result["weight"]
                            src_row["CogX"]   = result["cog"][0]
                            src_row["CogY"]   = result["cog"][1]
                            src_row["CogZ"]   = result["cog"][2]
                            src_row["Ixx"]    = result["inertia"][0][0]
                            src_row["Iyy"]    = result["inertia"][1][1]
                            src_row["Izz"]    = result["inertia"][2][2]
                            src_row["Ixy"]    = result["inertia"][0][1]
                            src_row["Ixz"]    = result["inertia"][0][2]
                            src_row["Iyz"]    = result["inertia"][1][2]

                # Step 3b：重新计算镜像数据并原地合并
                mi         = mirror_idx_by_id[child_id]
                mirror_row = rows[mi]

                new_data = self._make_mirror_row(src_idx)
                new_data["_is_mirror"]        = True
                new_data["_mirror_id"]        = mirror_row["_mirror_id"]
                new_data["_mirror_source_pn"] = mirror_row.get("_mirror_source_pn")
                if mirror_row.get("_excluded"):
                    new_data["_excluded"] = True

                mirror_row.update(new_data)

                # Step 4：刷新该对称件的 QTreeWidgetItem
                for vis_item in self._item_by_row:
                    if vis_item.data(0, _ROW_IDX_ROLE) != mi:
                        continue
                    src_type = str(src_row.get("Type", BomNodeType.PART))
                    # 原件为产品/部件时，None 值显示空字符串；原件为零件时显示"—"
                    empty_text = "" if src_type in BomNodeType.ASSEMBLY_TYPES else "—"
                    for ci, col in enumerate(self._columns):
                        if col == "Weight":
                            raw = mirror_row.get("Weight")
                            if raw is None:
                                vis_item.setText(ci, empty_text)
                            else:
                                vis_item.setText(ci, self._fmt_mass_val(raw))
                        elif col == "Density":
                            d_val = mirror_row.get("Density")
                            if d_val is None:
                                d_text = empty_text
                            elif d_val < 0:
                                d_text = "不统一"
                            else:
                                d_text = _fmt(d_val)
                            vis_item.setText(ci, d_text)
                        elif col in _COG_IDX:
                            # mirror_row 中的 CogX/Y/Z 已在 _make_mirror_row() 中
                            # 保留了 None 语义（原件无数据时为 None）
                            raw = mirror_row.get(col)
                            if raw is None:
                                vis_item.setText(ci, empty_text)
                            else:
                                vis_item.setText(ci, self._fmt_cog_val(raw))
                        elif col in _INERTIA_IDX:
                            # 同上，Ixx/Ixy/… 亦保留 None 语义
                            raw = mirror_row.get(col)
                            if raw is None:
                                vis_item.setText(ci, empty_text)
                            else:
                                vis_item.setText(ci, self._fmt_inertia_val(raw))
                    break
        finally:
            self._is_updating = False


    def _apply_excluded_style_for_indices(
        self,
        indices: set[int],
        excluded: bool,
    ) -> None:
        """更新指定行索引集合对应 QTreeWidgetItem 的排除/恢复样式。

        excluded=True：浅灰紫背景 + 灰色斜体前景，并覆盖 tooltip。
        excluded=False：重置背景/前景/字体为系统默认，清空排除 tooltip。
        """
        # Qt 的 setBackground/setForeground/setFont/setToolTip 均会触发
        # itemChanged 信号；设置 _is_updating=True 防止 _on_item_changed
        # 把格式化后的显示字符串当作用户编辑写回，导致密度/重量被污染。
        self._is_updating = True
        try:
            self._apply_excluded_style_impl(indices, excluded)
        finally:
            self._is_updating = False

    def _apply_excluded_style_impl(
        self,
        indices: set[int],
        excluded: bool,
    ) -> None:
        c             = _get_colors(theme_manager.current_mode())
        default_brush = QBrush()
        default_font  = QFont()

        for item in self._item_by_row:
            r_idx = item.data(0, _ROW_IDX_ROLE)
            if r_idx not in indices:
                continue
            item.setData(0, _EXCLUDED_ROLE, excluded)
            for ci in range(len(self._columns)):
                if excluded:
                    item.setBackground(ci, c.EXCL_BG)
                    item.setForeground(ci, c.EXCL_FG)
                    item.setFont(ci, _EXCL_FONT)
                    item.setToolTip(ci, "该行已被排除，不参与计算。")
                else:
                    item.setBackground(ci, default_brush)
                    item.setForeground(ci, default_brush)
                    item.setFont(ci, default_font)
                    item.setToolTip(ci, "")

            # 恢复时重新应用该行原有的背景（异常/产品/正常等状态）
            if not excluded:
                self._apply_row_state_style(item, self._rows[r_idx])

    def _apply_row_state_style(self, item: QTreeWidgetItem, row_data: dict) -> None:
        """根据行状态设置 item 的背景色和 tooltip。

        覆盖以下状态：对称件（浅蓝）、锁定行（not_found/meas_failed/unreadable）、
        未保存（no_file）、产品/部件（淡灰）。
        不处理"排除"状态，该状态由 _apply_excluded_style_impl 单独管理。
        """
        not_found   = bool(row_data.get("_not_found"))
        meas_failed = bool(row_data.get("_meas_failed"))
        unreadable  = bool(row_data.get("_unreadable"))
        no_file     = bool(row_data.get("_no_file"))
        node_type   = str(row_data.get("Type", ""))
        is_mirror   = bool(row_data.get("_is_mirror"))
        row_locked  = unreadable or not_found or meas_failed or is_mirror
        col_count   = len(self._columns)
        c = _get_colors(theme_manager.current_mode())
        if is_mirror:
            for ci in range(col_count):
                item.setBackground(ci, c.MIRROR_BG)
                item.setToolTip(ci, _MIRROR_TOOLTIP)
        elif row_locked:
            if not_found:
                bg  = c.ROW_NOT_FOUND_BG
                tip = "该零件/产品的文件未被 CATIA 检索到，行内容不可编辑。"
            elif meas_failed:
                bg  = c.ROW_MEAS_FAILED_BG
                tip = "该零件的质量特性测量失败，行内容不可编辑。"
            else:
                bg  = c.ROW_LIGHTWEIGHT_BG
                tip = "该零件/产品处于轻量化模式，无法读取属性。"
            for ci in range(col_count):
                item.setForeground(ci, c.ROW_LOCKED_FG)
                item.setBackground(ci, bg)
                item.setToolTip(ci, tip)
        elif no_file:
            no_file_tip = "该零件尚未保存到磁盘，质量特性数据可能不完整。"
            for ci in range(col_count):
                item.setBackground(ci, c.ROW_UNSAVED_BG)
                item.setToolTip(ci, no_file_tip)
        elif node_type in BomNodeType.ASSEMBLY_TYPES:
            for ci in range(col_count):
                item.setBackground(ci, c.ROW_PRODUCT_BG)

    # ── 右键上下文菜单 ─────────────────────────────────────────────────────

    def _on_tree_context_menu(self, pos) -> None:
        """显示表格行的右键上下文菜单。支持多选批量操作。"""
        clicked_item = self._table.itemAt(pos)
        if clicked_item is None:
            return
        clicked_row_idx = clicked_item.data(0, _ROW_IDX_ROLE)
        if clicked_row_idx is None:
            return

        # 若右键点击的行不在当前选中集中，则清空选中并单独选中该行（单选时的默认行为）
        if not clicked_item.isSelected():
            self._table.clearSelection()
            clicked_item.setSelected(True)

        # 收集全部选中行的 row_idx
        selected_idxs: list[int] = []
        for sel_item in self._table.selectedItems():
            ri = sel_item.data(0, _ROW_IDX_ROLE)
            if ri is not None:
                selected_idxs.append(ri)
        if not selected_idxs:
            return

        is_single = len(selected_idxs) == 1

        # ── 右键点击行的单行属性（用于单选专属菜单项）────────────────────
        row_data     = self._rows[clicked_row_idx]
        fp           = str(row_data.get("_filepath", ""))
        fp_path      = Path(fp) if fp else None
        is_component = row_data.get("Type") == BomNodeType.COMPONENT
        is_part      = row_data.get("Type") == BomNodeType.PART
        is_product   = row_data.get("Type") == BomNodeType.PRODUCT
        not_found    = bool(row_data.get("_not_found"))
        no_file      = bool(row_data.get("_no_file"))
        unreadable   = bool(row_data.get("_unreadable"))
        is_mirror    = bool(row_data.get("_is_mirror"))

        menu = QMenu(self)

        # ── 打开路径（仅单选）────────────────────────────────────────────
        act_open_path = menu.addAction("打开路径")
        path_available = (
            is_single
            and not is_mirror
            and bool(fp) and not no_file and fp_path is not None
            and (fp_path.exists() or fp_path.parent.exists())
        )
        act_open_path.setEnabled(path_available)

        # ── 复制路径（仅单选）────────────────────────────────────────────
        act_copy_path = menu.addAction("复制路径")
        act_copy_path.setEnabled(is_single and not is_mirror and bool(fp) and not no_file)

        # ── 在CATIA中打开（仅单选）───────────────────────────────────────
        act_open_catia = menu.addAction("在 CATIA 中打开")
        catia_available = (
            is_single
            and not is_mirror
            and not is_component and not not_found and not unreadable
            and fp_path is not None and fp_path.exists()
        )
        act_open_catia.setEnabled(catia_available)

        menu.addSeparator()

        # ── 刷新质量特性 ──────────────────────────────────────────────────────
        # 完整 BOM：刷新选中节点及其子树内所有零件（含 mat4 重读）+ 子树外同 PN 兄弟实例。
        # 汇总 BOM：仅刷新当前行对应零件的质量特性（不涉及 mat4）。
        # 两种模式均通过 _product COM 引用直接测量，无需文件已保存到磁盘。
        act_refresh = menu.addAction(
            "刷新质量特性（子树范围）" if not self._summarize else "刷新质量特性"
        )
        refresh_available = (
            is_single
            and not is_mirror
            and self._rows[clicked_row_idx].get("_product") is not None
        )
        act_refresh.setEnabled(refresh_available)
        if not self._summarize:
            act_refresh.setToolTip(
                "通过 CATIA COM 引用直接重新测量选中节点及其子树内所有零件的质量特性。\n"
                "按当前面板选择的「Analyze」或「惯量包络体」方式执行。\n"
                "无需零件文件已保存到磁盘，适用于尚未保存的新建零件。"
            )
        else:
            act_refresh.setToolTip(
                "通过 CATIA COM 引用重新测量当前零件的质量特性。\n"
                "按当前面板选择的「Analyze」或「惯量包络体」方式执行。\n"
                "无需零件文件已保存到磁盘。"
            )

        # ── 层级BOM专属：增加对称件 / 参与计算 / 删除 ─────────────────────
        # act_toggle / act_delete / act_add_mirror 预置 None，以便在条件块外统一分发 action
        act_toggle     = None
        act_delete     = None
        act_add_mirror = None
        if not self._summarize:
            menu.addSeparator()
            # 参与计算：全部已排除显示"×"，全部参与显示"√"，混合状态显示"切换"
            all_excluded = all(bool(self._rows[ri].get("_excluded", False)) for ri in selected_idxs)
            any_excluded = any(bool(self._rows[ri].get("_excluded", False)) for ri in selected_idxs)
            if all_excluded:
                toggle_label = "参与计算：×"
            elif any_excluded:
                toggle_label = "参与计算：切换"
            else:
                toggle_label = "参与计算：√"
            act_toggle = menu.addAction(toggle_label)
            act_delete = menu.addAction("删除本行" if is_single else f"删除选中 {len(selected_idxs)} 行")

            menu.addSeparator()
            act_add_mirror = menu.addAction("增加对称件")
            # 仅对单选非对称件的零件/产品/部件行有效；对称件自身不可再次对称；
            # 同一行已有对称件（_mirror_child_id 已设置）时也不允许重复添加。
            already_has_mirror = bool(
                is_single
                and self._rows[clicked_row_idx].get("_mirror_child_id")
            )
            mirror_eligible = (
                is_single
                and (is_part or is_product or is_component)
                and not is_mirror
                and not already_has_mirror
            )
            act_add_mirror.setEnabled(mirror_eligible)

        action = menu.exec(self._table.viewport().mapToGlobal(pos))

        if action == act_open_path:
            self._open_path(fp)
        elif action == act_copy_path:
            QApplication.clipboard().setText(fp)
        elif action == act_open_catia:
            self._open_in_catia(fp)
        elif action == act_refresh:
            if self._summarize:
                self._refresh_mass_props_single(clicked_row_idx)
            else:
                self._refresh_mass_props_subtree(clicked_row_idx)
        elif act_toggle is not None and action == act_toggle:
            self._toggle_excluded_multi(selected_idxs)
        elif act_delete is not None and action == act_delete:
            if is_single:
                pn_label = str(row_data.get("Part Number", "") or row_data.get("Filename", ""))
                msg = f"确定要删除「{pn_label}」及其子节点吗？\n此操作不可撤销。"
            else:
                msg = f"确定要删除选中的 {len(selected_idxs)} 行（及各自的子节点）吗？\n此操作不可撤销。"
            confirm = QMessageBox.question(
                self,
                "确认删除",
                msg,
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
                QMessageBox.StandardButton.No,
            )
            if confirm == QMessageBox.StandardButton.Yes:
                self._delete_rows_multi(selected_idxs)
        elif act_add_mirror is not None and action == act_add_mirror:
            self._add_mirror_row(clicked_row_idx)

    def _open_path(self, fp: str) -> None:
        """在 Windows 资源管理器中打开包含 *fp* 的文件夹，并高亮选中该文件。

        使用 ShellExecuteW（宽字符 Unicode API）调用 explorer，避免经过
        cmd.exe / PowerShell 时中文路径因 OEM 代码页转换而乱码。
        """
        p = Path(fp).resolve()
        try:
            if p.exists():
                ctypes.windll.shell32.ShellExecuteW(
                    None, "open", "explorer.exe", f'/select,"{p}"', None, 1
                )
            elif p.parent.exists():
                ctypes.windll.shell32.ShellExecuteW(
                    None, "open", "explorer.exe", f'"{p.parent}"', None, 1
                )
        except Exception as exc:
            logger.warning(f"无法在资源管理器中打开路径: {exc}")

    def _open_in_catia(self, fp: str) -> None:
        """在 CATIA 中打开 *fp* 指向的文档，并将 CATIA V5 主窗口置于前台。"""
        try:
            open_document(fp, foreground=True)
        except Exception as e:
            QMessageBox.warning(self, "在 CATIA 中打开失败", f"无法在 CATIA 中打开文件：\n{e}")

    # ── 刷新质量特性 ─────────────────────────────────────────────────────────

    def _refresh_mass_props_single(self, row_idx: int) -> None:
        """汇总 BOM 模式下，仅刷新当前行对应零件的质量特性（不更新 mat4）。

        汇总 BOM 展示零件自身坐标系下的值，_root_mp 用 identity placement 计算。
        刷新成功后同步所有同 PN 的行（汇总模式下通常只有一行，但保持一致性）。
        """
        if row_idx >= len(self._rows):
            return
        r       = self._rows[row_idx]
        product = r.get("_product")
        pn      = str(r.get("Part Number", ""))

        if product is None:
            QMessageBox.warning(
                self, "无 COM 引用",
                f"零件「{pn}」没有有效的 COM 引用（可能来自载入文件），无法刷新。",
            )
            return

        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            if self._source == "analyze":
                new_mp = _measure_part_mass_props_analyze(product)
            else:
                part_doc_com = product.ReferenceProduct.Parent
                part_com     = part_doc_com.Part
                new_mp = _measure_part_mass_props(part_com, pn, self._read_mode)
        except Exception as e:
            new_mp = None
            logger.warning("汇总 BOM 刷新：零件 %s 测量失败: %s", pn, e)
        finally:
            QApplication.restoreOverrideCursor()

        if new_mp is None:
            QMessageBox.warning(self, "刷新失败", f"零件「{pn}」质量特性测量失败。")
            return

        # 汇总 BOM 用零件自身坐标系，_placement 为 identity（不影响坐标变换）
        self._apply_new_mp(r, new_mp, r.get("_placement"))

        # 同步所有同 PN 的行（_rows 中可能存在多个，保持数据一致）
        for i, other_r in enumerate(self._rows):
            if i == row_idx:
                continue
            if str(other_r.get("Part Number", "")) == pn and other_r.get("Type") == BomNodeType.PART:
                self._apply_new_mp(other_r, new_mp, other_r.get("_placement"))

        # 刷新可见树节点
        all_updated_rows = {
            i for i, other_r in enumerate(self._rows)
            if str(other_r.get("Part Number", "")) == pn
            and other_r.get("Type") == BomNodeType.PART
        }
        self._is_updating = True
        try:
            for vis_item in self._pn_to_items.get(pn, []):
                ri = vis_item.data(0, _ROW_IDX_ROLE)
                if ri is not None and ri in all_updated_rows:
                    self._refresh_part_item_after_reread(vis_item, self._rows[ri])
        finally:
            self._is_updating = False

        self._rollup_result = None
        self._clear_summary_labels()
        self._calculate()
        QMessageBox.information(self, "刷新完成", f"零件「{pn}」质量特性已刷新。")

    @staticmethod
    def _apply_new_mp(row: dict, new_mp: dict, placement) -> None:
        """将新的质量特性写入行 dict，并用 placement 重算根坐标系质量特性。

        参数：
            row:       要更新的行 dict（in-place 修改）。
            new_mp:    新的质量特性字典（内部单位）。
            placement: 该行的 4×4 变换矩阵（None 时跳过坐标变换）。
        """
        cog_local = new_mp["cog"]
        I_local   = new_mp["inertia"]
        row["_mass_props"]  = new_mp
        row["Density"]      = new_mp.get("density", None)
        row["Weight"]       = new_mp["weight"]
        row["CogX"]         = cog_local[0]
        row["CogY"]         = cog_local[1]
        row["CogZ"]         = cog_local[2]
        row["Ixx"]          = I_local[0][0]
        row["Iyy"]          = I_local[1][1]
        row["Izz"]          = I_local[2][2]
        row["Ixy"]          = I_local[0][1]
        row["Ixz"]          = I_local[0][2]
        row["Iyz"]          = I_local[1][2]
        row["_meas_failed"] = False
        if placement is not None:
            row["_root_mp"] = _compute_root_mp_from_placement(placement, new_mp)
        else:
            row["_root_mp"] = {
                "weight":  new_mp["weight"],
                "cog":     list(cog_local),
                "inertia": [list(row_i) for row_i in I_local],
            }

    def _refresh_mass_props_subtree(self, root_row_idx: int) -> None:
        """通过 COM 引用直接重新测量选中节点及其子树内所有零件的质量特性。

        按当前面板选择的「Analyze」或「惯量包络体」方式（及读取模式）执行。
        通过行 dict 中缓存的 ``_product`` COM 引用直接调用测量函数，
        无需零件文件已保存到磁盘（未保存零件仍在 CATIA 内存中时同样有效）。
        由完整 BOM 模式下的右键菜单调用。

        子树内实例：重测质量特性 + 重读 mat4（装配位置）+ 重算根坐标系质量特性。
        子树外同零件兄弟实例（同 PartNumber）：复用已测量的质量特性（无需重测）
            + 保留各自原有 mat4 + 重算根坐标系质量特性。

        注：mat4 仅对子树内实例重读；若用户在 CATIA 中移动了零件位置，
        应重新加载整个 BOM 而非使用此刷新功能。

        参数：
            root_row_idx: 右键点击行在 ``self._rows`` 中的索引。
        """
        if root_row_idx >= len(self._rows):
            return

        # ── 收集子树范围内的零件行索引 ────────────────────────────────────
        root_level = int(self._rows[root_row_idx].get("Level", 0))
        subtree_idxs: list[int] = []
        for i in range(root_row_idx, len(self._rows)):
            row = self._rows[i]
            if i > root_row_idx and int(row.get("Level", 0)) <= root_level:
                break  # 已超出子树范围
            if row.get("Type") == BomNodeType.PART and not row.get("_is_mirror"):
                subtree_idxs.append(i)

        if not subtree_idxs:
            QMessageBox.information(
                self, "无零件行",
                "选中节点及其子树内没有可刷新的零件行。",
            )
            return

        # ── 逐行重新测量：刷新质量特性 + mat4 ────────────────────────────
        # new_mp_by_pn: PartNumber → new_mp，供同 PN 其他实例复用（避免重复 COM 调用）
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        updated: list[int] = []
        failed_pns: list[str] = []
        new_mp_by_pn: dict[str, dict] = {}
        try:
            for ri in subtree_idxs:
                r       = self._rows[ri]
                product = r.get("_product")
                pn      = str(r.get("Part Number", ""))
                if product is None:
                    failed_pns.append(f"{pn}（无 COM 引用，可能来自载入文件）")
                    continue

                # 重读 mat4：取父节点的累积矩阵 × 本节点局部矩阵
                # 父节点的累积矩阵存在父行的 _placement 里；
                # 根节点（Level 0）无父，用单位矩阵。
                try:
                    local_mat4 = _position_to_mat4(product)
                    # 向上找最近已有 _placement 的祖先行作为父矩阵
                    parent_mat4 = _identity_4x4()
                    cur_level = int(r.get("Level", 0))
                    for j in range(ri - 1, -1, -1):
                        anc = self._rows[j]
                        if int(anc.get("Level", 0)) < cur_level:
                            p_mat = anc.get("_placement")
                            if p_mat is not None:
                                parent_mat4 = p_mat
                            break
                    new_abs_mat4 = _mat4_mul(parent_mat4, local_mat4)
                    r["_placement"] = new_abs_mat4
                except Exception as e:
                    logger.warning("子树刷新：零件 %s mat4 刷新失败: %s", pn, e)
                    new_abs_mat4 = r.get("_placement") or _identity_4x4()

                # 重测质量特性（同 PN 则复用，避免对同一文件重复 COM 调用）
                if pn and pn in new_mp_by_pn:
                    new_mp = new_mp_by_pn[pn]
                else:
                    try:
                        if self._source == "analyze":
                            new_mp = _measure_part_mass_props_analyze(product)
                        else:
                            part_doc_com = product.ReferenceProduct.Parent
                            part_com     = part_doc_com.Part
                            new_mp = _measure_part_mass_props(part_com, pn, self._read_mode)
                    except Exception as e:
                        logger.warning("子树刷新：零件 %s 测量失败: %s", pn, e)
                        new_mp = None

                if new_mp is None:
                    failed_pns.append(pn)
                    continue

                if pn:
                    new_mp_by_pn[pn] = new_mp

                # 写回 row dict
                self._apply_new_mp(r, new_mp, new_abs_mat4)
                updated.append(ri)
        finally:
            QApplication.restoreOverrideCursor()

        if not updated:
            QMessageBox.warning(
                self, "刷新失败",
                "子树内所有零件均未能重新测量。\n\n失败零件：\n"
                + "\n".join(f"  • {p}" for p in failed_pns[:10]),
            )
            return

        # ── 子树外同零件兄弟实例：复用 new_mp + 保留各自 mat4 + 重算 _root_mp ──
        subtree_set = set(subtree_idxs)
        sibling_updated: list[int] = []
        for i, r in enumerate(self._rows):
            if i in subtree_set:
                continue
            if r.get("Type") != BomNodeType.PART or r.get("_is_mirror"):
                continue
            pn_r = str(r.get("Part Number", ""))
            if not pn_r or pn_r not in new_mp_by_pn:
                continue
            self._apply_new_mp(r, new_mp_by_pn[pn_r], r.get("_placement"))  # 保留原 mat4
            sibling_updated.append(i)

        all_updated = set(updated) | set(sibling_updated)

        # ── 刷新可见树节点 ────────────────────────────────────────────────
        updated_pns: set[str] = {
            str(self._rows[ri].get("Part Number", "")) for ri in all_updated
        }
        self._is_updating = True
        try:
            for pn in updated_pns:
                for vis_item in self._pn_to_items.get(pn, []):
                    ri = vis_item.data(0, _ROW_IDX_ROLE)
                    if ri is not None and ri in all_updated:
                        vis_row = self._rows[ri]
                        if vis_row.get("Type") == BomNodeType.PART:
                            self._refresh_part_item_after_reread(vis_item, vis_row)
        finally:
            self._is_updating = False

        # ── 重新汇总并刷新底部结果 ─────────────────────────────────────────
        self._rollup_result = None
        self._clear_summary_labels()
        self._calculate()

        msg = f"已刷新 {len(updated)} 个子树内零件节点"
        if sibling_updated:
            msg += f"，另同步 {len(sibling_updated)} 个子树外同零件实例"
        msg += "。"
        if failed_pns:
            msg += "\n\n以下零件刷新失败：\n" + "\n".join(f"  • {p}" for p in failed_pns[:10])
            QMessageBox.warning(self, "部分刷新失败", msg)
        else:
            QMessageBox.information(self, "刷新完成", msg)

    def _refresh_part_item_after_reread(
        self,
        item: QTreeWidgetItem,
        row_data: dict,
    ) -> None:
        """重新读取质量特性成功后，更新零件行的视觉状态和显示值。

        清除之前因测量失败而设置的橙色背景和灰色文字，解除行锁定，
        并将新的质量特性数值写入各单元格。
        """
        default_brush = QBrush()  # 空画刷：传递给 setBackground/setForeground 时重置为系统默认样式

        # 恢复背景色、前景色和工具提示至默认状态
        for ci in range(len(self._columns)):
            item.setBackground(ci, default_brush)
            item.setForeground(ci, default_brush)
            item.setToolTip(ci, "")

        # 解除行锁定，允许编辑 Weight 列
        item.setData(0, _ITEM_LOCKED_ROLE, False)
        item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)

        # 更新密度列锁定状态
        density_val = row_data.get("Density")
        item.setData(0, _DENSITY_LOCKED_ROLE, (density_val is None) or (density_val < 0))

        # 更新各数值列的显示内容
        rmp = row_data.get("_root_mp")
        for col_idx, col_name in enumerate(self._columns):
            if col_name == "Density":
                d_val = row_data.get("Density")
                if d_val is None:
                    item.setText(col_idx, "—")
                elif d_val < 0:
                    item.setText(col_idx, "不统一")
                else:
                    item.setText(col_idx, _fmt(d_val))
            elif col_name == "Weight":
                item.setText(col_idx, self._fmt_mass_val(row_data.get("Weight")))
            elif col_name in ("CogX", "CogY", "CogZ"):
                if self._summarize:
                    # 汇总BOM：显示零件自身坐标系值
                    raw = row_data.get(col_name)
                else:
                    # 层级BOM：显示根产品坐标系值（来自 _root_mp）
                    cog_idx = ("CogX", "CogY", "CogZ").index(col_name)
                    raw = rmp["cog"][cog_idx] if rmp else row_data.get(col_name)
                item.setText(col_idx, self._fmt_cog_val(raw) if raw is not None else "—")
            elif col_name in _INERTIA_IDX:
                ir, ic = _INERTIA_IDX[col_name]
                if self._summarize:
                    raw = row_data.get(col_name)
                else:
                    raw = (
                        rmp["inertia"][ir][ic]
                        if rmp and rmp.get("inertia")
                        else row_data.get(col_name)
                    )
                item.setText(col_idx, self._fmt_inertia_val(raw) if raw is not None else "—")
