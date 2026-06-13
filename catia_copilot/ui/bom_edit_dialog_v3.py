"""
BOM 编辑对话框 V3 模块（part_master / instance 分离架构）。

提供：
- BomEditDialogV3 – 编辑即时写回版 BOM 表格（V3 架构）。
  每次单元格编辑后立即通过缓存 COM 引用写回 CATIA，
  无需点击"应用"或"完成"按钮批量提交。

  V3 相对 V2 的核心变化：
  - 用 _part_masters[pm_key] 替代 _canonical_data[inst_key]，
    PartMaster 级属性（PN/Nomenclature 等）只存一份，同文件多实例天然共享。
  - pm_key 永不变（基于 filepath/宿主），PN 改名只更新 pm["part_number"]。
  - 用 _pm_key_to_inst_keys 替代 _ref_to_insts，语义更清晰。
  - 删除 _inst_to_ref_unk 和 _sync_siblings_in_ui() 中的兄弟遍历，
    改用 _pm_key_to_inst_keys 直接覆盖所有同 pm_key 实例。
  - 撤销栈 key 类型区分：str → PartMaster 属性（pm_key），int → 实例属性（inst_key）。
"""

import csv
import ctypes
import logging
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from PySide6.QtWidgets import (
    QDialog, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QPushButton, QTreeWidgetItem, QHeaderView, QAbstractItemView,
    QComboBox, QCheckBox, QGroupBox, QMessageBox, QApplication,
    QFileDialog, QProgressDialog, QRadioButton, QButtonGroup, QSpinBox, QGridLayout,
    QMenu, QWidgetAction, QLineEdit,
)
from PySide6.QtGui import QPixmap, QKeySequence, QCloseEvent, QDesktopServices, QShortcut, QBrush, QPalette, QColor, QFont
from PySide6.QtCore import Qt, QSettings, QByteArray, QUrl

from catia_copilot.constants import (
    PRESET_USER_REF_PROPERTIES,
    PRESET_USER_REF_PROPERTY_OPTIONS,
    BOM_EDIT_COLUMN_ORDER,
    BOM_COLUMN_DISPLAY_NAMES,
    BOM_READONLY_COLUMNS,
    BOM_HIDEABLE_COLUMNS,
    BOM_ROW_NUMBER_COLUMN,
    BOM_INSTANCE_NAME_COLUMN,
    SOURCE_TO_DISPLAY,
    SOURCE_OPTIONS,
    PART_NUMBER_VALID_PATTERN,
    FILENAME_NOT_FOUND,
    FILENAME_UNSAVED,
    BOM_THUMBNAIL_MAX_SIZE,
    BomNodeType,
    TYPE_DISPLAY_NAMES,
)
from catia_copilot.catia.bom_collect import refresh_row_from_com
from catia_copilot.catia.bom_collect_v3 import (
    collect_bom_part_masters,
    iter_full_rows,
    iter_hierarchical_rows,
    flatten_bom_to_summary,
    get_part_master_attr,
    set_part_master_attr,
    rename_part_master,
)
from catia_copilot.catia.bom_write import write_bom_to_catia, write_cell
from catia_copilot.utils import read_catia_thumbnail
from catia_copilot.ui.bom_widgets import _BomTreeDelegate, _BomTreeWidget, _ITEM_LOCKED_ROLE, _BomSortItem
from catia_copilot.ui.bom_file_rename_dialog import _FileRenameDialog
from catia_copilot.ui.ui_colors import (
    MODIFIED_FG          as _MODIFIED_FG,
    ROW_LOCKED_FG, ROW_NOT_FOUND_BG, ROW_LIGHTWEIGHT_BG, ROW_UNSAVED_BG,
    get_colors as _get_colors,
)
from catia_copilot.ui.theme_manager import theme_manager, theme_signal
from catia_copilot.ui.ui_layout import L
from catia_copilot.catia.document import rename_document
from catia_copilot.catia.connection import open_document

logger = logging.getLogger(__name__)

_MAX_HISTORY = 10  # 撤销/重做最大步数


def _make_tree_combo(items: list[str]) -> QComboBox:
    """创建嵌入 QTreeWidget 行的 QComboBox。

    不使用 setStyleSheet，避免创建独立样式上下文导致下拉列表位置偏移。
    """
    combo = QComboBox()
    combo.setFocusPolicy(Qt.FocusPolicy.NoFocus)
    combo.setMaximumHeight(24)
    combo.addItems(items)
    return combo


class BomEditDialogV3(QDialog):
    """可编辑 BOM 表格（V3，part_master / instance 分离架构）。

    核心设计：
    - part_masters 是唯一真相源，不再依赖 full_rows / _canonical_data。
    - inst_info["product"] 存 COM 实例引用（防 GC），替代 _inst_to_product。
    - inst_info["instance_name"] 是实例名唯一真相，通过 _inst_key_to_info 快速访问。
    - 三个视图（完整/层级/汇总）从 part_masters 树临时生成，不缓存行列表。
    - 撤销栈 key 类型区分：str → PartMaster 属性（pn），int → 实例属性（inst_key）。
    - 每次单元格编辑后立即通过 inst_info["product"] 写回 CATIA（无批量提交）。
    """

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self.setWindowTitle("BOM 工作台 V3")
        self.setMinimumSize(900, 600)
        self.resize(1100, 700)
        self.setWindowFlags(
            self.windowFlags()
            | Qt.WindowType.WindowMaximizeButtonHint
            | Qt.WindowType.WindowMinimizeButtonHint
        )

        # ── 配置与持久化设置 ──────────────────────────────────────────────────
        # 与"导出BOM"对话框共享自定义列配置
        self._export_settings = QSettings("CATIACompanion", "ExportBOMDialog")
        self._last_browse_dir = self._export_settings.value("last_browse_dir", "")

        saved_custom = self._export_settings.value("custom_columns", [])
        if isinstance(saved_custom, str):
            saved_custom = [saved_custom]
        self._custom_columns: list[str] = list(saved_custom)

        # BomEditDialogV2 专用设置
        self._edit_settings  = QSettings("CATIACompanion", "BomEditDialogV2")
        saved_visible        = self._edit_settings.value("visible_preset_columns", [])
        if isinstance(saved_visible, str):
            saved_visible = [saved_visible]
        self._visible_preset_cols: list[str] = [
            c for c in saved_visible if c in PRESET_USER_REF_PROPERTIES
        ]

        # 可显示/隐藏的标准列（术语、版本、定义、来源）
        saved_hideable = self._edit_settings.value("visible_hideable_columns", BOM_HIDEABLE_COLUMNS)
        if isinstance(saved_hideable, str):
            saved_hideable = [saved_hideable]
        self._visible_hideable_cols: list[str] = [
            c for c in saved_hideable if c in BOM_HIDEABLE_COLUMNS
        ]

        self._full_bom: bool  = self._edit_settings.value("full_bom",  False, type=bool)
        self._summarize: bool = self._edit_settings.value("summarize", False, type=bool)
        # 完整 BOM 与汇总 BOM 互斥；若两者同时为 True（异常存档），以完整 BOM 优先
        if self._full_bom:
            self._summarize = False
        self._summary_include_assemblies: bool = self._edit_settings.value(
            "summary_include_assemblies", False, type=bool
        )
        # 包含所有预设的完整自定义列列表；从CATIA预读时覆盖所有列，不受当前可见性限制
        self._all_custom_columns: list[str] = list(dict.fromkeys(
            self._custom_columns + list(PRESET_USER_REF_PROPERTIES)
        ))

        self._show_filepath_col: bool = self._edit_settings.value(
            "show_filepath_column", False, type=bool,
        )
        self._show_filename_col: bool = self._edit_settings.value(
            "show_filename_column", True, type=bool,
        )
        # 完整 BOM 专有：实例名 / 实例描述列的可见性（其他 BOM 模式始终不显示）
        self._show_instance_name_col: bool = self._edit_settings.value(
            "show_instance_name_col", True, type=bool,
        )
        self._show_description_inst_col: bool = self._edit_settings.value(
            "show_description_inst_col", False, type=bool,
        )

        self._columns: list[str] = self._build_visible_columns()

        # ── 内部状态 ──────────────────────────────────────────────────────────
        # V3 核心数据结构：pm_key → part_master dict
        self._part_masters:      dict[str, dict] = {}
        # 根产品的 pm_key（== root PartNumber，根节点必为独立文件）
        self._root_pm_key:      str             = ""
        # inst_key → inst_info dict（O(1) 反向索引，仅包含非根节点实例）
        self._inst_key_to_info:  dict[int, dict] = {}
        # 按遍历顺序排列的显示行（由视图函数临时生成，_load_bom/_rebuild_rows 时重建）
        self._rows:             list[dict]      = []
        # 上次即时写回的状态文本
        self._last_write_status: str = ""
        # 防止变更处理回调重入的标志
        self._is_updating: bool = False
        # 与 self._rows 平行：self._item_by_row[i] 对应 self._rows[i] 的树形控件项
        self._item_by_row: list[QTreeWidgetItem] = []
        # BOM 成功加载至少一次后置为 True
        self._bom_loaded: bool = False
        # id(product) → 树形项列表，快速联动更新
        self._inst_to_items: dict = {}
        # pm_key → list[inst_key]，跨实例同步界面用
        self._pm_key_to_inst_keys: dict[str, list[int]] = {}
        # inst_key → product，补充 inst_key_to_info 里没有的节点（如根节点）
        self._inst_key_to_product: dict[int, object] = {}
        # 列名→像素宽度缓存；在列可见性切换时保留用户调整的列宽
        self._col_widths: dict[str, int] = {}
        # 撤销/重做历史栈（最多 _MAX_HISTORY 步）
        # key 为 str → PartMaster 属性（pm_key）；key 为 int → 实例属性（inst_key）
        self._undo_stack: list[list[tuple]] = []
        self._redo_stack: list[list[tuple]] = []
        # 当前搜索过滤文本（全小写）
        self._filter_text: str = ""

        # ── 界面布局 ──────────────────────────────────────────────────────────
        layout = QVBoxLayout(self)
        layout.setSpacing(10)
        layout.setContentsMargins(16, 16, 16, 16)

        # 数据来源选择行
        self._use_active_chk = QCheckBox("使用当前 CATIA 活动文档（无需手动选择文件）")
        self._use_active_chk.toggled.connect(self._toggle_file_row)
        layout.addWidget(self._use_active_chk)

        file_row = QHBoxLayout()
        self._file_edit       = QLineEdit()
        self._file_edit.setPlaceholderText("选择一个 CATProduct 文件...")
        self._file_edit.setReadOnly(True)
        self._file_browse_btn = QPushButton("浏览...")
        self._file_browse_btn.clicked.connect(self._browse_file)
        self._load_btn        = QPushButton("加载 BOM")
        self._load_btn.setToolTip("从文件或当前活动文档加载 BOM （F5）")
        self._load_btn.clicked.connect(self._load_bom)
        file_row.addWidget(self._file_edit)
        file_row.addWidget(self._file_browse_btn)
        file_row.addWidget(self._load_btn)
        layout.addLayout(file_row)

        # ── BOM类型与显示选项 ╳ 属性列（左右并排）────────────────────────────
        groups_row = QHBoxLayout()
        groups_row.setSpacing(8)

        # ── 左侧：BOM类型与显示选项（紧凑分组）──────────────────────────────
        display_group  = QGroupBox("BOM 类型与显示选项")
        display_layout = QVBoxLayout(display_group)
        display_layout.setSpacing(4)
        display_layout.setContentsMargins(8, 6, 8, 6)

        # 第一行：单选按钮 + 汇总选项
        bom_type_row = QHBoxLayout()
        self._bom_type_btn_group = QButtonGroup(self)
        self._radio_full_bom     = QRadioButton("完整 BOM")
        self._radio_hierarchical = QRadioButton("层级 BOM")
        self._radio_summary_bom  = QRadioButton("汇总 BOM")
        self._radio_full_bom.setToolTip(
            "显示完整产品树，每个实例单独一行，包含实例名列"
        )
        self._radio_full_bom.setMinimumHeight(24)
        if self._full_bom:
            self._radio_full_bom.setChecked(True)
        elif self._summarize:
            self._radio_summary_bom.setChecked(True)
        else:
            self._radio_hierarchical.setChecked(True)
        self._bom_type_btn_group.addButton(self._radio_full_bom)
        self._bom_type_btn_group.addButton(self._radio_hierarchical)
        self._bom_type_btn_group.addButton(self._radio_summary_bom)
        self._radio_full_bom.toggled.connect(self._on_full_bom_toggled)
        self._radio_hierarchical.toggled.connect(self._on_hierarchical_bom_toggled)
        self._radio_summary_bom.toggled.connect(self._on_summary_bom_toggled)
        bom_type_row.addWidget(self._radio_full_bom)
        bom_type_row.addWidget(self._radio_hierarchical)
        bom_type_row.addWidget(self._radio_summary_bom)

        self._summary_opts_widget = QWidget()
        summary_opts_layout = QHBoxLayout(self._summary_opts_widget)
        summary_opts_layout.setContentsMargins(0, 0, 0, 0)
        summary_opts_layout.setSpacing(8)

        self._include_assemblies_chk = QCheckBox("包含产品和部件")
        self._include_assemblies_chk.setToolTip(
            "勾选后，汇总 BOM 中也会列出产品和部件（子产品），而不仅限于零件。"
        )
        self._include_assemblies_chk.setChecked(self._summary_include_assemblies)
        self._include_assemblies_chk.toggled.connect(self._on_include_assemblies_toggled)
        summary_opts_layout.addWidget(self._include_assemblies_chk)
        self._summary_opts_widget.setVisible(self._summarize)
        bom_type_row.addWidget(self._summary_opts_widget)

        # 完整 BOM 专有选项：实例名 / 实例描述列可见性（其他模式隐藏）
        self._full_bom_opts_widget = QWidget()
        full_bom_opts_layout = QHBoxLayout(self._full_bom_opts_widget)
        full_bom_opts_layout.setContentsMargins(0, 0, 0, 0)
        full_bom_opts_layout.setSpacing(8)

        self._show_instance_name_chk = QCheckBox("实例名")
        self._show_instance_name_chk.setToolTip("完整 BOM 模式下显示实例名列（product.Name）")
        self._show_instance_name_chk.setChecked(self._show_instance_name_col)
        self._show_instance_name_chk.toggled.connect(self._on_full_bom_cols_toggled)
        full_bom_opts_layout.addWidget(self._show_instance_name_chk)

        self._show_description_inst_chk = QCheckBox("实例描述")
        self._show_description_inst_chk.setToolTip("完整 BOM 模式下显示实例描述列（product.DescriptionInst）")
        self._show_description_inst_chk.setChecked(self._show_description_inst_col)
        self._show_description_inst_chk.toggled.connect(self._on_full_bom_cols_toggled)
        full_bom_opts_layout.addWidget(self._show_description_inst_chk)

        self._full_bom_opts_widget.setVisible(self._full_bom)
        bom_type_row.addWidget(self._full_bom_opts_widget)

        bom_type_row.addStretch()
        display_layout.addLayout(bom_type_row)

        # 第二行：筛选框
        filter_row = QHBoxLayout()
        filter_row.setSpacing(6)
        filter_row.addWidget(QLabel("筛选:"))
        self._filter_edit = QLineEdit()
        self._filter_edit.setPlaceholderText("按零件编号、术语、文件名等关键字搜索行…")
        self._filter_edit.setClearButtonEnabled(True)
        self._filter_edit.textChanged.connect(self._on_filter_changed)
        filter_row.addWidget(self._filter_edit)
        display_layout.addLayout(filter_row)

        groups_row.addWidget(display_group, 0)

        # ── 右侧：属性列（勾选以显示）────────────────────────────────────────
        preset_group  = QGroupBox("属性列（勾选以显示）")
        preset_main_layout = QVBoxLayout(preset_group)
        preset_main_layout.setSpacing(4)
        preset_main_layout.setContentsMargins(8, 6, 8, 6)

        self._preset_checkboxes: dict[str, QCheckBox] = {}

        # 第一行：文件名 + 显示完整路径 + 可隐藏标准列
        row0 = QHBoxLayout()
        row0.setSpacing(12)

        fn_cb = QCheckBox(BOM_COLUMN_DISPLAY_NAMES.get("Filename", "Filename"))
        fn_cb.setChecked(self._show_filename_col)
        fn_cb.toggled.connect(self._on_preset_col_toggled)
        row0.addWidget(fn_cb)
        self._preset_checkboxes["Filename"] = fn_cb

        self._filepath_chk = QCheckBox("显示完整路径")
        self._filepath_chk.setToolTip("勾选后文件名列将显示文件完整路径（含目录），而非仅文件名")
        self._filepath_chk.setChecked(self._show_filepath_col)
        self._filepath_chk.toggled.connect(self._on_show_filepath_toggled)
        row0.addWidget(self._filepath_chk)

        for col_name in BOM_HIDEABLE_COLUMNS:
            cb = QCheckBox(BOM_COLUMN_DISPLAY_NAMES.get(col_name, col_name))
            cb.setChecked(col_name in self._visible_hideable_cols)
            cb.toggled.connect(self._on_hideable_col_toggled)
            row0.addWidget(cb)
            self._preset_checkboxes[col_name] = cb

        row0.addStretch()
        preset_main_layout.addLayout(row0)

        # 第二行：预设用户自定义属性（物料编码、物料名称等）
        row1 = QHBoxLayout()
        row1.setSpacing(12)
        for col_name in PRESET_USER_REF_PROPERTIES:
            cb = QCheckBox(col_name)
            cb.setChecked(col_name in self._visible_preset_cols)
            cb.toggled.connect(self._on_preset_col_toggled)
            row1.addWidget(cb)
            self._preset_checkboxes[col_name] = cb
        row1.addStretch()
        preset_main_layout.addLayout(row1)

        groups_row.addWidget(preset_group, 1)

        layout.addLayout(groups_row)

        hint = QLabel(
            "层级 / 类型 / 数量 为结构属性，不可编辑，"
            "零件编号可编辑但不能与其他行冲突，"
            "文件名/路径可编辑。"
        )
        hint.setWordWrap(True)
        hint.setObjectName("hintLabel")
        layout.addWidget(hint)

        # BOM树形控件（替代 QTableWidget，原生支持展开/折叠）
        self._table = _BomTreeWidget()
        _init_headers = self._display_headers()
        self._table.setColumnCount(len(_init_headers))
        self._table.setHeaderLabels(_init_headers)
        hdr = self._table.header()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.Interactive)
        hdr.setStretchLastSection(True)
        hdr.setSectionsMovable(True)
        hdr.setFixedHeight(L.TABLE_ROW_HEIGHT)
        self._table.setUniformRowHeights(True)
        self._table.setRootIsDecorated(True)
        self._table.setSortingEnabled(False)
        self._table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._table.setSelectionMode(QAbstractItemView.SelectionMode.ExtendedSelection)
        self._table.setAlternatingRowColors(True)
        self._table.setIndentation(12)
        self._table.itemChanged.connect(self._on_item_changed)
        hdr.sectionResized.connect(self._on_section_resized)
        _delegate = _BomTreeDelegate(lambda: self._columns, self._table)
        self._table.setItemDelegate(_delegate)
        self._table.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        self._table.customContextMenuRequested.connect(self._on_tree_context_menu)
        layout.addWidget(self._table, 1)

        # 底部按钮行
        btn_row = QHBoxLayout()

        autofit_btn = QPushButton("自适应列宽")
        autofit_btn.setToolTip("根据内容自动调整所有列的宽度")
        autofit_btn.clicked.connect(self._autofit_columns)
        btn_row.addWidget(autofit_btn)

        expand_btn = QPushButton("全部展开")
        expand_btn.setToolTip("展开结构树中的所有节点")
        expand_btn.clicked.connect(self._table.expandAll)
        btn_row.addWidget(expand_btn)

        collapse_btn = QPushButton("全部折叠")
        collapse_btn.setToolTip("折叠结构树中的所有节点")
        collapse_btn.clicked.connect(self._table.collapseAll)
        btn_row.addWidget(collapse_btn)

        self._undo_btn = QPushButton("↶")
        self._undo_btn.setAccessibleName("撤销")
        self._undo_btn.setToolTip("撤销上一步字段编辑（Ctrl+Z）")
        self._undo_btn.setShortcut(QKeySequence("Ctrl+Z"))
        self._undo_btn.setEnabled(False)
        self._undo_btn.clicked.connect(self._undo)
        _undo_redo_font = QFont("Segoe UI Emoji")
        _undo_redo_font.setPointSize(11)
        self._undo_btn.setFont(_undo_redo_font)
        btn_row.addWidget(self._undo_btn)

        self._redo_btn = QPushButton("↷")
        self._redo_btn.setAccessibleName("重做")
        self._redo_btn.setToolTip("重做上一步撤销的编辑（Ctrl+Y）")
        self._redo_btn.setShortcut(QKeySequence("Ctrl+Y"))
        self._redo_btn.setEnabled(False)
        self._redo_btn.clicked.connect(self._redo)
        self._redo_btn.setFont(_undo_redo_font)
        btn_row.addWidget(self._redo_btn)

        # 状态标签
        btn_row.addSpacing(8)
        self._status_label = QLabel("")
        self._status_label.setObjectName("hintLabel")
        btn_row.addWidget(self._status_label)

        btn_row.addStretch()

        self._export_btn = QPushButton("导出表格")
        self._export_btn.setToolTip("将当前表格导出为 Excel （.xlsx）或 CSV 文件（Ctrl+E）")
        self._export_btn.setEnabled(False)
        self._export_btn.setShortcut(QKeySequence("Ctrl+E"))
        self._export_btn.clicked.connect(self._export_table)
        btn_row.addWidget(self._export_btn)

        self._finish_btn = QPushButton("关闭")
        self._finish_btn.setDefault(False)
        self._finish_btn.setEnabled(True)
        self._finish_btn.setToolTip("关闭对话框（所有修改已即时写回 CATIA ）")
        self._finish_btn.setShortcut(QKeySequence("Ctrl+Return"))
        self._finish_btn.clicked.connect(self.accept)

        btn_row.addWidget(self._finish_btn)
        layout.addLayout(btn_row)

        # ── 快捷键 ────────────────────────────────────────────────────────────
        QShortcut(QKeySequence("F5"), self, self._load_bom)
        QShortcut(QKeySequence("Ctrl+F"), self, self._focus_filter)

        # ── 恢复窗口几何（位置与尺寸）────────────────────────────────────────
        saved_geom = self._edit_settings.value("geometry")
        if isinstance(saved_geom, QByteArray) and not saved_geom.isEmpty():
            self.restoreGeometry(saved_geom)

        # ── 主题切换：重新着色所有行 ──────────────────────────────────────────
        theme_signal.theme_changed.connect(self._on_theme_changed)

        # ── 默认使用活动文档（所有控件构建完毕后再触发 toggled）────────────────
        self._use_active_chk.setChecked(True)

    # ── 文件/活动文档切换 ─────────────────────────────────────────────────────

    def _toggle_file_row(self, use_active: bool) -> None:
        self._file_edit.setEnabled(not use_active)
        self._file_browse_btn.setEnabled(not use_active)

    # ── BOM类型切换 ───────────────────────────────────────────────────────────

    def _on_summary_bom_toggled(self, checked: bool) -> None:
        """处理"汇总 BOM"单选按钮切换。"""
        if not checked:
            return
        self._full_bom  = False
        self._summarize = True
        self._edit_settings.setValue("full_bom", False)
        self._edit_settings.setValue("summarize", True)
        self._summary_opts_widget.setVisible(True)
        self._full_bom_opts_widget.setVisible(False)
        if self._bom_loaded:
            self._rebuild_rows()
            self._rebuild_columns_and_repopulate()

    def _on_full_bom_toggled(self, checked: bool) -> None:
        """处理"完整 BOM"单选按钮切换。"""
        if not checked:
            return
        self._full_bom  = True
        self._summarize = False
        self._edit_settings.setValue("full_bom", True)
        self._edit_settings.setValue("summarize", False)
        self._summary_opts_widget.setVisible(False)
        self._full_bom_opts_widget.setVisible(True)
        if self._bom_loaded:
            self._rebuild_rows()
            self._rebuild_columns_and_repopulate()

    def _on_hierarchical_bom_toggled(self, checked: bool) -> None:
        """处理"层级 BOM"单选按钮切换。"""
        if not checked:
            return
        self._full_bom  = False
        self._summarize = False
        self._edit_settings.setValue("full_bom", False)
        self._edit_settings.setValue("summarize", False)
        self._summary_opts_widget.setVisible(False)
        self._full_bom_opts_widget.setVisible(False)
        if self._bom_loaded:
            self._rebuild_rows()
            self._rebuild_columns_and_repopulate()

    def _on_include_assemblies_toggled(self, checked: bool) -> None:
        self._summary_include_assemblies = checked
        self._edit_settings.setValue("summary_include_assemblies", checked)
        if self._summarize and self._bom_loaded:
            self._rebuild_rows()
            self._rebuild_columns_and_repopulate()

    def _rebuild_rows(self) -> None:
        """根据当前模式从 part_masters 树重新生成 _rows。"""
        if not self._root_pm_key or not self._part_masters:
            return
        if self._full_bom:
            self._rows = iter_full_rows(self._root_pm_key, self._part_masters)
        elif self._summarize:
            hierarchical = iter_hierarchical_rows(self._root_pm_key, self._part_masters)
            summary      = flatten_bom_to_summary(
                hierarchical,
                include_assemblies=self._summary_include_assemblies,
                sort_column=None,
            )
            # flatten_bom_to_summary 只保留标准字段，丢失了 _inst_key / _product 等
            # V3 内部字段。从层级行按 pm_key 建索引，将内部字段回填到汇总行。
            _v3_fields = ("_inst_key", "_product", "_pm_key",
                          "_filepath", "_not_found", "_no_file", "_unreadable")
            pm_key_to_hier: dict[str, dict] = {}
            for hr in hierarchical:
                bk = str(hr.get("_pm_key", ""))
                if bk and bk not in pm_key_to_hier:
                    pm_key_to_hier[bk] = hr
            for row in summary:
                bk = str(row.get("Part Number", ""))   # 汇总行此时 _pm_key 尚未回填，用 PN 找
                # 先尝试精确 pm_key 匹配（独立文件节点 pm_key == pn）
                hier_row = pm_key_to_hier.get(bk)
                if hier_row is not None:
                    for f in _v3_fields:
                        if f in hier_row:
                            row[f] = hier_row[f]
            self._rows = summary
        else:
            self._rows = iter_hierarchical_rows(self._root_pm_key, self._part_masters)

    # ── 表格辅助方法 ──────────────────────────────────────────────────────────

    def _autofit_columns(self) -> None:
        """根据内容自动调整所有列宽，设有最小宽度下限。"""
        # QTreeWidget 使用 resizeColumnToContents(int)，而非 resizeColumnsToContents()
        min_width = 60
        for col_idx, col_name in enumerate(self._columns):
            self._table.resizeColumnToContents(col_idx)
            if self._table.columnWidth(col_idx) < min_width:
                self._table.setColumnWidth(col_idx, min_width)
            # 更新缓存，使后续列可见性切换能保留此列宽
            self._col_widths[col_name] = self._table.columnWidth(col_idx)

    def _on_section_resized(self, logical_index: int, _old_size: int, new_size: int) -> None:
        """用户或代码调整列宽时，将新宽度写入缓存。"""
        if logical_index < len(self._columns):
            self._col_widths[self._columns[logical_index]] = new_size

    def _rebuild_columns_and_repopulate(self) -> None:
        """重建可见列列表，更新表头，若已有行数据则刷新表格。"""
        # --- 不可变宽度快照（同时用于锚点计算和宽度恢复）---
        #
        # 在任何 Qt 树/表头操作之前先取快照。快照将持久缓存
        # （为此前隐藏的列提供列宽）与当前 columnWidth() 值
        # （对可见列具有权威性）合并，使锚点计算和宽度恢复都从
        # 稳定副本读取，而非从可能在重建过程中被修改的
        # self._col_widths 读取。
        width_snapshot: dict[str, int] = dict(self._col_widths)
        for col_idx, col_name in enumerate(self._columns):
            w = self._table.columnWidth(col_idx)
            width_snapshot[col_name] = w
            self._col_widths[col_name] = w  # keep persistent cache current

        # --- 锚定列水平滚动位置记录 ---
        # 目标：在添加/删除列后，保持视口最左侧可见列不变。
        # 跨列数变化的原始像素滚动值并不稳定，因为 Qt 可能重置滚动条，
        # 且删除视口左侧的列会使所有剩余可见列的像素位置发生偏移。
        #
        # 策略：在重建前确定视口左边缘所在的列名（及偏入该列的像素数），
        # 重建后再根据列宽重新计算目标水平滚动值。
        old_columns = list(self._columns)
        old_hscroll = self._table.horizontalScrollBar().value()

        anchor_col_name: str | None = None  # 视口左边缘所在列名
        anchor_offset: int = 0              # 偏入该列的像素数

        x = 0
        for col_name in old_columns:
            w = width_snapshot[col_name]    # 使用快照，而非实时 columnWidth
            if x + w > old_hscroll:
                anchor_col_name = col_name
                anchor_offset = old_hscroll - x
                break
            x += w

        vscroll = self._table.verticalScrollBar().value()

        self._columns = self._build_visible_columns()
        if self._rows:
            self._populate_table()  # 内部已设置列数和表头
            # 从不可变快照恢复列宽；对从未出现过的新列执行自适应宽度
            for col_idx, col_name in enumerate(self._columns):
                if col_name in width_snapshot:
                    self._table.setColumnWidth(col_idx, width_snapshot[col_name])
                else:
                    self._table.resizeColumnToContents(col_idx)

            # 根据锚定列的新像素位置计算新的水平滚动值
            new_hscroll = 0
            if anchor_col_name is not None:
                # 在新布局中查找锚定列
                x = 0
                found = False
                for col_idx, col_name in enumerate(self._columns):
                    if col_name == anchor_col_name:
                        new_hscroll = x + anchor_offset
                        found = True
                        break
                    x += self._table.columnWidth(col_idx)

                if not found:
                    # 锚定列已被隐藏；滚动到旧布局中锚定列右侧第一个
                    # 仍存在的列（即被删除列右边第一个幸存列）
                    old_col_order = {c: i for i, c in enumerate(old_columns)}
                    anchor_old_idx = old_col_order.get(anchor_col_name, -1)
                    x = 0
                    for col_idx, col_name in enumerate(self._columns):
                        if old_col_order.get(col_name, -1) > anchor_old_idx:
                            new_hscroll = x
                            break
                        x += self._table.columnWidth(col_idx)
                    else:
                        # 所有剩余列均在锚定列原位置左侧；滚动到末尾
                        new_hscroll = self._table.horizontalScrollBar().maximum()

            new_hscroll = max(0, min(new_hscroll, self._table.horizontalScrollBar().maximum()))
            self._table.verticalScrollBar().setValue(vscroll)
            self._table.horizontalScrollBar().setValue(new_hscroll)
        else:
            # 尚无行数据：仅更新列数和表头，以便在加载BOM前也能反映最新列选择
            _headers = self._display_headers()
            self._table.setColumnCount(len(_headers))
            self._table.setHeaderLabels(_headers)

    # ── 列可见性管理 ──────────────────────────────────────────────────────────

    def _display_headers(self) -> list[str]:
        """返回当前列列表的显示表头标签。

        当"显示完整路径"选项激活时，文件名列的表头显示为"完整路径"，
        以便用户直观区分。
        """
        result = []
        for c in self._columns:
            if c == "Filename" and self._show_filepath_col:
                result.append("完整路径")
            else:
                result.append(BOM_COLUMN_DISPLAY_NAMES.get(c, c))
        return result

    def _build_visible_columns(self) -> list[str]:
        base = list(BOM_EDIT_COLUMN_ORDER)
        # 过滤隐藏列（文件名列和可隐藏列）
        if not self._show_filename_col:
            base = [c for c in base if c != "Filename"]
        # 过滤已隐藏的可隐藏列
        base = [c for c in base if c not in BOM_HIDEABLE_COLUMNS or c in self._visible_hideable_cols]
        if self._summarize:
            # 汇总模式下层级列无意义；不含产品时也隐藏类型列
            cols_to_hide = {"Level"}
            if not self._summary_include_assemblies:
                cols_to_hide.add("Type")
            base = [c for c in base if c not in cols_to_hide]
        visible_preset = [
            c for c in PRESET_USER_REF_PROPERTIES if c in self._visible_preset_cols
        ]
        other_custom   = [
            c for c in self._custom_columns
            if c not in BOM_EDIT_COLUMN_ORDER and c not in PRESET_USER_REF_PROPERTIES
        ]
        # 将"#"紧插在"Level"之后（逻辑索引0→Level，逻辑索引1→"#"），
        # 使 QTreeWidget 的树形装饰（分支线）保留在 Level 列（逻辑索引0）。
        # 汇总模式下 Level 被隐藏，"#"自然落到第0列，无需特殊处理。
        result = base + visible_preset + other_custom
        if "Level" in result:
            level_idx = result.index("Level")
            result.insert(level_idx + 1, BOM_ROW_NUMBER_COLUMN)
        else:
            result.insert(0, BOM_ROW_NUMBER_COLUMN)
        # 完整 BOM 模式：在"零件编号"列后紧插"实例名"列
        if self._full_bom and "Part Number" in result:
            pn_idx = result.index("Part Number")
            # 倒序插入，保证顺序：Part Number → 实例名 → 实例描述
            if self._show_description_inst_col:
                result.insert(pn_idx + 1, "description_inst")
            if self._show_instance_name_col:
                result.insert(pn_idx + 1, BOM_INSTANCE_NAME_COLUMN)
        return result

    def _on_preset_col_toggled(self) -> None:
        # "文件名"复选框控制内置文件名列的可见性
        if "Filename" in self._preset_checkboxes:
            new_show_fn = self._preset_checkboxes["Filename"].isChecked()
            if new_show_fn != self._show_filename_col:
                self._show_filename_col = new_show_fn
                self._edit_settings.setValue("show_filename_column", self._show_filename_col)
        self._visible_preset_cols = [
            name for name, cb in self._preset_checkboxes.items()
            if name != "Filename" and name not in BOM_HIDEABLE_COLUMNS and cb.isChecked()
        ]
        self._edit_settings.setValue("visible_preset_columns", self._visible_preset_cols)
        self._rebuild_columns_and_repopulate()

    def _on_hideable_col_toggled(self) -> None:
        """处理可隐藏列复选框切换（术语、版本、定义、来源）。"""
        self._visible_hideable_cols = [
            name for name, cb in self._preset_checkboxes.items()
            if name in BOM_HIDEABLE_COLUMNS and cb.isChecked()
        ]
        self._edit_settings.setValue("visible_hideable_columns", self._visible_hideable_cols)
        self._rebuild_columns_and_repopulate()

    def _on_show_filepath_toggled(self, checked: bool) -> None:
        self._show_filepath_col = checked
        self._edit_settings.setValue("show_filepath_column", checked)
        self._rebuild_columns_and_repopulate()

    def _on_full_bom_cols_toggled(self) -> None:
        """处理完整 BOM 专有列（实例名 / 实例描述）的可见性切换。"""
        self._show_instance_name_col    = self._show_instance_name_chk.isChecked()
        self._show_description_inst_col = self._show_description_inst_chk.isChecked()
        self._edit_settings.setValue("show_instance_name_col",    self._show_instance_name_col)
        self._edit_settings.setValue("show_description_inst_col", self._show_description_inst_col)
        self._rebuild_columns_and_repopulate()

    # ── 文件选择 ──────────────────────────────────────────────────────────────

    def _browse_file(self) -> None:
        file, _ = QFileDialog.getOpenFileName(
            self, "选择 CATProduct 文件",
            self._last_browse_dir,
            "*.CATProduct (*.CATProduct);;All Files (*)",
        )
        if file:
            self._file_edit.setText(file)
            self._last_browse_dir = str(Path(file).parent)
            self._export_settings.setValue("last_browse_dir", self._last_browse_dir)

    # ── 加载BOM ───────────────────────────────────────────────────────────────

    def _load_bom(self) -> None:
        if self._use_active_chk.isChecked():
            file_path = None
        else:
            file_path = self._file_edit.text().strip()
            if not file_path:
                QMessageBox.warning(self, "未选择文件", "请先选择一个 CATProduct 文件。")
                return
            if not Path(file_path).exists():
                QMessageBox.warning(self, "文件不存在", f"文件不存在：\n{file_path}")
                return

        self._load_btn.setEnabled(False)
        self._load_btn.setText("加载中…")
        QApplication.processEvents()

        progress = QProgressDialog("正在加载 BOM ，请稍候…", None, 0, 0, self)
        progress.setWindowTitle("加载 BOM")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(300)
        progress.setValue(0)

        def _on_row_collected(count: int) -> None:
            progress.setLabelText(f"正在加载 BOM ，请稍候… 已读取 {count} 个节点")
            progress.repaint()
            QApplication.processEvents()

        try:
            all_read_cols = list(dict.fromkeys(
                BOM_EDIT_COLUMN_ORDER
                + [c for c in self._all_custom_columns if c not in BOM_EDIT_COLUMN_ORDER]
            ))
            root_pm_key, part_masters, inst_key_to_info = collect_bom_part_masters(
                file_path, all_read_cols, self._all_custom_columns,
                progress_callback=_on_row_collected,
            )
        except Exception as e:
            progress.close()
            logger.error(f"Failed to load BOM for edit: {e}")
            QMessageBox.critical(
                self, "加载失败",
                f"加载 BOM 时出错：\n{e}\n\n请确保 CATIA 已启动。",
            )
            self._load_btn.setEnabled(True)
            self._load_btn.setText("加载 BOM")
            return
        finally:
            progress.close()

        self._load_btn.setEnabled(True)
        self._load_btn.setText("重新加载 BOM")

        # 保存 part_masters 属性仓库、根 pm_key 和反向索引
        self._root_pm_key     = root_pm_key
        self._part_masters     = part_masters
        self._inst_key_to_info = inst_key_to_info

        # 根据当前模式生成显示行
        self._rebuild_rows()

        # 加载新BOM时清空撤销/重做历史
        self._undo_stack.clear()
        self._redo_stack.clear()
        self._update_undo_redo_buttons()

        # 刷新前按列名保存当前列宽
        if self._bom_loaded:
            for col_idx, col_name in enumerate(self._columns):
                self._col_widths[col_name] = self._table.columnWidth(col_idx)

        self._populate_table()
        if not self._bom_loaded:
            # 首次加载：自适应所有列宽并初始化缓存；
            # "#"行号列固定默认宽度40像素（之后可像其他列一样调整）
            for _c, col_name in enumerate(self._columns):
                if col_name == BOM_ROW_NUMBER_COLUMN:
                    self._table.setColumnWidth(_c, 40)
                    self._col_widths[col_name] = 40
                else:
                    self._table.resizeColumnToContents(_c)
                    self._col_widths[col_name] = self._table.columnWidth(_c)
            self._bom_loaded = True
        else:
            # 后续重新加载：按列名恢复已保存的列宽
            for col_idx, col_name in enumerate(self._columns):
                if col_name in self._col_widths:
                    self._table.setColumnWidth(col_idx, self._col_widths[col_name])

        self._export_btn.setEnabled(True)

    def _populate_table(self) -> None:
        self._is_updating = True
        self._table.blockSignals(True)

        # 汇总模式：所有行均为无子项的顶层项。
        # 若保持 setRootIsDecorated(True)，每行都会预留展开箭头的空间，
        # 使第0列内容向右偏移。汇总模式下禁用，层级模式下重新启用以显示展开箭头。
        self._table.setRootIsDecorated(not self._summarize)

        # 插入行前必须先禁用排序，否则 addTopLevelItem/addChild 会立即按当前排序列重排，
        # 导致完整 BOM / 层级 BOM 的树形顺序被破坏。末尾再按模式决定是否重新启用。
        self._table.setSortingEnabled(False)
        self._table.clear()                          # 删除所有项；表头保留
        headers = self._display_headers()
        self._table.setColumnCount(len(headers))     # Qt 不会自动缩减列数
        self._table.setHeaderLabels(headers)
        self._item_by_row = []
        self._inst_to_items.clear()        # 重置 inst_key→树形项索引
        self._pm_key_to_inst_keys.clear() # 重置 pm_key→[inst_key] 跨实例索引
        self._inst_key_to_product.clear()  # 重置 inst_key→product 补充索引

        # parent_stack：(层级, 树形项|None) 的列表
        # 索引0处的哨兵代表不可见根节点（层级为−1）
        parent_stack: list[tuple[int, QTreeWidgetItem | None]] = [(-1, None)]

        for row_idx, row_data in enumerate(self._rows):
            level = 0 if self._summarize else int(row_data.get("Level", 0))

            # 弹出栈，直到栈顶层级严格低于当前行
            while len(parent_stack) > 1 and parent_stack[-1][0] >= level:
                parent_stack.pop()

            parent_item = parent_stack[-1][1]
            item = _BomSortItem()
            # 将 row_idx 存入第0列的 UserRole，用于反向查找
            item.setData(0, Qt.ItemDataRole.UserRole, row_idx)

            if parent_item is None:
                self._table.addTopLevelItem(item)
            else:
                parent_item.addChild(item)

            parent_stack.append((level, item))
            self._item_by_row.append(item)

            pn         = str(row_data.get("Part Number", ""))
            not_found  = bool(row_data.get("_not_found"))
            no_file    = bool(row_data.get("_no_file"))
            unreadable = bool(row_data.get("_unreadable"))
            row_locked = unreadable or not_found

            # 注册 inst_key→树形项 和 pm_key→inst_key 索引
            inst_key = row_data.get("_inst_key")
            if inst_key is not None:
                self._inst_to_items.setdefault(inst_key, []).append(item)
                pm_key_reg = str(row_data.get("_pm_key", ""))
                if pm_key_reg:
                    self._pm_key_to_inst_keys.setdefault(pm_key_reg, []).append(inst_key)
                # 根节点无 inst_info，需补充 inst_key→product 映射供写回使用
                if inst_key not in self._inst_key_to_info:
                    prod_ref = row_data.get("_product")
                    if prod_ref is not None:
                        self._inst_key_to_product[inst_key] = prod_ref

            for col_idx, col_name in enumerate(self._columns):

                # 来源列 → QComboBox（覆盖控件；不存储为项文本）
                if col_name == "Source":
                    raw    = str(row_data.get("Source", ""))
                    pm_key_src = str(row_data.get("_pm_key", ""))
                    pn_val = get_part_master_attr(self._part_masters, pm_key_src, "Source", raw)
                    # source 存原始值，转换为显示值
                    pn_val = SOURCE_TO_DISPLAY.get(pn_val, SOURCE_OPTIONS[0])
                    if pn_val not in SOURCE_OPTIONS:
                        pn_val = SOURCE_OPTIONS[0]
                    combo = _make_tree_combo(SOURCE_OPTIONS)
                    combo.blockSignals(True)
                    combo.setCurrentText(pn_val)
                    combo.blockSignals(False)
                    if row_locked:
                        combo.setEnabled(False)
                    else:
                        combo.currentTextChanged.connect(
                            lambda text, r=row_idx: self._on_source_changed(r, text)
                        )
                    self._table.setItemWidget(item, col_idx, combo)
                    continue

                # 具有受限选项的用户自定义属性列 → QComboBox
                opts = PRESET_USER_REF_PROPERTY_OPTIONS.get(col_name)
                if opts is not None:
                    pm_key_opt = str(row_data.get("_pm_key", ""))
                    pn_val = get_part_master_attr(
                        self._part_masters, pm_key_opt, col_name,
                        str(row_data.get(col_name, ""))
                    )
                    # 构建有效选项列表：
                    #   • 始终在开头插入""，使未设置的属性显示为空白
                    #   • 若存储值不在允许列表中且非空，则追加以保留原始值可见性
                    display_opts = [""] + list(opts)
                    if pn_val and pn_val not in opts:
                        logger.debug(
                            "属性 '%s' 的值 '%s' 不在可选列表中，将以原始值显示（零件编号: %s）",
                            col_name, pn_val, pn,
                        )
                        display_opts.append(pn_val)
                    combo = _make_tree_combo(display_opts)
                    combo.blockSignals(True)
                    combo.setCurrentText(pn_val)
                    combo.blockSignals(False)
                    if row_locked:
                        combo.setEnabled(False)
                    else:
                        combo.currentTextChanged.connect(
                            lambda text, r=row_idx, c=col_name: self._on_option_col_changed(r, c, text)
                        )
                    self._table.setItemWidget(item, col_idx, combo)
                    continue

                # 其他所有列 → 项文本
                if col_name == BOM_ROW_NUMBER_COLUMN:
                    value = str(row_idx + 1)
                elif col_name == "Quantity":
                    value = str(row_data.get("Quantity", "1"))
                elif col_name == "Filename":
                    fp = str(row_data.get("_filepath", ""))
                    fn = str(row_data.get("Filename", ""))
                    if no_file:
                        # 文件未保存到磁盘：固定显示哨兵文本
                        value = FILENAME_UNSAVED
                    elif self._show_filepath_col:
                        value = fp if fp else fn
                    else:
                        # 已知路径时显示带扩展名的文件名；
                        # 未知路径时回退到存储的文件名茎（可能等于 FILENAME_NOT_FOUND）
                        value = Path(fp).name if fp else fn
                elif col_name == "Filepath":
                    value = str(row_data.get("_filepath", ""))
                elif col_name in BOM_READONLY_COLUMNS:
                    raw = str(row_data.get(col_name, ""))
                    # Type 列存储英文 key，显示时转为中文
                    value = TYPE_DISPLAY_NAMES.get(raw, raw) if col_name == "Type" else raw
                elif col_name == BOM_INSTANCE_NAME_COLUMN:
                    # 实例名：实例级属性，从 _inst_key_to_info 取
                    inst_info = self._inst_key_to_info.get(inst_key) if inst_key is not None else None
                    value = inst_info["instance_name"] if inst_info is not None else str(row_data.get(BOM_INSTANCE_NAME_COLUMN, ""))
                elif col_name == "description_inst":
                    # 实例描述：实例级属性，从 _inst_key_to_info 取
                    inst_info = self._inst_key_to_info.get(inst_key) if inst_key is not None else None
                    value = inst_info.get("description_inst", "") if inst_info is not None else str(row_data.get("description_inst", ""))
                else:
                    # PartMaster 级可写属性（Nomenclature/Revision/Definition/Description/自定义列等）
                    pm_key_cell = str(row_data.get("_pm_key", ""))
                    value = get_part_master_attr(
                        self._part_masters, pm_key_cell, col_name,
                        str(row_data.get(col_name, ""))
                    )
                item.setText(col_idx, value)

                if col_name == "Filename":
                    fp = str(row_data.get("_filepath", ""))
                    fn = str(row_data.get("Filename", ""))
                    if no_file:
                        pass  # tooltip 由下方 no_file 块统一设置
                    elif fp:
                        if self._show_filepath_col:
                            # 列显示完整路径；工具提示显示文件名+扩展名
                            name_with_ext = Path(fp).name
                            if name_with_ext and name_with_ext != FILENAME_NOT_FOUND:
                                item.setToolTip(col_idx, name_with_ext)
                        else:
                            # 列显示文件名+扩展名；工具提示显示完整路径
                            item.setToolTip(col_idx, fp)

            # 未锁定行：允许就地编辑（代理阻止只读列）
            if not row_locked:
                item.setFlags(item.flags() | Qt.ItemFlag.ItemIsEditable)
                item.setData(0, _ITEM_LOCKED_ROLE, False)
            else:
                c    = _get_colors(theme_manager.current_mode())
                grey = c.ROW_LOCKED_FG
                bg   = c.ROW_NOT_FOUND_BG if not_found else c.ROW_LIGHTWEIGHT_BG
                item.setData(0, _ITEM_LOCKED_ROLE, True)
                for ci in range(len(self._columns)):
                    item.setForeground(ci, grey)
                    item.setBackground(ci, bg)
                tip = (
                    "该零件/产品的文件未被 CATIA 检索到，行内容不可编辑。"
                    if not_found else
                    "该零件/产品处于轻量化模式，无法读取属性。"
                )
                for ci in range(len(self._columns)):
                    item.setToolTip(ci, tip)

            # _no_file 行（文件未保存到磁盘）：不锁定，但以淡黄背景和专属提示标识
            if no_file:
                c = _get_colors(theme_manager.current_mode())
                no_file_tip = "该零件尚未保存到磁盘，可通过右键菜单「另存为」将其保存。"
                for ci in range(len(self._columns)):
                    item.setBackground(ci, c.ROW_UNSAVED_BG)
                    item.setToolTip(ci, no_file_tip)

        self._table.expandAll()
        self._table.blockSignals(False)
        self._is_updating = False

        # 汇总模式：启用表头点击排序；层级模式：禁用（保持树结构）
        self._table.setSortingEnabled(self._summarize)

        # 若当前有过滤文本，对重新填充的行重新应用过滤
        if self._filter_text:
            self._apply_filter()

        self._update_status()

    # ── 树形遍历辅助 ──────────────────────────────────────────────────────────

    def _iter_all_items(self):
        """以深度优先前序遍历方式逐个产出所有 QTreeWidgetItem。"""
        def _walk(parent: QTreeWidgetItem):
            yield parent
            for i in range(parent.childCount()):
                yield from _walk(parent.child(i))
        for i in range(self._table.topLevelItemCount()):
            item = self._table.topLevelItem(i)
            if item is not None:
                yield from _walk(item)

    # ── 主题切换响应 ──────────────────────────────────────────────────────────

    def _on_theme_changed(self, mode: str) -> None:
        """主题切换时重新着色所有行。"""
        if not self._bom_loaded:
            return
        self._recolor_all_items(mode)

    def _recolor_all_items(self, mode: str) -> None:
        """按当前主题颜色重新为所有树形行设置背景/前景色。

        仅处理行状态色（locked/unsaved）；已修改字段的橙色前景由
        _refresh_pns_appearance 单独处理。
        """
        c = _get_colors(mode)
        col_count = len(self._columns)
        self._is_updating = True
        try:
            for item in self._iter_all_items():
                row_idx = item.data(0, Qt.ItemDataRole.UserRole)
                if row_idx is None:
                    continue
                row_data   = self._rows[row_idx]
                not_found  = bool(row_data.get("_not_found"))
                unreadable = bool(row_data.get("_unreadable"))
                no_file    = bool(row_data.get("_no_file"))
                row_locked = not_found or unreadable
                # 先清除现有行级颜色（保留列级修改色，由后续 _refresh_pns_appearance 处理）
                default_brush = QBrush()
                for ci in range(col_count):
                    item.setBackground(ci, default_brush)
                    # 前景只清锁定行颜色；已修改字段颜色由 _refresh_pns_appearance 负责
                    if item.data(0, _ITEM_LOCKED_ROLE):
                        item.setData(ci, Qt.ItemDataRole.ForegroundRole, None)
                if row_locked:
                    for ci in range(col_count):
                        item.setForeground(ci, c.ROW_LOCKED_FG)
                        item.setBackground(ci, c.ROW_NOT_FOUND_BG if not_found else c.ROW_LIGHTWEIGHT_BG)
                elif no_file:
                    for ci in range(col_count):
                        item.setBackground(ci, c.ROW_UNSAVED_BG)
        finally:
            self._is_updating = False

    # ── 即时写回核心方法 ──────────────────────────────────────────────────────

    def _get_product(self, inst_key: int):
        """根据 inst_key 取 COM product 引用。
        
        优先从 inst_key_to_info 取（子节点路径），
        找不到则从 _inst_key_to_product 取（根节点补充索引）。
        """
        inst_info = self._inst_key_to_info.get(inst_key)
        if inst_info is not None:
            return inst_info.get("product")
        return self._inst_key_to_product.get(inst_key)

    def _write_cell_to_catia(self, inst_key, col_name: str, value: str,
                             label: str = "已写回") -> bool:
        """通过缓存 COM 引用将单个单元格值立即写入 CATIA。返回是否成功。

        参数：
            inst_key: ``id(product)``，COM 实例对象的唯一标识。
            col_name / value: 列名与新值。
            label: 状态栏前缀（"已写回" / "已撤销" / "已重做"）。

        写入目标通过 ``_get_product(inst_key)`` 取得：
        - 独立文件：写一次，``write_cell`` 内部经由 ``ReferenceProduct`` 覆盖所有实例。
        - Component：写对应的实例 COM 对象，不走 ``ReferenceProduct``。
        """
        inst_info = self._inst_key_to_info.get(inst_key)
        product   = self._get_product(inst_key)
        if product is None:
            self._last_write_status = f"⚠ 未找到 COM 引用（inst_key={inst_key!r}）"
            self._update_status()
            return False

        # pn 用于状态栏：优先从 inst_info 取，根节点从 part_masters 反查
        pn = inst_info.get("pn", "") if inst_info is not None else ""
        if not pn:
            for pm in self._part_masters.values():
                if pm.get("_product") is product:
                    pn = pm.get("part_number", "")
                    break
        try:
            success = write_cell(product, col_name, value, self._all_custom_columns)
            # bom_write.py 内部已 warning 记录底层原因，此处只更新状态栏
            if success:
                self._last_write_status = f"{label}：{pn!r}.{col_name} = {value!r}"
            else:
                self._last_write_status = f"⚠ 写入失败：{pn!r}.{col_name} = {value!r}"
        except Exception as e:
            success = False
            self._last_write_status = f"⚠ 写入异常：{pn!r}.{col_name}: {e}"
            logger.error("_write_cell_to_catia: 异常 inst_key=%r col=%r: %s",
                         inst_key, col_name, e)
        self._update_status()
        return success

    def _sync_pn_siblings_in_ui(
        self,
        pm_key: str,
        col_name: str,
        value: str,
        col_idx: int | None = None,
        exclude_inst_keys: set | None = None,
    ) -> None:
        """将属性变更同步到 UI 中所有同 pm_key 的实例行。

        V3 版本：通过 _pm_key_to_inst_keys[pm_key] 直接找所有同 pm_key 实例。

        参数：
            pm_key:          PartMaster 的 pm_key。
            col_name:         被编辑的列名。
            value:            新值（显示值，Source 列为中文显示值）。
            col_idx:          列在 _columns 中的索引，None 时自动查找。
            exclude_inst_keys: 已直接写入的 inst_key 集合（这些行跳过界面更新）。
        """
        if col_idx is None:
            col_idx = self._columns.index(col_name) if col_name in self._columns else None
        if col_idx is None:
            return

        all_insts = self._pm_key_to_inst_keys.get(pm_key, [])
        if not all_insts:
            return

        excluded = exclude_inst_keys or set()
        self._is_updating = True
        try:
            for inst_key in all_insts:
                if inst_key in excluded:
                    continue
                for tree_item in self._inst_to_items.get(inst_key, []):
                    combo = self._table.itemWidget(tree_item, col_idx)
                    if isinstance(combo, QComboBox):
                        if combo.currentText() != value:
                            combo.blockSignals(True)
                            combo.setCurrentText(value)
                            combo.blockSignals(False)
                    else:
                        if tree_item.text(col_idx) != value:
                            tree_item.setText(col_idx, value)
        finally:
            self._is_updating = False

    # ── "来源"下拉框变更 ──────────────────────────────────────────────────────

    def _on_source_changed(self, row_idx: int, text: str) -> None:
        if self._is_updating:
            return
        # Source 显示值→原始值（"0"/"1"/"2"），写入 part_master 存原始值
        raw = {"未知": "0", "自制": "1", "外购": "2"}.get(text, text)
        self._handle_combo_col_change(row_idx, "Source", display_value=text, store_value=raw)

    # ── 用户自定义选项列变更 ──────────────────────────────────────────────────

    def _on_option_col_changed(self, row_idx: int, col_name: str, text: str) -> None:
        if self._is_updating:
            return
        self._handle_combo_col_change(row_idx, col_name, display_value=text, store_value=text)

    def _handle_combo_col_change(
        self,
        row_idx: int,
        col_name: str,
        display_value: str,
        store_value: str,
    ) -> None:
        """处理 QComboBox 列（Source / 受限选项列）的变更。

        display_value: 界面显示值（同步到 UI 和撤销栈）。
        store_value:   写入 part_master 的值（Source 为原始值 "0"/"1"/"2"，其余同 display_value）。
        """
        if col_name not in self._columns:
            return
        col_idx = self._columns.index(col_name)

        selected_row_indices = {
            it.data(0, Qt.ItemDataRole.UserRole)
            for it in self._table.selectedItems()
            if it.data(0, Qt.ItemDataRole.UserRole) is not None
        }
        direct_rows = selected_row_indices if row_idx in selected_row_indices else {row_idx}

        insts_to_update: set = set()
        for r in direct_rows:
            ik = self._rows[r].get("_inst_key")
            if ik is not None:
                insts_to_update.add(ik)

        old_vals: dict = {}
        affected_pm_keys: set[str] = set()
        for r in direct_rows:
            pm_key = str(self._rows[r].get("_pm_key", "")).strip()
            if not pm_key or pm_key not in self._part_masters:
                continue
            if pm_key in affected_pm_keys:
                continue
            # 旧值记录为显示值（撤销时恢复到界面）
            old_store   = get_part_master_attr(self._part_masters, pm_key, col_name, "")
            old_display = SOURCE_TO_DISPLAY.get(old_store, old_store) if col_name == "Source" else old_store
            old_vals[pm_key] = old_display
            set_part_master_attr(self._part_masters, pm_key, col_name, store_value)
            write_ok = False
            for ik in self._pm_key_to_inst_keys.get(pm_key, []):
                if self._get_product(ik) is not None:
                    write_ok = self._write_cell_to_catia(ik, col_name, display_value)
                    insts_to_update.add(ik)
                    break
            if write_ok:
                affected_pm_keys.add(pm_key)
            else:
                # 写入失败：回滚内存，不加入 affected_pm_keys（不更新界面为新值）
                set_part_master_attr(self._part_masters, pm_key, col_name, old_store)
                # 回滚触发行的 combo 显示值
                self._is_updating = True
                try:
                    for r in direct_rows:
                        if str(self._rows[r].get("_pm_key", "")).strip() == pm_key:
                            src_item = self._item_by_row[r] if r < len(self._item_by_row) else None
                            if src_item is not None:
                                combo = self._table.itemWidget(src_item, col_idx)
                                if isinstance(combo, QComboBox):
                                    combo.blockSignals(True)
                                    combo.setCurrentText(old_display)
                                    combo.blockSignals(False)
                finally:
                    self._is_updating = False

        for pm_key in affected_pm_keys:
            self._sync_pn_siblings_in_ui(pm_key, col_name, display_value, col_idx=col_idx)

        undo_actions = [
            (pm_key, col_name, old_vals[pm_key], display_value)
            for pm_key in affected_pm_keys
            if pm_key in old_vals and old_vals[pm_key] != display_value
        ]
        if undo_actions:
            self._push_undo(undo_actions)
        self._refresh_keys_appearance(insts_to_update)

    # ── 普通单元格编辑 ────────────────────────────────────────────────────────

    def _on_item_changed(self, item: QTreeWidgetItem, col_idx: int) -> None:
        if self._is_updating:
            return
        row_idx = item.data(0, Qt.ItemDataRole.UserRole)
        if row_idx is None:
            return
        col_name = self._columns[col_idx]

        if col_name in BOM_READONLY_COLUMNS or col_name == "Source" or col_name in PRESET_USER_REF_PROPERTY_OPTIONS:
            return

        new_value = item.text(col_idx)
        pm_key   = str(self._rows[row_idx].get("_pm_key", ""))
        pn        = str(self._rows[row_idx].get("Part Number", ""))
        _this_p   = self._rows[row_idx].get("_product")
        this_inst = id(_this_p) if _this_p is not None else None

        # ── 实例名称（完整 BOM 模式专用，按行而非按 PN 处理）──────────────────
        if col_name == BOM_INSTANCE_NAME_COLUMN:
            # 根节点（Level=0）没有实例名，拒绝编辑
            # 注意：不能用 _parent_inst_key is None 判断——根节点直接子实例的
            # _parent_inst_key 也是 None（iter_full_rows 第一次调用传 parent_inst_key=None）
            if int(self._rows[row_idx].get("Level", 0)) == 0:
                self._is_updating = True
                item.setText(col_idx, "")
                self._is_updating = False
                return
            self._handle_instance_name_changed(item, col_idx, row_idx, new_value)
            return

        # ── 实例描述（完整 BOM 模式专用）────────────────────────────────
        if col_name == "description_inst":
            self._handle_description_inst_changed(item, col_idx, row_idx, new_value)
            return

        if col_name == "Part Number":
            # ── 零件编号为空或仅含空格 ────────────────────────────────────────
            if not new_value.strip():
                QMessageBox.warning(
                    self, "零件编号不能为空",
                    "零件编号不能为空或仅含空格，请输入有效的零件编号。",
                )
                self._is_updating = True
                item.setText(col_idx, pn)   # 回退到行中存储的旧 PN
                self._is_updating = False
                return

            # ── 静默去除首尾空格 ──────────────────────────────────────────────
            if new_value != new_value.strip():
                new_value = new_value.strip()
                self._is_updating = True
                item.setText(col_idx, new_value)
                self._is_updating = False

            # ── 字符合法性校验 ────────────────────────────────────────────────
            if not PART_NUMBER_VALID_PATTERN.fullmatch(new_value):
                QMessageBox.warning(
                    self, "零件编号含非法字符",
                    f"零件编号 \"{new_value}\" 含有非法字符。\n"
                    "不允许：控制字符、非 ASCII 字符，以及 Windows 文件名禁用字符"
                    "（\\ / : * ? \" < > |）。",
                )
                self._is_updating = True
                item.setText(col_idx, pn)
                self._is_updating = False
                return

            # ── V3：PN 冲突检查（在下方 direct_rows 循环内逐行检查，此处仅做语法校验）
            # 冲突检查移入循环内以正确处理多选多 PN 场景

        selected_row_indices = {
            it.data(0, Qt.ItemDataRole.UserRole)
            for it in self._table.selectedItems()
            if it.data(0, Qt.ItemDataRole.UserRole) is not None
        }
        direct_rows = selected_row_indices if row_idx in selected_row_indices else {row_idx}

        # PN 冲突检查：在循环内逐行检查，任一行冲突则整体拒绝并回退触发行的显示
        if col_name == "Part Number":
            for r in direct_rows:
                row_pm_key = str(self._rows[r].get("_pm_key", "")).strip()
                if self._is_pn_conflicting(row_pm_key, new_value):
                    msg = (
                        f"零件编号 \"{new_value}\" 与同一产品文件内已有的零件编号冲突，"
                        "CATIA 不允许。"
                        if ":" in row_pm_key else
                        f"零件编号 \"{new_value}\" 与现有零件编号冲突，不允许修改。"
                    )
                    QMessageBox.warning(self, "零件编号冲突", msg)
                    self._is_updating = True
                    item.setText(col_idx, pn)
                    self._is_updating = False
                    return

        # V3：以 pm_key 为单位操作 part_masters；收集所有涉及的 pm_key 和对应的写入实例
        # 每个 pm_key 只写一个实例（CATIA 自动同步），界面更新通过 _pm_key_to_inst_keys 覆盖所有实例
        old_pm_key_vals: dict[str, str] = {}   # pm_key → old_value（撤销用）
        affected_pm_keys: set[str]      = set()
        insts_to_update:   set[int]      = set()  # 用于 _refresh_keys_appearance

        for r in direct_rows:
            row_pm_key = str(self._rows[r].get("_pm_key", "")).strip()
            if not row_pm_key or row_pm_key not in self._part_masters:
                continue
            if row_pm_key in affected_pm_keys:
                continue   # 同 pm_key 只处理一次

            old_val = get_part_master_attr(self._part_masters, row_pm_key, col_name, "")
            if old_val == new_value:
                continue   # 无变化，跳过

            old_pm_key_vals[row_pm_key] = old_val
            affected_pm_keys.add(row_pm_key)

            if col_name == "Part Number":
                # PN 修改：pm_key 永不变，只更新 pm["part_number"] 和 inst_info["pn"]
                success = rename_part_master(
                    self._part_masters,
                    self._pm_key_to_inst_keys,
                    row_pm_key,
                    new_value,
                )
                if not success:
                    logger.error(
                        "_on_item_changed: rename_part_master 失败 pm_key=%r new_pn=%r",
                        row_pm_key, new_value,
                    )
                    self._last_write_status = f"⚠ 零件编号修改失败（pm_key={row_pm_key!r}）"
                    self._update_status()
                    self._is_updating = True
                    item.setText(col_idx, old_val)
                    self._is_updating = False
                    continue
                # 同步更新 _rows 里所有匹配此 pm_key 的行的 Part Number 显示字段
                for row in self._rows:
                    if row.get("_pm_key") == row_pm_key:
                        row["Part Number"] = new_value
                # 写回 CATIA：用任意一个有效实例写一次
                write_ok = False
                for ik in self._pm_key_to_inst_keys.get(row_pm_key, []):
                    if self._get_product(ik) is not None:
                        write_ok = self._write_cell_to_catia(ik, col_name, new_value)
                        insts_to_update.add(ik)
                        break
                if not write_ok:
                    # 写入失败：回滚内存和 _rows 的 Part Number 字段
                    rename_part_master(self._part_masters, self._pm_key_to_inst_keys,
                                       row_pm_key, old_pm_key_vals[row_pm_key])
                    for row in self._rows:
                        if row.get("_pm_key") == row_pm_key:
                            row["Part Number"] = old_pm_key_vals[row_pm_key]
                    affected_pm_keys.discard(row_pm_key)
                    del old_pm_key_vals[row_pm_key]
            else:
                set_part_master_attr(self._part_masters, row_pm_key, col_name, new_value)
                # 写回 CATIA：用任意一个有效实例写一次
                write_ok = False
                for ik in self._pm_key_to_inst_keys.get(row_pm_key, []):
                    if self._get_product(ik) is not None:
                        write_ok = self._write_cell_to_catia(ik, col_name, new_value)
                        insts_to_update.add(ik)
                        break
                if not write_ok:
                    # 写入失败：回滚内存
                    set_part_master_attr(self._part_masters, row_pm_key, col_name,
                                         old_pm_key_vals[row_pm_key])
                    affected_pm_keys.discard(row_pm_key)
                    del old_pm_key_vals[row_pm_key]

        if not affected_pm_keys:
            # 所有写入均失败，回滚触发行的界面显示为旧值
            # 从已回滚的 part_master 重新读取旧值（内存已恢复）
            pm_key_trigger = str(self._rows[row_idx].get("_pm_key", ""))
            old_display = get_part_master_attr(
                self._part_masters, pm_key_trigger, col_name, pn
            )
            self._is_updating = True
            item.setText(col_idx, old_display)
            self._is_updating = False
            return

        # 更新所有同 pm_key 实例的界面（通过 _pm_key_to_inst_keys 遍历）
        self._is_updating = True
        try:
            for bk in affected_pm_keys:
                for ik in self._pm_key_to_inst_keys.get(bk, []):
                    for tree_item in self._inst_to_items.get(ik, []):
                        if tree_item.text(col_idx) != new_value:
                            tree_item.setText(col_idx, new_value)
        finally:
            self._is_updating = False

        # 推入撤销栈（key 为 str = pm_key）
        undo_actions = [
            (bk, col_name, old_pm_key_vals[bk], new_value)
            for bk in affected_pm_keys
            if bk in old_pm_key_vals
        ]
        if undo_actions:
            self._push_undo(undo_actions)
        self._refresh_keys_appearance(insts_to_update)

    # ── 实例名称编辑（完整 BOM 模式） ────────────────────────────────────────

    def _handle_instance_name_changed(
        self,
        item: QTreeWidgetItem,
        col_idx: int,
        row_idx: int,
        new_value: str,
    ) -> None:
        """处理"实例名"列的就地编辑（完整 BOM 模式专用）。

        - 不允许空值。
        - 同父节点下的兄弟实例名称不能重复（CATIA 约束）。
        - 通过 ``product.Name = value`` 写入 CATIA。
        - 推入撤销栈，支持 Ctrl+Z 撤销。
        - 实例名唯一真相：_inst_key_to_info[inst_key]["instance_name"]（part_masters 树的引用）。
        """
        row_data         = self._rows[row_idx]
        cur_product      = row_data.get("_product")
        inst_key         = row_data.get("_inst_key") or (id(cur_product) if cur_product else None)
        inst_info        = self._inst_key_to_info.get(inst_key) if inst_key is not None else None
        parent_inst_key  = row_data.get("_parent_inst_key")
        parent_inst_info = self._inst_key_to_info.get(parent_inst_key) if parent_inst_key is not None else None
        parent_pm_key   = parent_inst_info["pm_key"] if parent_inst_info is not None else self._root_pm_key

        # inst_info 为 None 表示数据层异常（非根节点却找不到实例信息），拒绝编辑
        if inst_info is None:
            logger.error("_handle_instance_name_changed: inst_info 为 None，inst_key=%r，拒绝编辑",
                         inst_key)
            self._is_updating = True
            item.setText(col_idx, str(row_data.get(BOM_INSTANCE_NAME_COLUMN, "")))
            self._is_updating = False
            return

        # old_val 提前计算，确保 _rollback 总能取到正确的旧值（inst_info 已确认非 None）
        old_val = inst_info["instance_name"]

        def _rollback() -> None:
            self._is_updating = True
            item.setText(col_idx, old_val)
            self._is_updating = False
            self._update_status()   # 立即刷新状态栏，不等外层调用

        if not new_value.strip():
            QMessageBox.warning(self, "实例名称不能为空",
                                "实例名称不能为空或仅含空格。")
            _rollback()
            return

        # ── 同父唯一性检查：同父节点下所有子实例的实例名必须唯一（CATIA 约束）──
        # 不限于同 pm_key，跨不同 Part/Component 的实例名也不能重复
        parent_pm_check = self._part_masters.get(parent_pm_key, {})
        for sib in parent_pm_check.get("instances", []):
            if sib.get("inst_key") == inst_key:
                continue   # 跳过自身
            if sib.get("instance_name") == new_value:
                QMessageBox.warning(
                    self, "实例名称冲突",
                    f"同一父节点下已存在实例名称 \"{new_value}\"，\n"
                    "请使用不同的实例名称。",
                )
                _rollback()
                return

        # ── 即时写回 CATIA ────────────────────────────────────────────────────
        if cur_product is not None:
            try:
                ok = write_cell(cur_product, BOM_INSTANCE_NAME_COLUMN,
                                new_value, self._all_custom_columns)
                if not ok:
                    # bom_write.py 已 warning 记录底层原因，此处只更新状态栏并回滚
                    self._last_write_status = f"⚠ 写入失败：实例名 {old_val!r} → {new_value!r}"
                    _rollback()
                else:
                    self._last_write_status = f"已写回：实例名 {old_val!r} → {new_value!r}"

                    # 更新 inst_info（唯一真相）
                    inst_info["instance_name"] = new_value

                    # 刷新所有共享同一 inst_key 的树形项
                    col_idx_inst = self._columns.index(BOM_INSTANCE_NAME_COLUMN) if BOM_INSTANCE_NAME_COLUMN in self._columns else -1
                    if col_idx_inst >= 0:
                        self._is_updating = True
                        try:
                            for other_item in self._inst_to_items.get(inst_key, []):
                                if other_item is not item and other_item.text(col_idx_inst) != new_value:
                                    other_item.setText(col_idx_inst, new_value)
                        finally:
                            self._is_updating = False

                    # 推入撤销栈（仅写入成功时）
                    self._push_undo([(inst_key, BOM_INSTANCE_NAME_COLUMN, old_val, new_value)])

            except Exception as e:
                self._last_write_status = f"⚠ 写入异常：实例名 {old_val!r}: {e}"
                logger.error("_handle_instance_name_changed: 异常 inst_key=%r: %s", inst_key, e)
                _rollback()
        else:
            self._last_write_status = f"⚠ 未找到实例 COM 引用（行 {row_idx + 1}）"
        self._update_status()

    # ── 批量单元格写入（内部辅助） ────────────────────────────────────────────

    def _handle_description_inst_changed(
        self,
        item: QTreeWidgetItem,
        col_idx: int,
        row_idx: int,
        new_value: str,
    ) -> None:
        """处理“实例描述”列的就地编辑（完整 BOM 模式专用）。

        - 允许空字符串（与实例名不同，实例描述无要求非空）。
        - 无唯一性限制（允许多个实例有相同描述）。
        - 通过 ``product.DescriptionInst = value`` 写入 CATIA。
        - 推入撤销栈，支持 Ctrl+Z 撤销。
        - 实例描述唯一真相：_inst_key_to_info[inst_key]["description_inst"]。
        """
        row_data    = self._rows[row_idx]
        cur_product = row_data.get("_product")
        inst_key    = row_data.get("_inst_key") or (id(cur_product) if cur_product else None)
        inst_info   = self._inst_key_to_info.get(inst_key) if inst_key is not None else None

        if inst_info is None:
            logger.error("_handle_description_inst_changed: inst_info 为 None，inst_key=%r，拒绝编辑",
                         inst_key)
            self._is_updating = True
            item.setText(col_idx, "")
            self._is_updating = False
            return

        old_val = inst_info.get("description_inst", "")

        def _rollback() -> None:
            self._is_updating = True
            item.setText(col_idx, old_val)
            self._is_updating = False
            self._update_status()

        # ── 即时写回 CATIA ──────────────────────────────────────────
        if cur_product is not None:
            try:
                cur_product.DescriptionInst = new_value
                self._last_write_status = f"已写回：实例描述 {old_val!r} → {new_value!r}"

                # 更新 inst_info（唯一真相）
                inst_info["description_inst"] = new_value

                # 刷新所有共享同一 inst_key 的树形项
                if col_idx >= 0:
                    self._is_updating = True
                    try:
                        for other_item in self._inst_to_items.get(inst_key, []):
                            if other_item is not item and other_item.text(col_idx) != new_value:
                                other_item.setText(col_idx, new_value)
                    finally:
                        self._is_updating = False

                # 推入撤销栈
                self._push_undo([(inst_key, "description_inst", old_val, new_value)])

            except Exception as e:
                self._last_write_status = f"⚠ 写入异常：实例描述 {old_val!r}: {e}"
                logger.error("_handle_description_inst_changed: 异常 inst_key=%r: %s", inst_key, e)
                _rollback()
        else:
            self._last_write_status = f"⚠ 未找到实例 COM 引用（行 {row_idx + 1}）"
        self._update_status()

    def _is_pn_conflicting(self, pm_key: str, new_pn: str) -> bool:
        """检查将 pm_key 对应的零件重命名为 new_pn 是否与现有零件编号冲突。

        基于当前内存状态（_part_masters），在批量操作中可逐步调用以感知中间写入状态。
        返回 True 表示冲突（不允许修改），False 表示无冲突。
        """
        if ":" not in pm_key:
            # 独立文件节点：new_pn 不能与其他任何独立节点的 part_number 相同
            for bk2, pm2 in self._part_masters.items():
                if bk2 == pm_key:
                    continue
                if ":" in bk2:
                    continue   # 跳过嵌入部件
                if pm2.get("part_number") == new_pn:
                    return True
        else:
            # 嵌入部件（pm_key = "pn:host_file_pn"）：
            # 同一宿主文件内所有节点（根节点 + 所有 Component）的 PN 不能重复
            cur_pm       = self._part_masters.get(pm_key, {})
            host_file_pn = cur_pm.get("host_file_pn", "")
            occupied: set[str] = set()
            host_pm = self._part_masters.get(host_file_pn)
            if host_pm:
                occupied.add(host_pm.get("part_number", ""))
            for bk, pm in self._part_masters.items():
                if bk == pm_key:
                    continue
                if pm.get("host_file_pn") == host_file_pn and host_file_pn:
                    occupied.add(pm.get("part_number", ""))
            if new_pn in occupied:
                return True
        return False

    def _apply_cell_values(
        self,
        assignments: list[tuple[int, str, str]],
    ) -> int:
        """批量将 (row_idx, col_name, new_value) 写入数据层和界面，并推入撤销栈。

        - 每个 row_idx 对应 ``self._rows`` 的索引。
        - 只写入可编辑列（不在 BOM_READONLY_COLUMNS 中）且行未锁定的单元格。
        - 同一零件编号的多行通过 ``_key_to_items`` 一并刷新（包括 combo widget）。
        - 整批 assignments 作为一个原子步骤推入 ``_undo_stack``。

        返回写入失败的次数（CATIA 拒绝或 COM 异常）；0 表示全部成功。
        """
        if not assignments:
            return 0

        old_vals: dict = {}   # (pm_key, col_name) → old_value
        insts_to_update: set = set()
        affected_pm_keys: set[str] = set()
        fail_count: int = 0

        for row_idx, col_name, new_value in assignments:
            if row_idx < 0 or row_idx >= len(self._rows):
                continue
            row = self._rows[row_idx]
            if row.get("_not_found") or row.get("_unreadable"):
                continue
            pm_key = str(row.get("_pm_key", "")).strip()
            if not pm_key or pm_key not in self._part_masters:
                continue
            if col_name in BOM_READONLY_COLUMNS or col_name in (BOM_INSTANCE_NAME_COLUMN, "description_inst"):
                # 只读列、实例名列和实例描述列不走 part_master 路径
                continue

            old_val = get_part_master_attr(self._part_masters, pm_key, col_name, "")
            key = (pm_key, col_name)
            if key not in old_vals:   # 只记录每个 pm_key+列的第一次旧值
                old_vals[key] = old_val

            # PN 修改：pm_key 永不变，通过 rename_part_master 同步 inst_info["pn"]
            if col_name == "Part Number":
                # 冲突前置检查（基于当前内存状态，感知批量中间写入状态）
                if self._is_pn_conflicting(pm_key, new_value):
                    fail_count += 1
                    old_vals.pop(key, None)
                    continue
                if not rename_part_master(
                    self._part_masters, self._pm_key_to_inst_keys, pm_key, new_value,
                ):
                    old_vals.pop(key, None)
                    continue
                for row in self._rows:
                    if row.get("_pm_key") == pm_key:
                        row["Part Number"] = new_value
            else:
                set_part_master_attr(self._part_masters, pm_key, col_name, new_value)
            affected_pm_keys.add(pm_key)

            # 写回 CATIA：用任意一个有效实例写一次
            write_ok = False
            for ik in self._pm_key_to_inst_keys.get(pm_key, []):
                if self._get_product(ik) is not None:
                    write_ok = self._write_cell_to_catia(ik, col_name, new_value)
                    insts_to_update.add(ik)
                    break
            if not write_ok:
                # 写入失败：回滚内存，不更新界面
                fail_count += 1
                if col_name == "Part Number":
                    rename_part_master(
                        self._part_masters, self._pm_key_to_inst_keys, pm_key, old_val,
                    )
                    for row in self._rows:
                        if row.get("_pm_key") == pm_key:
                            row["Part Number"] = old_val
                else:
                    set_part_master_attr(self._part_masters, pm_key, col_name, old_val)
                affected_pm_keys.discard(pm_key)
                old_vals.pop(key, None)

        if not affected_pm_keys:
            return fail_count

        # 刷新界面（V3：通过 _pm_key_to_inst_keys 覆盖所有同 pm_key 实例）
        self._is_updating = True
        try:
            for row_idx, col_name, new_value in assignments:
                if row_idx < 0 or row_idx >= len(self._rows):
                    continue
                pm_key = str(self._rows[row_idx].get("_pm_key", "")).strip()
                if not pm_key or pm_key not in affected_pm_keys:
                    continue
                col_idx = self._columns.index(col_name) if col_name in self._columns else -1
                if col_idx < 0:
                    continue
                for ik in self._pm_key_to_inst_keys.get(pm_key, []):
                    for tree_item in self._inst_to_items.get(ik, []):
                        widget = self._table.itemWidget(tree_item, col_idx)
                        if isinstance(widget, QComboBox):
                            if widget.currentText() != new_value:
                                widget.blockSignals(True)
                                widget.setCurrentText(new_value)
                                widget.blockSignals(False)
                        else:
                            if tree_item.text(col_idx) != new_value:
                                tree_item.setText(col_idx, new_value)
        finally:
            self._is_updating = False

        # 构建撤销动作（按 pm_key+col_name 去重，只保留真正改变的条目）
        seen: set = set()
        undo_actions: list = []
        for row_idx, col_name, new_value in assignments:
            if row_idx < 0 or row_idx >= len(self._rows):
                continue
            pm_key = str(self._rows[row_idx].get("_pm_key", "")).strip()
            key = (pm_key, col_name)
            if key in seen:
                continue
            seen.add(key)
            old_val = old_vals.get(key)
            if old_val is not None and old_val != new_value:
                undo_actions.append((pm_key, col_name, old_val, new_value))

        if undo_actions:
            self._push_undo(undo_actions)

        self._refresh_keys_appearance(insts_to_update)
        self._update_status()
        return fail_count

    # ── 撤销/重做 ─────────────────────────────────────────────────────────────

    def _push_undo(self, actions: list) -> None:
        """将一组字段变更推入撤销栈（最多保留 _MAX_HISTORY 步）。

        Args:
            actions: 每项为 ``(key, col_name, old_val, new_val)``。
                     key 为 str → PartMaster 属性（pm_key）；key 为 int → 实例属性（inst_key）。
        """
        if not actions:
            return
        self._undo_stack.append(actions)
        if len(self._undo_stack) > _MAX_HISTORY:
            self._undo_stack.pop(0)
        self._redo_stack.clear()
        self._update_undo_redo_buttons()

    def _undo(self) -> None:
        """撤销最近一步字段变更。"""
        if not self._undo_stack:
            return
        actions = self._undo_stack.pop()
        self._apply_field_changes(actions, forward=False)
        self._redo_stack.append(actions)
        self._update_undo_redo_buttons()

    def _redo(self) -> None:
        """重做最近一步撤销的字段变更。"""
        if not self._redo_stack:
            return
        actions = self._redo_stack.pop()
        self._apply_field_changes(actions, forward=True)
        self._undo_stack.append(actions)
        self._update_undo_redo_buttons()

    def _apply_field_changes(self, actions: list, *, forward: bool) -> None:
        """将一组字段变更应用到 part_masters / 实例内存和界面。

        V3 版本：
        - key 为 str → PartMaster 属性（pm_key），写 part_masters[pm_key][col]，
          通过 _pm_key_to_inst_keys[pm_key] 更新所有实例界面。
        - key 为 int → 实例属性（inst_key），走 product.Name = value 路径。

        Args:
            actions: 每项为 ``(key, col_name, old_val, new_val)``。
            forward: True 应用 new_val（重做），False 应用 old_val（撤销）。
        """
        insts_affected: set = set()
        _label = "已重做" if forward else "已撤销"
        self._is_updating = True
        try:
            for key, col_name, old_val, new_val in actions:
                value = new_val if forward else old_val

                if isinstance(key, int):
                    # ── 实例属性（Instance Name / description_inst）：key = inst_key ──
                    inst_key  = key
                    inst_info = self._inst_key_to_info.get(inst_key)
                    product   = inst_info["product"] if inst_info is not None else None
                    if product is not None:
                        try:
                            if col_name == BOM_INSTANCE_NAME_COLUMN:
                                product.Name = value
                                self._last_write_status = f"{_label}：实例名 → {value!r}"
                            elif col_name == "description_inst":
                                product.DescriptionInst = value
                                self._last_write_status = f"{_label}：实例描述 → {value!r}"
                        except Exception as e:
                            logger.error("undo/redo instance attr write error col=%r: %s", col_name, e)
                    # 唯一真相：更新 _inst_key_to_info
                    inst_info = self._inst_key_to_info.get(inst_key)
                    if inst_info is not None:
                        if col_name == BOM_INSTANCE_NAME_COLUMN:
                            inst_info["instance_name"] = value
                        elif col_name == "description_inst":
                            inst_info["description_inst"] = value
                    # 更新界面单元格
                    col_idx = self._columns.index(col_name) if col_name in self._columns else -1
                    if col_idx >= 0:
                        for tree_item in self._inst_to_items.get(inst_key, []):
                            tree_item.setText(col_idx, value)
                    insts_affected.add(inst_key)

                else:
                    # ── PartMaster 属性：key = pm_key (str)，永不变 ────────────
                    bk = key

                    if col_name == "Part Number":
                        # PN 改名：pm_key 永不变，直接用 bk 查找并更新 part_number
                        dst_pn = new_val if forward else old_val
                        if bk not in self._part_masters:
                            logger.warning("_apply_field_changes: pm_key=%r 不在 part_masters 中，跳过", bk)
                            continue
                        success = rename_part_master(
                            self._part_masters, self._pm_key_to_inst_keys, bk, dst_pn)
                        if not success:
                            logger.error(
                                "_apply_field_changes: rename_part_master 失败 pm_key=%r dst_pn=%r",
                                bk, dst_pn,
                            )
                            continue
                        # 同步更新 _rows 里匹配 pm_key 的行的 Part Number 显示字段
                        for row in self._rows:
                            if row.get("_pm_key") == bk:
                                row["Part Number"] = dst_pn
                        # 更新界面
                        col_idx = self._columns.index(col_name) if col_name in self._columns else -1
                        if col_idx >= 0:
                            for ik in self._pm_key_to_inst_keys.get(bk, []):
                                for tree_item in self._inst_to_items.get(ik, []):
                                    tree_item.setText(col_idx, dst_pn)
                                insts_affected.add(ik)
                        # 写回 CATIA
                        for ik in self._pm_key_to_inst_keys.get(bk, []):
                            if self._get_product(ik) is not None:
                                self._write_cell_to_catia(ik, col_name, dst_pn, label=_label)
                                break
                    else:
                        # 非 PN 的 PartMaster 属性（Nomenclature/Revision 等）
                        if bk not in self._part_masters:
                            continue
                        set_part_master_attr(self._part_masters, bk, col_name, value)
                        # 写回 CATIA（任意一个有效实例）
                        for ik in self._pm_key_to_inst_keys.get(bk, []):
                            if self._get_product(ik) is not None:
                                self._write_cell_to_catia(ik, col_name, value, label=_label)
                                insts_affected.add(ik)
                                break
                        # 更新所有同 pm_key 实例的界面
                        col_idx = self._columns.index(col_name) if col_name in self._columns else -1
                        if col_idx >= 0:
                            for ik in self._pm_key_to_inst_keys.get(bk, []):
                                for tree_item in self._inst_to_items.get(ik, []):
                                    widget = self._table.itemWidget(tree_item, col_idx)
                                    if isinstance(widget, QComboBox):
                                        widget.blockSignals(True)
                                        widget.setCurrentText(value)
                                        widget.blockSignals(False)
                                    else:
                                        tree_item.setText(col_idx, value)
                                insts_affected.add(ik)

        finally:
            self._is_updating = False

        if insts_affected:
            self._refresh_keys_appearance(insts_affected)

    # ── 相同内容填充 ──────────────────────────────────────────────────────────

    def _fill_same_value(
        self,
        col_name: str,
        row_indices: list[int],
    ) -> None:
        """将选中行中视觉最靠前（row_idx 最小）的行在 col_name 列的值填充到其余行。

        - row_indices 无需排序，方法内部按 row_idx 升序排列。
        - 源行取排序后第一行（表格中最靠上的行）。
        - 来源值从 _part_masters 中读取（优先），combo 列从 widget 读取。
        - 锁定行或只读列静默跳过。
        - 整批变更作为一个原子步骤推入撤销栈。
        """
        if len(row_indices) < 2:
            return
        if col_name in BOM_READONLY_COLUMNS:
            return

        # 按视觉顺序排序（row_idx 即 _rows 中的索引，对应表格从上到下的顺序）
        sorted_indices = sorted(row_indices)
        source_row_idx = sorted_indices[0]
        # 目标行：排除锁定行（未找到文件 / 轻量化），_apply_cell_values 虽然也会跳过，
        # 但提前过滤可避免锁定行占用 source 位置（若第一行是锁定行，整体直接跳过）
        target_row_indices = [
            r for r in sorted_indices[1:]
            if r < len(self._rows)
            and not self._rows[r].get("_not_found")
            and not self._rows[r].get("_unreadable")
        ]

        if source_row_idx >= len(self._rows):
            return
        # 源行本身若锁定，无意义填充，直接返回
        if (self._rows[source_row_idx].get("_not_found")
                or self._rows[source_row_idx].get("_unreadable")):
            return
        if not target_row_indices:
            return

        # 取源值（从 part_masters 读取，优先）
        src_pm_key = str(self._rows[source_row_idx].get("_pm_key", "")).strip()
        if src_pm_key and src_pm_key in self._part_masters:
            src_value = get_part_master_attr(self._part_masters, src_pm_key, col_name, "")
        else:
            src_item = self._item_by_row[source_row_idx] if source_row_idx < len(self._item_by_row) else None
            if src_item is not None:
                col_idx = self._columns.index(col_name) if col_name in self._columns else -1
                if col_idx >= 0:
                    widget = self._table.itemWidget(src_item, col_idx)
                    src_value = widget.currentText() if isinstance(widget, QComboBox) else src_item.text(col_idx)
                else:
                    return
            else:
                return

        assignments = [(r, col_name, src_value) for r in target_row_indices]
        fail_count = self._apply_cell_values(assignments)
        if fail_count:
            QMessageBox.warning(
                self, "写入失败",
                f"有 {fail_count} 处写入未被 CATIA 接受。\n详情见底部状态栏及日志。",
            )

    # ── 序列填充 ──────────────────────────────────────────────────────────────

    def _fill_sequence(
        self,
        col_name: str,
        row_indices: list[int],
    ) -> None:
        """弹出序列填充对话框，按视觉顺序（row_idx 升序）在 col_name 列写入递增序列值。

        - row_indices 无需排序，方法内部按 row_idx 升序排列后再依次赋值。
        - 序列格式：``{前缀}{数字:0位数}``，例如前缀 "A-"、起始 1、步长 1、位数 3
          生成 A-001, A-002, A-003...
        - 写入失败（CATIA 拒绝）时不中止，全部尝试完成后以弹窗汇总失败数。
        """
        if col_name in BOM_READONLY_COLUMNS:
            return
        if len(row_indices) < 2:
            return

        # 按视觉顺序排序，过滤锁定行（未找到文件 / 轻量化）。
        # 锁定行不可编辑，若不过滤会占用序列槽位导致序列与行错位。
        ordered_row_indices: list[int] = []
        for r in sorted(row_indices):
            if r >= len(self._rows):
                continue
            row = self._rows[r]
            if row.get("_not_found") or row.get("_unreadable"):
                continue
            ordered_row_indices.append(r)

        if len(ordered_row_indices) < 2:
            return

        # ── 内联对话框 ────────────────────────────────────────────────────────

        _ROW_H    = 24   # 输入控件统一行高（px）
        _DLG_W    = 400  # 固定宽度（px）
        _SETT_PFX = "fill_seq/"  # QSettings key 前缀

        # ── 字母序列辅助函数（对话框之外也要用） ─────────────────────────────
        def _num_to_alpha(n: int) -> str:
            """1-based 整数 → Excel 列名风格：1→A, 26→Z, 27→AA…"""
            result = ""
            while n > 0:
                n, rem = divmod(n - 1, 26)
                result = chr(ord("A") + rem) + result
            return result

        def _alpha_to_num(s: str) -> int:
            """字母串 → 1-based 整数，与 _num_to_alpha 互逆。"""
            result = 0
            for ch in s.upper():
                result = result * 26 + (ord(ch) - ord("A") + 1)
            return result

        # ── 读取持久化设置 ────────────────────────────────────────────────────
        saved_prefix = self._edit_settings.value(_SETT_PFX + "prefix", "")
        saved_suffix = self._edit_settings.value(_SETT_PFX + "suffix", "")

        dlg = QDialog(self)
        dlg.setWindowTitle("序列填充")
        dlg.setFixedWidth(_DLG_W)
        root = QVBoxLayout(dlg)
        root.setSpacing(10)
        root.setContentsMargins(16, 12, 16, 12)

        # ── 统一 QGridLayout（7列：左padding + 3数据列×(标题+输入) + 右padding） ──
        # 实际布局：4列数据列，左右各1列 stretch 列用于居中
        # col: 0=stretch  1=前缀/起始  2=序列值/步长  3=后缀/位数  4=stretch

        _COL_W = 100  # 所有输入框统一宽度（px）

        main_grid = QGridLayout()
        main_grid.setHorizontalSpacing(20)
        main_grid.setVerticalSpacing(4)
        # 左右 stretch 列撑开，三个数据列等宽固定
        main_grid.setColumnStretch(0, 1)
        main_grid.setColumnStretch(4, 1)
        for c in (1, 2, 3):
            main_grid.setColumnMinimumWidth(c, _COL_W)
            main_grid.setColumnStretch(c, 0)

        # ── 行0：前缀 / 序列值（第1项） / 后缀 标题 ──────────────────────────
        lbl_prefix = QLabel("前缀")
        lbl_seq1h  = QLabel("序列值（第 1 项）")
        lbl_suffix = QLabel("后缀")
        for lbl in (lbl_prefix, lbl_seq1h, lbl_suffix):
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(lbl_prefix, 0, 1, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(lbl_seq1h,  0, 2, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(lbl_suffix, 0, 3, Qt.AlignmentFlag.AlignCenter)

        # ── 行1：前缀输入 / 序列值只读预览 / 后缀输入 ────────────────────────
        prefix_edit = QLineEdit(saved_prefix)
        prefix_edit.setPlaceholderText("（可为空）")
        prefix_edit.setFixedHeight(_ROW_H)
        prefix_edit.setFixedWidth(_COL_W)

        seq1_label = QLabel()
        seq1_label.setFixedHeight(_ROW_H)
        seq1_label.setFixedWidth(_COL_W)
        seq1_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        seq1_label.setToolTip("第 1 项序列值（不含前后缀）")
        seq1_label.setFrameShape(seq1_label.Shape.StyledPanel)  # 视觉上像输入框
        seq1_label.setFrameShadow(seq1_label.Shadow.Sunken)

        suffix_edit = QLineEdit(saved_suffix)
        suffix_edit.setPlaceholderText("（可为空）")
        suffix_edit.setFixedHeight(_ROW_H)
        suffix_edit.setFixedWidth(_COL_W)

        main_grid.addWidget(prefix_edit, 1, 1, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(seq1_label,  1, 2, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(suffix_edit, 1, 3, Qt.AlignmentFlag.AlignCenter)

        # ── 行2：起始 / 步长 / 位数（填零） 标题 ────────────────────────────
        lbl_start  = QLabel("起始数字")
        lbl_step   = QLabel("步长")
        lbl_digits = QLabel("位数（填零）")
        for lbl in (lbl_start, lbl_step, lbl_digits):
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(lbl_start,  2, 1, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(lbl_step,   2, 2, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(lbl_digits, 2, 3, Qt.AlignmentFlag.AlignCenter)

        # ── 行3：起始输入 / 步长输入 / 位数输入 ──────────────────────────────
        # 起始数字（数字模式）
        start_spin = QSpinBox()
        start_spin.setRange(0, 999999)
        start_spin.setValue(1)
        start_spin.setFixedHeight(_ROW_H)
        start_spin.setFixedWidth(_COL_W)

        # 起始字母（字母模式），初始隐藏
        start_alpha_edit = QLineEdit("A")
        start_alpha_edit.setPlaceholderText("A")
        start_alpha_edit.setFixedHeight(_ROW_H)
        start_alpha_edit.setFixedWidth(_COL_W)
        start_alpha_edit.setAlignment(Qt.AlignmentFlag.AlignCenter)
        start_alpha_edit.setToolTip("起始字母（A~ZZ），仅允许大写英文字母")
        start_alpha_edit.setVisible(False)

        # 用容器把 start_spin / start_alpha_edit 叠放，切换时 setVisible
        start_cell = QWidget()
        start_cell.setFixedWidth(_COL_W)
        start_cell.setFixedHeight(_ROW_H)
        start_cell_layout = QHBoxLayout(start_cell)
        start_cell_layout.setContentsMargins(0, 0, 0, 0)
        start_cell_layout.setSpacing(0)
        start_cell_layout.addWidget(start_spin)
        start_cell_layout.addWidget(start_alpha_edit)

        step_spin = QSpinBox()
        step_spin.setRange(0, 9)
        step_spin.setValue(1)
        step_spin.setFixedHeight(_ROW_H)
        step_spin.setFixedWidth(_COL_W)

        digits_spin = QSpinBox()
        digits_spin.setRange(0, 6)
        digits_spin.setValue(0)
        digits_spin.setToolTip("数字最小位数（不足时补前导零）；0 = 不补零")
        digits_spin.setFixedHeight(_ROW_H)
        digits_spin.setFixedWidth(_COL_W)

        main_grid.addWidget(start_cell,  3, 1, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(step_spin,   3, 2, Qt.AlignmentFlag.AlignCenter)
        main_grid.addWidget(digits_spin, 3, 3, Qt.AlignmentFlag.AlignCenter)

        # ── 行4：模式 toggle（跨3列） ─────────────────────────────────────────
        mode_group = QButtonGroup(dlg)
        rb_numeric = QRadioButton("数字序列")
        rb_alpha   = QRadioButton("字母序列（A, B…, Z, AA…, ZZ）")
        rb_numeric.setChecked(True)
        mode_group.addButton(rb_numeric, 0)
        mode_group.addButton(rb_alpha,   1)
        toggle_widget = QWidget()
        toggle_layout = QHBoxLayout(toggle_widget)
        toggle_layout.setContentsMargins(0, 4, 0, 0)
        toggle_layout.setSpacing(16)
        toggle_layout.addWidget(rb_numeric)
        toggle_layout.addWidget(rb_alpha)
        toggle_layout.addStretch()
        main_grid.addWidget(toggle_widget, 4, 1, 1, 3)  # 跨 col 1-3

        root.addLayout(main_grid)

        # ── 预览行（3行，左对齐） ─────────────────────────────────────────────
        n_preview = min(3, len(ordered_row_indices))
        preview_labels: list[QLabel] = []
        for _ in range(n_preview):
            lbl = QLabel()
            lbl.setObjectName("hintLabel")
            lbl.setAlignment(Qt.AlignmentFlag.AlignHCenter | Qt.AlignmentFlag.AlignVCenter)
            root.addWidget(lbl)
            preview_labels.append(lbl)

        # ── 更新逻辑 ──────────────────────────────────────────────────────────
        def _seq_value_numeric(i: int) -> str:
            num    = start_spin.value() + i * step_spin.value()
            digits = digits_spin.value()
            return str(num).zfill(digits) if digits > 0 else str(num)

        def _seq_value_alpha(i: int) -> str:
            raw  = start_alpha_edit.text().strip().upper()
            base = _alpha_to_num(raw) if raw else 1
            return _num_to_alpha(base + i * step_spin.value())

        def _make_seq_value(i: int) -> str:
            return _seq_value_numeric(i) if rb_numeric.isChecked() else _seq_value_alpha(i)

        def _update_all() -> None:
            # 序列值预览框（不含前后缀）
            seq1_label.setText(_make_seq_value(0))
            # 预览行（前缀 + 序列 + 后缀），每行一条
            pfx = prefix_edit.text()
            sfx = suffix_edit.text()
            for i, lbl in enumerate(preview_labels):
                is_last = (i == n_preview - 1) and (len(ordered_row_indices) > n_preview)
                val  = pfx + _make_seq_value(i) + sfx
                text = f"预览 {i + 1}：{val}"
                lbl.setText(text)

        def _on_mode_changed() -> None:
            is_numeric = rb_numeric.isChecked()
            start_spin.setVisible(is_numeric)
            start_alpha_edit.setVisible(not is_numeric)
            lbl_start.setText("起始数字" if is_numeric else "起始字母")
            lbl_digits.setEnabled(is_numeric)
            digits_spin.setEnabled(is_numeric)
            _update_all()

        def _on_alpha_text_changed(text: str) -> None:
            cleaned = "".join(c for c in text.upper() if c.isalpha() and c.isascii())
            if len(cleaned) > 2:
                cleaned = cleaned[:2]
            if cleaned != text:
                pos = start_alpha_edit.cursorPosition()
                start_alpha_edit.blockSignals(True)
                start_alpha_edit.setText(cleaned)
                start_alpha_edit.setCursorPosition(min(pos, len(cleaned)))
                start_alpha_edit.blockSignals(False)
            _update_all()

        rb_numeric.toggled.connect(lambda _: _on_mode_changed())
        rb_alpha.toggled.connect(lambda _: _on_mode_changed())
        prefix_edit.textChanged.connect(lambda _: _update_all())
        suffix_edit.textChanged.connect(lambda _: _update_all())
        start_spin.valueChanged.connect(lambda _: _update_all())
        start_alpha_edit.textChanged.connect(_on_alpha_text_changed)
        step_spin.valueChanged.connect(lambda _: _update_all())
        digits_spin.valueChanged.connect(lambda _: _update_all())
        _update_all()

        # ── 行6：按钮 ─────────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        ok_btn     = QPushButton("确定")
        cancel_btn = QPushButton("取消")
        ok_btn.setDefault(True)
        ok_btn.clicked.connect(dlg.accept)
        cancel_btn.clicked.connect(dlg.reject)
        btn_row.addWidget(ok_btn)
        btn_row.addWidget(cancel_btn)
        root.addLayout(btn_row)

        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        # 持久化前缀/后缀
        self._edit_settings.setValue(_SETT_PFX + "prefix", prefix_edit.text())
        self._edit_settings.setValue(_SETT_PFX + "suffix", suffix_edit.text())

        pfx = prefix_edit.text()
        sfx = suffix_edit.text()

        # 生成所有值
        if rb_numeric.isChecked():
            start  = start_spin.value()
            step   = step_spin.value()
            digits = digits_spin.value()

            def _make_value(i: int) -> str:
                num = start + i * step
                num_str = str(num).zfill(digits) if digits > 0 else str(num)
                return pfx + num_str + sfx
        else:
            raw_alpha  = start_alpha_edit.text().strip().upper()
            alpha_base = _alpha_to_num(raw_alpha) if raw_alpha else 1
            alpha_step = step_spin.value()

            def _make_value(i: int) -> str:  # type: ignore[misc]
                return pfx + _num_to_alpha(alpha_base + i * alpha_step) + sfx

        generated = [_make_value(i) for i in range(len(ordered_row_indices))]

        assignments = list(zip(ordered_row_indices, [col_name] * len(ordered_row_indices), generated))
        fail_count = self._apply_cell_values(assignments)
        if fail_count:
            QMessageBox.warning(
                self, "写入失败",
                f"有 {fail_count} 处写入未被 CATIA 接受。\n详情见底部状态栏及日志。",
            )

    def _refresh_keys_appearance(self, inst_keys: set) -> None:
        """刷新指定 inst_key 对应所有行的视觉外观。

        无脏标记——所有单元格始终显示默认外观（已即时写回 CATIA）。
        锁定行（文件未找到/轻量化）不受影响。
        """
        # 纯视觉更新（setFont/setForeground/setData role）会触发 itemChanged 信号，
        # 进而回调 _on_item_changed 并错误地将字段重新标记为已修改。
        # 用 _is_updating 标志屏蔽这些信号，避免循环触发。
        self._is_updating = True
        try:
            for inst_key in inst_keys:
                items = self._inst_to_items.get(inst_key, [])
                for item in items:
                    if item.data(0, _ITEM_LOCKED_ROLE):
                        continue  # 锁定行保持固定的灰色/红色样式
                    for col_idx, col_name in enumerate(self._columns):
                        if col_name in BOM_READONLY_COLUMNS or col_name == BOM_ROW_NUMBER_COLUMN:
                            continue
                        widget = self._table.itemWidget(item, col_idx)
                        if isinstance(widget, QComboBox):
                            widget.setFont(QApplication.font(widget))
                            # 清除之前可能被设置的显式 palette，让控件重新继承应用 palette
                            widget.setPalette(QPalette())
                        else:
                            # 恢复默认外观（清除 ForegroundRole 和 FontRole 的自定义数据）
                            item.setData(col_idx, Qt.ItemDataRole.ForegroundRole, None)
                            item.setData(col_idx, Qt.ItemDataRole.FontRole, None)
        finally:
            self._is_updating = False
        self._update_status()

    def _update_undo_redo_buttons(self) -> None:
        """根据撤销/重做栈的状态启用或禁用对应按钮。"""
        self._undo_btn.setEnabled(bool(self._undo_stack))
        self._redo_btn.setEnabled(bool(self._redo_stack))

    def _update_status(self) -> None:
        """刷新底部状态标签，显示上次写入结果。"""
        if not self._bom_loaded:
            self._status_label.setText("")
            return
        self._status_label.setText(self._last_write_status)

    # ── 搜索/过滤 ─────────────────────────────────────────────────────────────

    def _focus_filter(self) -> None:
        """将输入焦点移至搜索框（Ctrl+F）。"""
        self._filter_edit.setFocus()
        self._filter_edit.selectAll()

    def _on_filter_changed(self, text: str) -> None:
        """搜索框文本变化时，重新过滤表格中的行。"""
        self._filter_text = text.strip().lower()
        self._apply_filter()

    def _apply_filter(self) -> None:
        """根据 _filter_text 显示/隐藏树形控件中的各行。"""
        pattern = self._filter_text
        root = self._table.invisibleRootItem()
        for i in range(root.childCount()):
            self._apply_filter_to_item(root.child(i), pattern)

    def _apply_filter_to_item(self, item: QTreeWidgetItem, pattern: str) -> bool:
        """递归地对 *item* 及其子项应用过滤。

        返回 True 表示该项或其任意后代匹配 *pattern*，
        同时设置 item.setHidden() 的可见性。
        """
        if not pattern:
            item.setHidden(False)
            for i in range(item.childCount()):
                self._apply_filter_to_item(item.child(i), pattern)
            return True
        self_match = self._item_matches_filter(item, pattern)
        child_match = False
        for i in range(item.childCount()):
            child_match |= self._apply_filter_to_item(item.child(i), pattern)
        visible = self_match or child_match
        item.setHidden(not visible)
        return visible

    def _item_matches_filter(self, item: QTreeWidgetItem, pattern: str) -> bool:
        """判断 *item* 任意列的文本是否包含 *pattern*（大小写不敏感）。"""
        for col_idx in range(len(self._columns)):
            if pattern in item.text(col_idx).lower():
                return True
        return False

    # ── 关闭 ───────────────────────────────────────────────────────────────────

    def _on_cancel(self) -> None:
        """所有修改均已即时写回，直接关闭。"""
        self.reject()

    def done(self, result: int) -> None:
        """保存窗口几何后执行标准关闭流程。"""
        self._edit_settings.setValue("geometry", self.saveGeometry())
        super().done(result)

    def closeEvent(self, event: QCloseEvent) -> None:
        """所有修改均已即时写回，直接关闭（无需确认）。"""
        super().closeEvent(event)

    # ── 写回CATIA ─────────────────────────────────────────────────────────────

    def _auto_rename_instance_names(self, row_idx: int) -> None:
        """将选中产品/部件的子树内所有实例名批量改为 PartNumber.X 格式。

        规则：
        - 对指定节点的**直接子节点**，按 PartNumber 分组，在各组内从 1 递增
          赋值实例名：``PartNumber.1``、``PartNumber.2`` …
        - 若子节点为产品（Product）或部件（Component），则递归执行同样操作。
        - 叶节点（Part）不再递归。
        - 通过缓存的 ``_product.Name = value`` 即时写回 CATIA。
        - 同 PartNumber 的其他父节点实例（CATIA 端已自动同步）仅更新内存，不重写 COM。
        - 操作完成后重派生 ``_hierarchical_rows``，刷新表格；恢复滚动位置。
        - 写入成功的部分推入撤销栈，支持 Ctrl+Z 整体撤销。
        """
        target_row       = self._rows[row_idx]
        target_inst_key  = target_row.get("_inst_key")
        target_pm_key   = str(target_row.get("_pm_key", "")).strip()
        target_inst_info = self._inst_key_to_info.get(target_inst_key) if target_inst_key is not None else None

        # 根节点行（_inst_key 为 None）：用 part_masters[root_pm_key] 本身代表，
        # 其 instances 就是要批量改名的直接子节点
        if target_inst_info is None:
            pm = self._part_masters.get(target_pm_key)
            if pm is not None:
                target_inst_info = pm   # part_master 本身，有 instances 字段
            else:
                QMessageBox.warning(self, "无 COM 引用", "选中行没有有效的 COM 引用，无法执行操作。")
                return

        # ── 辅助：对 pm_key 对应 part_master 的 instances 生成改名计划 ──────────
        def _collect(pm_key: str, out: list) -> None:
            """按 PartNumber.n 规则为 part_masters[pm_key].instances 生成改名计划。
            instances 是文件视角，唯一一份，修改后所有引用自动同步。
            """
            pm = self._part_masters.get(pm_key, {})
            pn_counter: dict[str, int] = {}
            for inst_info in pm.get("instances", []):
                child_pn   = inst_info["pn"]
                child_bk   = inst_info["pm_key"]
                child_type = self._part_masters.get(child_bk, {}).get("type", "")
                pn_counter[child_pn] = pn_counter.get(child_pn, 0) + 1
                out.append((inst_info, f"{child_pn}.{pn_counter[child_pn]}"))
                if child_type in BomNodeType.ASSEMBLY_TYPES:
                    _collect(child_bk, out)

        # ── 收集目标子树改名计划 ──────────────────────────────────────────────────
        plan: list[tuple[dict, str]] = []
        _collect(target_pm_key, plan)

        if not plan:
            QMessageBox.information(self, "无子节点", "选中节点下没有子节点。")
            return

        # ── 确认对话框 ────────────────────────────────────────────────────────────
        already_ok  = sum(1 for ii, n in plan if ii.get("instance_name") == n)
        need_change = len(plan) - already_ok
        if need_change == 0:
            QMessageBox.information(self, "无需修改", "所有实例名已符合 PartNumber.n 规则。")
            return

        reply = QMessageBox.question(
            self, "自动修改实例名（子树范围）",
            f"共 {len(plan)} 个实例，其中 {need_change} 个需要修改。\n\n是否继续？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return

        # ── 写入 CATIA 并更新 part_masters 树 ────────────────────────────────────
        # instances 是文件视角唯一一份，修改后所有引用自动同步，不需要兄弟同步逻辑
        errors: list[str] = []
        undo_actions: list[tuple] = []
        changed = 0
        for inst_inf, new_name in plan:
            old_name = inst_inf.get("instance_name", "")
            if old_name == new_name:
                continue
            product = inst_inf.get("product")
            if product is None:
                continue
            ik = inst_inf.get("inst_key")
            try:
                product.Name = new_name
                inst_inf["instance_name"] = new_name
                if ik is not None:
                    undo_actions.append((ik, BOM_INSTANCE_NAME_COLUMN, old_name, new_name))
                changed += 1
            except Exception as e:
                errors.append(f"{inst_inf.get('inst_key', '?')}: {e}")
                logger.error("auto_rename_instance_names error inst_key=%s: %s", ik, e)

        # ── 推入撤销栈（整批作为一个原子步骤）────────────────────────────────────
        if undo_actions:
            self._push_undo(undo_actions)

        # ── 重建行列表，刷新表格（保存/恢复滚动位置）─────────────────────────────
        vscroll = self._table.verticalScrollBar().value()
        self._rebuild_rows()
        self._populate_table()
        self._table.verticalScrollBar().setValue(vscroll)
        # 尝试滚动回操作节点所在位置
        if row_idx < len(self._item_by_row):
            self._table.scrollToItem(
                self._item_by_row[row_idx],
                QAbstractItemView.ScrollHint.EnsureVisible,
            )

        if errors:
            QMessageBox.warning(
                self, "部分写入失败",
                f"已修改 {changed} 个实例名，以下条目写入失败：\n" + "\n".join(errors[:10]),
            )
        else:
            self._last_write_status = f"已自动修改 {changed} 个实例名"
            self._update_status()

    def _auto_rename_files(self, row_idx: int) -> None:
        """将选中节点及其子树内所有文件名与零件编号不符的文件批量另存为改名。

        规则：
        - 作用范围：选中节点（包含自身）的整个子树中 ``Level >=`` 当前层级的所有行。
          若选中节点是零件（无子行），或当前处于汇总 BOM 状态，则仅作用于自身。
        - 跳过部件（Component）—— 部件共享父产品文件，没有独立文件可改名。
        - 跳过 ``_not_found`` / ``_no_file`` 行。
        - 同一文件路径只处理一次（去重）。
        - 改名完成后同步更新 ``_rows``、``_hierarchical_rows``、``_full_rows`` 并刷新表格。
        """
        target_row   = self._rows[row_idx]
        target_level = int(target_row.get("Level", 0))

        # ── 收集子树行（含 target_row 本身） ─────────────────────────────────
        # 两种情况只改自身：
        #   1. 汇总 BOM 模式：各行不展开子树，遍历子行无意义
        #   2. 选中节点是零件（下一行不存在或 Level 未增大）：本就无子行
        is_self_only = self._summarize or (
            row_idx + 1 >= len(self._rows)
            or int(self._rows[row_idx + 1].get("Level", 0)) <= target_level
        )

        subtree: list[dict] = []
        if is_self_only:
            subtree = [target_row]
        else:
            for i in range(row_idx, len(self._rows)):
                r = self._rows[i]
                if i > row_idx and int(r.get("Level", 0)) < target_level:
                    break
                subtree.append(r)

        # ── 过滤出需要改名的 (filepath, new_stem) ────────────────────────────
        to_rename: list[tuple[str, str]] = []
        seen_fps: set[str] = set()
        for row in subtree:
            if row.get("Type") == BomNodeType.COMPONENT:
                continue           # 部件无独立文件
            if row.get("_not_found") or row.get("_no_file"):
                continue
            fp = str(row.get("_filepath", ""))
            if not fp or fp in seen_fps:
                continue
            seen_fps.add(fp)
            orig_pn = str(row.get("Part Number", ""))
            pn      = get_part_master_attr(self._part_masters, orig_pn, "Part Number", orig_pn)
            if pn and Path(fp).stem != pn:
                to_rename.append((fp, pn))

        if not to_rename:
            QMessageBox.information(self, "无需改名", "选中节点及其子树内所有文件名已与零件编号一致。")
            return

        delete_old = (
            QMessageBox.question(
                self, "是否删除旧文件",
                f"将对子树内 {len(to_rename)} 个文件执行另存为改名。\n"
                "另存为完成后，是否删除旧文件？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes
        )

        QMessageBox.information(self, "请在 CATIA 中继续操作", "准备就绪，请在 CATIA 中确认后续操作。")

        renamed_count = 0
        for fp, pn in reversed(to_rename):
            if not PART_NUMBER_VALID_PATTERN.fullmatch(pn):
                QMessageBox.warning(
                    self, "零件编号含非法字符",
                    f"零件编号 「{pn}」 含有非法字符。\n"
                    "不允许：控制字符、非 ASCII 字符，以及 Windows 文件名禁用字符"
                    "（\\ / : * ? \" < > |）。\n请在表格中修改此零件编号后重试。",
                )
                continue
            if not Path(fp).exists():
                continue
            try:
                new_fp, skipped = rename_document(fp, pn, delete_old=delete_old)
                if skipped:
                    logger.info("SaveAs skipped for %s (user cancelled)", Path(fp).name)
                    continue
            except FileNotFoundError as e:
                QMessageBox.warning(self, "无法找到文档", str(e))
                continue
            except Exception as e:
                QMessageBox.warning(self, "另存为失败", f"文件「{Path(fp).name}」另存为失败：\n{e}")
                continue

            # 同步更新 part_masters 中的 filepath / filename（唯一真相）
            # 以及当前显示行的对应字段
            for pm in self._part_masters.values():
                if pm.get("filepath") == fp:
                    pm["filepath"] = new_fp
                    pm["filename"] = pn
            for row in self._rows:
                if str(row.get("_filepath", "")) == fp:
                    row["_filepath"] = new_fp
                    row["Filename"]  = pn
            renamed_count += 1

        if renamed_count > 0:
            QMessageBox.information(
                self, "改名完成",
                f"已成功将 {renamed_count} 个文件通过 CATIA 另存为功能改名。",
            )
            self._populate_table()

    def _rename_selected_file(self) -> None:
        """通过 CATIA 另存为功能，对选中的单行 BOM 记录执行重命名或移动操作。"""
        selected_row_indices = {
            it.data(0, Qt.ItemDataRole.UserRole)
            for it in self._table.selectedItems()
            if it.data(0, Qt.ItemDataRole.UserRole) is not None
        }
        if len(selected_row_indices) != 1:
            QMessageBox.warning(
                self, "请选择单行",
                "请在表格中选中恰好一行，再执行此操作。",
            )
            return

        row_idx  = next(iter(selected_row_indices))
        row_data = self._rows[row_idx]
        fp       = str(row_data.get("_filepath", ""))

        if not fp or row_data.get("_not_found"):
            QMessageBox.warning(self, "无有效路径", "该行没有可用的文件路径，无法执行重命名/移动。")
            return
        # 注意：此处不检查 Path(fp).exists()；
        # 未保存过的零件（文件尚不在磁盘上但在CATIA内存中打开）同样允许另存为。

        # 属性已即时写回，无需前置写回检查
        dlg = _FileRenameDialog(fp, parent=self)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return

        new_fp                = dlg.new_path

        # 仅当旧文件实际存在于磁盘时才询问是否删除，否则无从删除
        file_on_disk = Path(fp).exists()
        delete_old = file_on_disk and (
            QMessageBox.question(
                self, "是否删除旧文件",
                f"另存为完成后，是否删除旧文件？\n\n旧文件：{fp}",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            ) == QMessageBox.StandardButton.Yes
        )

        QMessageBox.information(self, "请在 CATIA 中继续操作", "准备就绪，请在 CATIA 中确认后续操作。")

        try:
            new_fp, skipped = rename_document(
                fp, Path(new_fp).stem,
                delete_old=delete_old,
                target_path=new_fp,
            )
            if skipped:
                logger.info(
                    "SaveAs skipped for %s (user cancelled)", Path(fp).name
                )
                return
        except FileNotFoundError as e:
            QMessageBox.warning(self, "无法找到文档", str(e))
            return
        except Exception as e:
            QMessageBox.warning(self, "另存为失败", f"文件操作失败：\n{e}")
            return

        new_stem = Path(new_fp).stem
        # 同步更新 part_masters（唯一真相）和当前显示行
        for pm in self._part_masters.values():
            if pm.get("filepath") == fp:
                pm["filepath"] = new_fp
                pm["filename"] = new_stem
                pm["_no_file"] = False
        for row in self._rows:
            if str(row.get("_filepath", "")) == fp:
                row["_filepath"] = new_fp
                row["Filename"]  = new_stem
                row["_no_file"]  = False
        self._populate_table()
        QMessageBox.information(
            self, "操作成功",
            f"文件已成功另存为：\n{new_fp}",
        )
    
    # ── 导出表格 ──────────────────────────────────────────────────────────────

    def _export_table(self) -> None:
        """将当前显示的 BOM 表格导出为 Excel 或 CSV 文件。"""
        if not self._bom_loaded or not self._rows:
            QMessageBox.warning(self, "无数据", "请先加载 BOM 。")
            return

        # 根据根产品零件编号建议默认文件名
        suffix_hint = "_汇总BOM" if self._summarize else "_BOM"
        root_pm_display = self._part_masters.get(self._root_pm_key, {})
        root_pn_display = root_pm_display.get("part_number", self._root_pm_key)
        # 去除 Windows 文件名中不合法的字符（本工具目标平台为 Windows）
        invalid_chars = r'\/:*?"<>|'
        safe_stem = "".join(c if c not in invalid_chars else "_" for c in root_pn_display)
        base_name = (safe_stem + suffix_hint) if safe_stem else ""

        initial_name = ""
        if base_name:
            # 尝试沿用源文件所在目录（仅在使用文件模式时）
            if not self._use_active_chk.isChecked():
                fp_src = self._file_edit.text().strip()
                if fp_src:
                    initial_name = str(Path(fp_src).with_name(base_name))
                else:
                    initial_name = base_name
            else:
                initial_name = base_name

        dest, selected_filter = QFileDialog.getSaveFileName(
            self,
            "导出 BOM 表格",
            initial_name,
            "Excel 工作簿 (*.xlsx);;CSV 文件 (*.csv)",
        )
        if not dest:
            return

        dest_path = Path(dest)
        # 根据扩展名推断格式；不明确时回退为 xlsx
        suffix = dest_path.suffix.lower()
        if suffix not in (".xlsx", ".csv"):
            dest_path = dest_path.with_suffix(".xlsx")
            suffix = ".xlsx"

        # 导出列：当前可见列，排除行号"#"列
        export_cols = [c for c in self._columns if c != BOM_ROW_NUMBER_COLUMN]

        # 从当前表格快照列宽（像素）
        col_px_widths: dict[str, int] = {}
        for col_idx, col_name in enumerate(self._columns):
            if col_name != BOM_ROW_NUMBER_COLUMN:
                col_px_widths[col_name] = self._table.columnWidth(col_idx)

        # 使用与 _populate_table 相同的取值逻辑收集行数据
        rows_data: list[dict] = []
        for row_data in self._rows:
            pm_key_exp = str(row_data.get("_pm_key", "")).strip()
            row_out: dict = {}
            for col_name in export_cols:
                if col_name == "Source":
                    raw = get_part_master_attr(self._part_masters, pm_key_exp, "Source", "")
                    val = SOURCE_TO_DISPLAY.get(raw, raw)
                elif col_name in PRESET_USER_REF_PROPERTY_OPTIONS:
                    val = get_part_master_attr(
                        self._part_masters, pm_key_exp, col_name,
                        str(row_data.get(col_name, ""))
                    )
                elif col_name == "Filename":
                    fp_val = str(row_data.get("_filepath", ""))
                    fn_val = str(row_data.get("Filename", ""))
                    if self._show_filepath_col:
                        val = fp_val if fp_val else fn_val
                    else:
                        val = Path(fp_val).name if fp_val else fn_val
                elif col_name in BOM_READONLY_COLUMNS:
                    val = str(row_data.get(col_name, ""))
                else:
                    val = get_part_master_attr(
                        self._part_masters, pm_key_exp, col_name,
                        str(row_data.get(col_name, ""))
                    )
                row_out[col_name] = val
            rows_data.append(row_out)

        try:
            if suffix == ".xlsx":
                self._write_xlsx(dest_path, export_cols, col_px_widths, rows_data)
            else:
                self._write_csv(dest_path, export_cols, rows_data)
        except PermissionError:
            QMessageBox.critical(
                self, "导出失败",
                f"无法写入文件（文件可能已在其他程序中打开）：\n{dest_path}",
            )
            return
        except Exception as e:
            logger.error(f"BOM table export failed: {e}")
            QMessageBox.critical(self, "导出失败", f"导出时出错：\n{e}")
            return

        self._show_export_success(dest_path)

    def _show_export_success(self, dest_path: Path) -> None:
        """导出成功后弹出含"打开文件"和"打开所在文件夹"按钮的提示框。"""
        msg = QMessageBox(self)
        msg.setWindowTitle("导出成功")
        msg.setText(f"BOM 已成功导出：\n{dest_path}")
        msg.setIcon(QMessageBox.Icon.Information)
        open_file_btn   = msg.addButton("打开文件", QMessageBox.ButtonRole.ActionRole)
        open_folder_btn = msg.addButton("打开所在文件夹", QMessageBox.ButtonRole.ActionRole)
        msg.addButton(QMessageBox.StandardButton.Ok)
        msg.exec()
        clicked = msg.clickedButton()
        if clicked == open_file_btn:
            try:
                QDesktopServices.openUrl(QUrl.fromLocalFile(str(dest_path)))
            except Exception as exc:
                logger.warning(f"Failed to open exported file: {exc}")
        elif clicked == open_folder_btn:
            self._open_path(str(dest_path))

    def _export_header(self, col_name: str) -> str:
        """返回列的显示表头字符串，与当前表格保持一致。"""
        if col_name == "Filename" and self._show_filepath_col:
            return "完整路径"
        return BOM_COLUMN_DISPLAY_NAMES.get(col_name, col_name)

    def _write_xlsx(
        self,
        dest: Path,
        cols: list[str],
        px_widths: dict[str, int],
        rows: list[dict],
    ) -> None:
        """将 *rows* 写入 *dest* 路径的 .xlsx 工作簿。"""

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None  # 新建工作簿 active 必然非 None
        ws.title = "汇总 BOM" if self._summarize else "BOM"

        center      = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
        thin_side   = Side(style="thin")
        thin_border = Border(
            left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
        )

        # 表头行
        for col_idx, col_name in enumerate(cols, start=1):
            cell        = ws.cell(row=1, column=col_idx, value=self._export_header(col_name))
            cell.font   = Font(bold=True)
            cell.fill   = header_fill
            cell.border = thin_border

        # 数据行
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(cols, start=1):
                raw_val = row.get(col_name, "")
                # 将数字存为整数，以便 Excel 排序/筛选
                if col_name in ("Level", "Quantity"):
                    try:
                        value = int(raw_val)
                    except (ValueError, TypeError):
                        logger.debug(
                            "Could not convert %r to int for column '%s'", raw_val, col_name
                        )
                        value = raw_val
                else:
                    # Type 列存储英文 key，导出时转为中文显示
                    value = TYPE_DISPLAY_NAMES.get(raw_val, raw_val) if col_name == "Type" else raw_val
                cell        = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                if col_name in ("Level", "Quantity", "Type"):
                    cell.alignment = center

        # 冻结表头行并启用自动筛选
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # 列宽：像素→Excel字符单位换算
        # Calibri 11pt 默认字体约为每字符7像素，作为近似换算基准
        PX_PER_CHAR = 7.0
        for col_idx, col_name in enumerate(cols, start=1):
            col_letter = ws.cell(row=1, column=col_idx).column_letter  # type: ignore[union-attr]
            px = px_widths.get(col_name, 80)
            char_width = max(8.0, px / PX_PER_CHAR)
            ws.column_dimensions[col_letter].width = round(char_width, 1)

        wb.save(str(dest))
        logger.info(f"BOM table exported (xlsx) -> {dest}")

    def _write_csv(
        self,
        dest: Path,
        cols: list[str],
        rows: list[dict],
    ) -> None:
        """将 *rows* 写入 *dest* 路径的 UTF-8 CSV 文件（带 BOM 头）。"""

        with open(dest, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([self._export_header(c) for c in cols])
            for row in rows:
                writer.writerow([row.get(c, "") for c in cols])
        logger.info(f"BOM table exported (csv) -> {dest}")
    
    def _finish_and_close(self) -> None:
        """所有修改均已即时写回，直接关闭。"""
        self.accept()

    # ── 右键上下文菜单 ────────────────────────────────────────────────────────

    def _on_tree_context_menu(self, pos) -> None:
        """显示右键点击的 BOM 行对应的上下文菜单。

        若关联文件内嵌了缩略图，则在菜单顶部以非交互式图片控件展示。
        """
        item = self._table.itemAt(pos)
        if item is None:
            return
        row_idx = item.data(0, Qt.ItemDataRole.UserRole)
        if row_idx is None:
            return

        row_data     = self._rows[row_idx]
        fp           = str(row_data.get("_filepath", ""))
        fp_path      = Path(fp) if fp else None
        is_component = row_data.get("Type") == BomNodeType.COMPONENT
        is_assembly  = row_data.get("Type") in BomNodeType.ASSEMBLY_TYPES
        not_found    = bool(row_data.get("_not_found"))
        no_file      = bool(row_data.get("_no_file"))
        unreadable   = bool(row_data.get("_unreadable"))
        pn           = str(row_data.get("Part Number", ""))

        # 确保右键点击的行已被选中，以便下游方法（_rename_selected_file 等）能找到它
        if not item.isSelected():
            self._table.clearSelection()
            item.setSelected(True)

        menu = QMenu(self)

        # ── 嵌入缩略图（可用时在菜单顶部内联显示）─────────────────────────────
        # 以下情况跳过缩略图提取：
        #   • 部件：文件路径属于父产品，而非此组件
        #   • not_found：CATIA无法解析该文件
        #   • 文件在磁盘上不存在（未保存或丢失）
        if fp and not is_component and not not_found and fp_path is not None and fp_path.exists():
            img_bytes = read_catia_thumbnail(fp)
            if img_bytes:
                pixmap = QPixmap()
                loaded = pixmap.loadFromData(img_bytes)
                if loaded and not pixmap.isNull():
                    if (pixmap.width() > BOM_THUMBNAIL_MAX_SIZE
                            or pixmap.height() > BOM_THUMBNAIL_MAX_SIZE):
                        pixmap = pixmap.scaled(
                            BOM_THUMBNAIL_MAX_SIZE,
                            BOM_THUMBNAIL_MAX_SIZE,
                            Qt.AspectRatioMode.KeepAspectRatio,
                            Qt.TransformationMode.SmoothTransformation,
                        )
                    thumb_label = QLabel()
                    thumb_label.setPixmap(pixmap)
                    thumb_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
                    thumb_label.setContentsMargins(6, 4, 6, 4)
                    thumb_action = QWidgetAction(menu)
                    thumb_action.setDefaultWidget(thumb_label)
                    menu.addAction(thumb_action)
                    menu.addSeparator()

        # ── 打开路径 ──────────────────────────────────────────────────────────
        act_open_path = menu.addAction("打开路径")
        path_available = (
            bool(fp) and not no_file and fp_path is not None
            and (fp_path.exists() or fp_path.parent.exists())
        )
        act_open_path.setEnabled(path_available)

        # ── 复制路径 ──────────────────────────────────────────────────────────
        act_copy_path = menu.addAction("复制路径")
        act_copy_path.setEnabled(bool(fp) and not no_file)

        # ── 复制单元格内容 ────────────────────────────────────────────────────
        # 根据鼠标所在列动态获取单元格文本
        clicked_col_idx = self._table.columnAt(pos.x())
        cell_text: str = ""
        if 0 <= clicked_col_idx < len(self._columns):
            widget = self._table.itemWidget(item, clicked_col_idx)
            if isinstance(widget, QComboBox):
                cell_text = widget.currentText()
            else:
                cell_text = item.text(clicked_col_idx)
        act_copy_cell = menu.addAction("复制单元格内容")
        act_copy_cell.setEnabled(bool(cell_text))

        # ── 在CATIA中打开 ─────────────────────────────────────────────────────
        # 仅当文件在磁盘上存在且不是损坏/轻量化引用时启用。
        # 部件行共享父产品的文件路径，因此也排除在外。
        act_open_catia = menu.addAction("在 CATIA 中打开")
        catia_available = (
            not is_component and not not_found and not unreadable
            and fp_path is not None and fp_path.exists()
        )
        act_open_catia.setEnabled(catia_available)

        menu.addSeparator()

        # ── 填充操作（多选 + 可编辑列时启用） ────────────────────────────────
        # 收集所有选中行的 row_idx，按 row_idx 升序排列（视觉顺序）
        selected_row_indices: list[int] = sorted(
            {
                it.data(0, Qt.ItemDataRole.UserRole)
                for it in self._table.selectedItems()
                if it.data(0, Qt.ItemDataRole.UserRole) is not None
            }
        )
        # 右键点击列是否可编辑（非只读列）
        fill_col_name: str = (
            self._columns[clicked_col_idx]
            if 0 <= clicked_col_idx < len(self._columns)
            else ""
        )
        fill_enabled = (
            len(selected_row_indices) >= 2
            and bool(fill_col_name)
            and fill_col_name not in BOM_READONLY_COLUMNS
            and fill_col_name != BOM_INSTANCE_NAME_COLUMN
            # 序列填充仅对文本列启用（Source / Preset 选项列是 combo，不支持序列）
        )
        fill_seq_enabled = fill_enabled and fill_col_name not in PRESET_USER_REF_PROPERTY_OPTIONS and fill_col_name != "Source"

        _fill_col_display = BOM_COLUMN_DISPLAY_NAMES.get(fill_col_name, fill_col_name) if fill_col_name else ""

        act_fill_same = menu.addAction(
            f"首行内容填充（{_fill_col_display}）" if fill_enabled else "首行内容填充"
        )
        act_fill_same.setEnabled(fill_enabled)

        act_fill_seq = menu.addAction(
            f"序列填充（{_fill_col_display}）" if fill_seq_enabled else "序列填充"
        )
        act_fill_seq.setEnabled(fill_seq_enabled)

        menu.addSeparator()

        # ── 另存为 ────────────────────────────────────────────────────────────
        act_edit_path = menu.addAction("另存为")
        # 允许对未保存过的零件（文件不在磁盘上但在CATIA内存中）执行另存为；
        # 仅排除没有路径或CATIA无法找到的节点。
        act_edit_path.setEnabled(bool(fp) and not is_component and not not_found)

        menu.addSeparator()

        # ── 自动修改文件名（子树范围）────────────────────────────────────────
        act_auto_rename_files = menu.addAction("自动修改文件名（子树范围）")
        act_auto_rename_files.setToolTip(
            "将选中节点及其子树内所有文件名与零件编号不符的文件批量另存为改名\n"
            "（部件共享父产品文件，自动跳过）"
        )
        act_auto_rename_files.setEnabled(self._bom_loaded)

        # ── 自动修改实例名（子树范围）（仅完整 BOM 模式）────────────────────────────────
        act_auto_rename_instances = menu.addAction("自动修改实例名（子树范围）")
        act_auto_rename_instances.setToolTip(
            "将选中产品/部件的子节点实例名批量改为 PartNumber.X 格式，并递归处理子装配\n"
            "（仅完整 BOM 模式可用）"
        )
        act_auto_rename_instances.setEnabled(
            self._bom_loaded and is_assembly and self._full_bom
        )

        # ── 刷新属性值（从 CATIA 重新读取）───────────────────────────────────
        act_refresh = menu.addAction("刷新属性值（子树范围）")
        act_refresh.setToolTip("从 CATIA COM 重新读取选中节点及其子树内所有行的属性值，覆盖表格中的当前显示值")
        act_refresh.setEnabled(self._bom_loaded and bool(selected_row_indices))

        action = menu.exec(self._table.viewport().mapToGlobal(pos))

        if action == act_open_path:
            self._open_path(fp)
        elif action == act_copy_path:
            QApplication.clipboard().setText(str(Path(fp).parent))
        elif action == act_copy_cell:
            QApplication.clipboard().setText(cell_text)
        elif action == act_open_catia:
            self._open_in_catia(fp)
        elif action == act_fill_same:
            self._fill_same_value(fill_col_name, selected_row_indices)
        elif action == act_fill_seq:
            self._fill_sequence(fill_col_name, selected_row_indices)
        elif action == act_edit_path:
            self._rename_selected_file()
        elif action == act_auto_rename_files:
            self._auto_rename_files(row_idx)
        elif action == act_auto_rename_instances:
            self._auto_rename_instance_names(row_idx)
        elif action == act_refresh:
            # 子树递归由 _refresh_rows_from_catia 内部沿 part_masters["instances"] 完成，
            # 此处只需传入右键所在行的行号作为起始节点。
            self._refresh_rows_from_catia([row_idx])

    def _refresh_rows_from_catia(self, row_indices: list[int]) -> None:
        """从 CATIA COM 重新读取 part_master 属性并增量刷新表格。

        以 part_masters 为唯一数据源，参照 ``collect_bom_part_masters._traverse``
        的读取逻辑：

        1. 从各起始行的 ``pm_key`` 出发，沿
           ``part_masters[pm_key]["instances"]`` 递归遍历子树，收集所有唯一
           ``pm_key``（同一 pm_key 只处理一次）。
        2. 对每个 ``pm_key``，用 ``part_masters[pm_key]["_product"]`` 的 COM 引用
           重新读取属性，结果写回对应的 part_master dict。
        3. ``instance_name`` **不刷新**——它是实例级属性，存于父 part_master 的
           ``instances[*]["instance_name"]`` 中，不属于 part_master 本身。
        4. Part Number 若有变动，通过 ``rename_part_master`` 同步更新所有父节点
           的 ``inst_info["pn"]`` 及 ``_rows`` 中的对应记录；其余属性通过
           ``set_part_master_attr`` 写入。
        5. 刷新完成后通过 ``_pm_key_to_inst_keys`` / ``_inst_to_items`` 增量更新
           受影响行的显示值，**不重建整棵树**（无 ``_populate_table`` 调用）。
           只更新 PM 级属性列；结构列（Level / Type / Filename / Quantity /
           Instance Name）不随属性刷新变化，跳过。
        """
        if not row_indices:
            return

        # ── 1. 从起始行收集唯一 pm_key ────────────────────────────────────────
        seen_starts: set[str] = set()
        start_pm_keys: list[str] = []
        for row_idx in row_indices:
            if row_idx >= len(self._rows):
                continue
            pm_key = str(self._rows[row_idx].get("_pm_key", "")).strip()
            if pm_key and pm_key in self._part_masters and pm_key not in seen_starts:
                seen_starts.add(pm_key)
                start_pm_keys.append(pm_key)

        if not start_pm_keys:
            return

        # ── 2. 沿 instances 递归收集子树内全部唯一 pm_key（前序，同 _traverse）──
        ordered_pm_keys: list[str] = []
        visited: set[str] = set()

        def _collect(pk: str) -> None:
            if pk in visited or pk not in self._part_masters:
                return
            visited.add(pk)
            ordered_pm_keys.append(pk)
            for inst in self._part_masters[pk].get("instances", []):
                _collect(inst["pm_key"])

        for pk in start_pm_keys:
            _collect(pk)

        # ── 3. 逐 pm_key 从 COM 重新读取属性 ──────────────────────────────────
        all_read_cols = list(dict.fromkeys(
            BOM_EDIT_COLUMN_ORDER
            + [c for c in self._all_custom_columns if c not in BOM_EDIT_COLUMN_ORDER]
        ))

        progress = QProgressDialog(
            f"正在刷新 {len(ordered_pm_keys)} 个零件…",
            None, 0, len(ordered_pm_keys), self,
        )
        progress.setWindowTitle("刷新属性值")
        progress.setWindowModality(Qt.WindowModality.WindowModal)
        progress.setMinimumDuration(300)

        refreshed_pm_keys: set[str] = set()
        try:
            for i, pm_key in enumerate(ordered_pm_keys):
                progress.setValue(i)
                QApplication.processEvents()

                pm = self._part_masters.get(pm_key)
                if pm is None:
                    continue
                if pm.get("_not_found") or pm.get("_unreadable"):
                    continue
                product = pm.get("_product")
                if product is None:
                    continue

                # 嵌入部件（host_file_pn 非空）不走 ReferenceProduct，同 _traverse
                is_embedded = bool(pm.get("host_file_pn"))
                try:
                    new_props = refresh_row_from_com(
                        product, all_read_cols, self._all_custom_columns,
                        is_component=is_embedded,
                    )
                except Exception as e:
                    logger.warning("刷新 pm_key=%r 失败：%s", pm_key, e)
                    continue

                # 写回 part_master；不触碰 instances / _product 等结构字段
                # Part Number 改名需同步 inst_info["pn"] 及 _rows，单独走 rename_part_master
                for col, val in new_props.items():
                    if col == "Part Number":
                        if val != pm.get("part_number", ""):
                            rename_part_master(
                                self._part_masters,
                                self._pm_key_to_inst_keys,
                                pm_key,
                                val,
                            )
                            # _rows 同步，保证增量刷新路径与 _apply_cell_values 一致
                            for row in self._rows:
                                if row.get("_pm_key") == pm_key:
                                    row["Part Number"] = val
                    else:
                        set_part_master_attr(self._part_masters, pm_key, col, val)

                refreshed_pm_keys.add(pm_key)

            progress.setValue(len(ordered_pm_keys))
        finally:
            progress.close()

        if not refreshed_pm_keys:
            return

        # ── 4. 增量刷新受影响行的显示值（不重建整棵树）──────────────────────────
        # 结构列（Level / Type / Filename / Filepath / Quantity / Instance Name /
        # Row Number）不随属性刷新变化，跳过；只更新 PM 级属性列。
        _skip_cols: frozenset[str] = frozenset((
            BOM_ROW_NUMBER_COLUMN,
            BOM_INSTANCE_NAME_COLUMN,
            "description_inst",
            "Quantity", "Filename", "Filepath",
        )) | frozenset(BOM_READONLY_COLUMNS)

        self._is_updating = True
        try:
            for pm_key in refreshed_pm_keys:
                for ik in self._pm_key_to_inst_keys.get(pm_key, []):
                    for tree_item in self._inst_to_items.get(ik, []):
                        row_idx = tree_item.data(0, Qt.ItemDataRole.UserRole)
                        if row_idx is None or row_idx >= len(self._rows):
                            continue
                        row_data = self._rows[row_idx]
                        for col_idx, col_name in enumerate(self._columns):
                            if col_name in _skip_cols:
                                continue
                            if col_name == "Source":
                                raw = get_part_master_attr(
                                    self._part_masters, pm_key, "Source", "")
                                display_val = SOURCE_TO_DISPLAY.get(
                                    raw, SOURCE_OPTIONS[0])
                                widget = self._table.itemWidget(tree_item, col_idx)
                                if isinstance(widget, QComboBox):
                                    widget.blockSignals(True)
                                    widget.setCurrentText(display_val)
                                    widget.blockSignals(False)
                            elif PRESET_USER_REF_PROPERTY_OPTIONS.get(col_name) is not None:
                                val = get_part_master_attr(
                                    self._part_masters, pm_key, col_name,
                                    str(row_data.get(col_name, "")))
                                widget = self._table.itemWidget(tree_item, col_idx)
                                if isinstance(widget, QComboBox):
                                    widget.blockSignals(True)
                                    widget.setCurrentText(val)
                                    widget.blockSignals(False)
                            else:
                                val = get_part_master_attr(
                                    self._part_masters, pm_key, col_name,
                                    str(row_data.get(col_name, "")))
                                if tree_item.text(col_idx) != val:
                                    tree_item.setText(col_idx, val)
        finally:
            self._is_updating = False

        self._last_write_status = f"已刷新：{len(refreshed_pm_keys)} 个零件"
        self._update_status()

    def _open_path(self, fp: str) -> None:
        """在 Windows 资源管理器中打开包含 *fp* 的文件夹，并高亮选中该文件。

        使用 ShellExecuteW（宽字符 Unicode API）调用 explorer，避免经过
        cmd.exe / PowerShell 时中文路径因 OEM 代码页转换而乱码。
        """
        p = Path(fp).resolve()
        try:
            if p.exists():
                ctypes.windll.shell32.ShellExecuteW(
                    None, "open", "explorer.exe", f'/select,"{p}"', None, 1
                )
            elif p.parent.exists():
                ctypes.windll.shell32.ShellExecuteW(
                    None, "open", "explorer.exe", f'"{p.parent}"', None, 1
                )
        except Exception as exc:
            logger.warning(f"Failed to open path in Explorer: {exc}")

    def _open_in_catia(self, fp: str) -> None:
        """在 CATIA 中打开 *fp* 指向的文档，并将 CATIA V5 主窗口置于前台。"""
        try:
            open_document(fp, foreground=True)
        except Exception as e:
            QMessageBox.warning(self, "在 CATIA 中打开失败", f"无法在 CATIA 中打开文件：\n{e}")
